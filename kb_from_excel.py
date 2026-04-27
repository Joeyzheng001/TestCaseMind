#!/usr/bin/env python3
"""
kb_from_excel.py - 从表设计文档提取知识库

把华锐资管风控表设计.xlsx 里的关键内容转换为结构化 md 文件，
作为测试 Agent 的知识库来源（KB 标记）。

用法:
    python kb_from_excel.py <xlsx路径>
    python kb_from_excel.py 华锐资管_规则配置_开发设计文档-资管风控表设计.xlsx

输出（写入 knowledge_base/ 目录）:
    数据字典枚举值.md     — 446个字典代码的枚举值，测试边界值的黄金来源
    风控表字段说明.md     — 51张风控表的字段定义，测试字段约束的依据
    表依赖关系.md         — 表间依赖，测试数据初始化顺序的依据
    元数据字段定义.md     — 1363个元数据字段类型，测试字段类型边界
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("错误: 需要 openpyxl，请先安装: pip install openpyxl")
    sys.exit(1)

KB_DIR = Path(__file__).parent / "knowledge_base"
KB_DIR.mkdir(exist_ok=True)


# ── 1. 数据字典枚举值 ──────────────────────────────────────────────────────
def extract_data_dict(wb) -> str:
    """提取所有字典代码的枚举值，按业务分组。"""
    ws = wb['数据字典']
    dict_codes: dict = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        code  = row[0]
        name  = row[1]
        val   = row[2]
        label = row[3]
        if not (isinstance(code, int) and name and isinstance(val, int) and label):
            continue
        if code not in dict_codes:
            dict_codes[code] = {'name': name, 'values': []}
        dict_codes[code]['values'].append((val, label))

    lines = [
        "# 数据字典枚举值",
        "",
        "本文档列出系统所有枚举字段的合法值域。",
        "测试时，**每个枚举字段必须覆盖：合法值（各枚举值）、非法值（超出范围的整数）、空值**。",
        "",
        f"共 {len(dict_codes)} 个字典代码。",
        "",
    ]

    for code, info in sorted(dict_codes.items()):
        lines.append(f"## 字典代码 {code}：{info['name']}")
        for val, label in info['values']:
            lines.append(f"- {val} → {label}")
        # 自动推断测试建议
        vals = [v for v, _ in info['values']]
        if vals:
            lines.append(f"- **合法值范围**: {min(vals)} ~ {max(vals)}（共{len(vals)}项）")
            lines.append(f"- **测试边界**: 传入 {max(vals)+1} 或 0 或 -1 应报错或返回空")
        lines.append("")

    return "\n".join(lines)


# ── 2. 全库表字段说明（所有 sheet）─────────────────────────────────────────
# 不含表结构的 sheet，跳过
_SKIP_SHEETS = {
    '修订记录', '表列表', '表依赖', '元数据', '数据字典', '字典映射', '词根',
    '全局序号生成规则', '盘中实时接口', 'IBOR（含上场转换）', '账户（含上场转换）',
    '资产（含上场转换）', '交易（含上场转换）', '品种（含上场转换）',
    '主体（含上场转换）', '估值相关（含上场转换）', '公共（含上场转换）',
    '历史交易(含上场转换）', '证券分类', '网关接口', '交易', '风控参数定义表表',
}


def _parse_sheet_tables(ws, sheet_name: str, meda_map: dict) -> dict:
    """解析单个 sheet，返回 {表名: {cn, fields, sheet}} 字典。"""
    tables   = {}
    cur_name = None
    cur_cn   = None
    for row in ws.iter_rows(values_only=True):
        if not row or all(v is None for v in row[:4]):
            continue
        if row[0] == '表英文名':
            cur_name = row[1]
            cur_cn   = (row[5] if len(row) > 5 else '') or ''
            if cur_name:
                tables[cur_name] = {'cn': cur_cn, 'fields': [], 'sheet': sheet_name}
        elif cur_name and row[0] is None and row[1] and row[1] != '字段名':
            field_code = row[1]
            is_pk      = 'PK' if row[2] == 'Y' else ''
            nullable   = '可空' if row[3] != 'N' else '非空'
            meta       = meda_map.get(field_code, ('', '', ''))
            tables[cur_name]['fields'].append({
                'code': field_code, 'name': meta[0],
                'type': meta[1],   'nullable': nullable,
                'pk':   is_pk,     'memo': meta[2],
            })
    return tables


def extract_risk_tables(wb) -> str:
    """提取所有 sheet 的表字段定义，按模块分组输出。"""
    # 构建元数据映射 field_code -> (name, oracle_type, memo)
    meda_map = {}
    for row in wb['元数据'].iter_rows(min_row=3, values_only=True):
        if row[1]:
            meda_map[row[1]] = (row[2] or '', row[6] or '', row[7] or '')

    # 遍历所有含表结构的 sheet
    all_tables: dict = {}   # sheet_name -> {tname -> tinfo}
    for sname in wb.sheetnames:
        if sname in _SKIP_SHEETS:
            continue
        tables = _parse_sheet_tables(wb[sname], sname, meda_map)
        if tables:
            all_tables[sname] = tables

    total = sum(len(t) for t in all_tables.values())

    lines = [
        "# 全库表字段说明",
        "",
        "本文档包含所有模块的表字段定义（主键、可空性、字段类型）。",
        "测试时关注：**主键唯一性、非空字段强制校验、字段类型边界**。",
        "",
        f"共 {len(all_tables)} 个模块，{total} 张表。",
        "",
        "## 通用测试规则",
        "- [PK] 字段：必须测试重复值插入（应报唯一约束错误）",
        "- [非空] 字段：必须测试传入 NULL 的处理",
        "- RSKM_RULE 与 RSKM_FULL_RULE 系列：测试增量/全量数据一致性",
        "",
    ]

    for sname, tables in all_tables.items():
        lines.append(f"# 模块：{sname}（{len(tables)} 张表）")
        lines.append("")
        for tname, tinfo in tables.items():
            lines.append(f"## {tname}（{tinfo['cn']}）")
            if not tinfo['fields']:
                lines.append("_（字段信息待补充）_")
                lines.append("")
                continue
            lines.append("| 字段名 | 中文名 | 类型 | 约束 | 备注 |")
            lines.append("|--------|--------|------|------|------|")
            for f in tinfo['fields']:
                marks = []
                if f['pk']:                 marks.append('PK')
                if f['nullable'] == '非空': marks.append('非空')
                constraint = ' '.join(marks) if marks else '-'
                lines.append(
                    f"| {f['code']} | {f['name'] or '-'} | "
                    f"{f['type'] or '-'} | {constraint} | {f['memo'] or '-'} |"
                )
            lines.append("")

    return "\n".join(lines)


# ── 3. 表依赖关系 ──────────────────────────────────────────────────────────
def extract_table_deps(wb) -> str:
    """提取表依赖，用于测试数据准备顺序。"""
    ws = wb['表依赖']
    deps: dict = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue
        table    = row[2]
        is_main  = row[3] == 'Y'
        dep_table = row[4]
        if not dep_table or dep_table == '依赖表名':
            continue
        if table not in deps:
            deps[table] = {'main_deps': [], 'other_deps': []}
        if is_main:
            deps[table]['main_deps'].append(dep_table)
        else:
            deps[table]['other_deps'].append(dep_table)

    lines = [
        "# 表依赖关系",
        "",
        "本文档描述各数据表的依赖关系，用于确定测试数据的初始化顺序。",
        "",
        "**测试规则**:",
        "- 主依赖（Y标记）：必须先准备主依赖表的数据，否则关联查询返回空",
        "- 其他依赖：可选依赖，缺失时部分字段可能为空",
        "- 测试数据清理时，必须按依赖关系逆序删除（先删子表，再删父表）",
        "",
        f"共 {len(deps)} 张表有依赖关系。",
        "",
    ]

    for table, info in sorted(deps.items()):
        lines.append(f"## {table}")
        if info['main_deps']:
            lines.append(f"**主依赖（必须先初始化）**: {', '.join(info['main_deps'])}")
        if info['other_deps']:
            lines.append(f"**其他依赖**: {', '.join(info['other_deps'])}")
        lines.append("")

    return "\n".join(lines)


# ── 4. 元数据字段定义（核心字段摘要）──────────────────────────────────────
def extract_metadata_summary(wb) -> str:
    """提取元数据字段定义，聚焦有字典关联的字段（枚举字段）。"""
    ws = wb['元数据']
    enum_fields  = []  # 有 DICT_CATEGORY_CODE 的字段（枚举类型）
    string_fields = []  # VARCHAR 类型
    number_fields = []  # NUMBER 类型

    for row in ws.iter_rows(min_row=3, values_only=True):
        code  = row[1]
        name  = row[2]
        dtype = row[6]  # ORACLE_DEFN
        dict_code = row[4]  # DICT_CATEGORY_CODE
        if not code or not name:
            continue
        if dict_code:
            enum_fields.append((code, name, dtype, dict_code))
        elif dtype and 'VARCHAR' in str(dtype):
            string_fields.append((code, name, dtype))
        elif dtype and 'NUMBER' in str(dtype):
            number_fields.append((code, name, dtype))

    lines = [
        "# 元数据字段定义",
        "",
        "本文档聚焦三类高测试价值字段：枚举字段、字符串字段、数值字段。",
        "",
        f"共 {len(enum_fields)} 个枚举字段，"
        f"{len(string_fields)} 个字符串字段，"
        f"{len(number_fields)} 个数值字段。",
        "",
        "## 枚举字段（关联数据字典）",
        "这些字段的合法值由数据字典约束，测试时必须验证非法枚举值的处理。",
        "",
        "| 字段名 | 中文名 | 类型 | 关联字典代码 |",
        "|--------|--------|------|-------------|",
    ]
    for code, name, dtype, dict_code in enum_fields[:100]:  # 取前100个
        lines.append(f"| {code} | {name} | {dtype or '-'} | {dict_code} |")

    lines += [
        "",
        "## 字符串字段长度约束",
        "测试边界：传入超长字符串应截断或报错，不能静默截断。",
        "",
        "| 字段名 | 中文名 | 类型定义 |",
        "|--------|--------|---------|",
    ]
    for code, name, dtype in string_fields[:80]:
        lines.append(f"| {code} | {name} | {dtype} |")

    lines += [
        "",
        "## 数值字段精度约束",
        "测试边界：精度边界（如 NUMBER(8) 最大值为 99999999）、负值、0 值。",
        "",
        "| 字段名 | 中文名 | 类型定义 |",
        "|--------|--------|---------|",
    ]
    for code, name, dtype in number_fields[:80]:
        lines.append(f"| {code} | {name} | {dtype} |")

    return "\n".join(lines)


# ── 主函数 ─────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("用法: python kb_from_excel.py <xlsx路径>")
        print("示例: python kb_from_excel.py 华锐资管_规则配置_开发设计文档-资管风控表设计.xlsx")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"错误: 找不到文件 {xlsx_path}")
        sys.exit(1)

    print(f"读取: {xlsx_path.name}")
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    print(f"共 {len(wb.sheetnames)} 个 sheet\n")

    tasks = [
        ("数据字典枚举值.md",   extract_data_dict,       "数据字典"),
        ("全库表字段说明.md",   extract_risk_tables,     "全库表字段（228张）"),
        ("表依赖关系.md",       extract_table_deps,      "表依赖关系"),
        ("元数据字段定义.md",   extract_metadata_summary,"元数据字段"),
    ]

    for filename, func, label in tasks:
        out_path = KB_DIR / filename
        print(f"  生成 {label}...", end=" ", flush=True)
        try:
            content = func(wb)
            out_path.write_text(content, encoding="utf-8")
            lines = len(content.splitlines())
            size  = out_path.stat().st_size // 1024
            print(f"✓  {filename}  ({lines} 行, {size} KB)")
        except Exception as e:
            print(f"✗  失败: {e}")

    print(f"\n知识库已写入 knowledge_base/ 目录:")
    for f in sorted(KB_DIR.glob("*.md")):
        print(f"  {f.name}  ({f.stat().st_size // 1024} KB)")

    # 额外生成按模块拆分的表字段文件
    tables_dir = KB_DIR / "tables"
    tables_dir.mkdir(exist_ok=True)
    print("\n  生成按模块拆分的表字段说明...")

    # 重新用 data_only 打开获取计算值
    wb2 = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    meda_map2 = {}
    for row in wb2['元数据'].iter_rows(min_row=3, values_only=True):
        if row[1]:
            meda_map2[row[1]] = (row[2] or '', row[6] or '', row[7] or '')

    index_rows = []
    for sname, fname in SHEET_FILENAMES.items():
        if sname not in wb2.sheetnames:
            continue
        ws2 = wb2[sname]
        tables2 = {}
        cur2 = cur_cn2 = None
        for row in ws2.iter_rows(values_only=True):
            if not row or all(v is None for v in row[:4]):
                continue
            if row[0] == '表英文名':
                cur2 = row[1]; cur_cn2 = (row[5] if len(row) > 5 else '') or ''
                if cur2: tables2[cur2] = {'cn': cur_cn2, 'fields': []}
            elif cur2 and row[0] is None and row[1] and row[1] != '字段名':
                fc = row[1]; meta = meda_map2.get(fc, ('', '', ''))
                tables2[cur2]['fields'].append({
                    'code': fc, 'name': meta[0], 'type': meta[1],
                    'nullable': '可空' if row[3] != 'N' else '非空',
                    'pk': 'PK' if row[2] == 'Y' else '', 'memo': meta[2],
                })
        if not tables2:
            continue

        lines2 = [f"# {sname} 表字段说明", "",
                  f"共 {len(tables2)} 张表。",
                  "测试关注：[PK] 字段唯一性、[非空] 字段 NULL 处理、字段类型边界。", ""]
        for tname2, tinfo2 in tables2.items():
            lines2.append(f"## {tname2}（{tinfo2['cn']}）")
            if not tinfo2['fields']:
                lines2.append("_（字段信息待补充）_\n"); continue
            lines2 += ["| 字段名 | 中文名 | 类型 | 约束 | 备注 |",
                       "|--------|--------|------|------|------|"]
            for f2 in tinfo2['fields']:
                marks2 = (['PK'] if f2['pk'] else []) + (['非空'] if f2['nullable'] == '非空' else [])
                lines2.append(f"| {f2['code']} | {f2['name'] or '-'} | "
                              f"{f2['type'] or '-'} | {' '.join(marks2) or '-'} | {f2['memo'] or '-'} |")
            lines2.append("")

        out2 = tables_dir / f"{fname}.md"
        out2.write_text("\n".join(lines2), encoding="utf-8")
        size2 = out2.stat().st_size // 1024
        print(f"    ✓  {fname}.md  ({len(tables2)}张表, {size2}KB)")

        eg = ', '.join(list(tables2.keys())[:4])
        index_rows.append(f"| {fname}.md | {sname} | {len(tables2)}张表 | {eg}... |")

    # 写索引
    idx_lines = ["# 表字段说明索引", "",
                 "根据需求涉及的模块，用 read_file 读取对应文件。不要一次读取所有文件。", "",
                 "| 文件名 | 模块 | 规模 | 包含核心表（示例） |",
                 "|--------|------|------|-------------------|"] + index_rows + [""]
    (tables_dir / "00_索引.md").write_text("\n".join(idx_lines), encoding="utf-8")
    print(f"    ✓  00_索引.md（{len(index_rows)}个模块）")
    print(f"\n  表字段文件写入: knowledge_base/tables/")

    print("\n下一步:")
    print("  python agent.py <需求文档> --kb")
    print("  Agent 会先读 knowledge_base/tables/00_索引.md，再按需读具体模块文件")


if __name__ == "__main__":
    main()


# ── 5. 按模块拆分表字段说明 ────────────────────────────────────────────────
SHEET_FILENAMES = {
    '风控':                  '01_风控表',
    '盘中巡检表':             '02_盘中巡检表',
    '消息扩展元因子表结构':   '03_消息扩展元因子表',
    'jobdb表结构':            '04_jobdb',
    'userdb表结构':           '05_userdb',
    'domainconfigdb表结构':   '06_domainconfigdb',
    'commondb表结构':         '07_commondb',
    'basedb表结构':           '08_basedb',
    'bosdb表结构':            '09_bosdb',
    'operationdb表结构':      '10_operationdb',
    'ruledb表结构':           '11_ruledb',
    'tradingdb表结构':        '12_tradingdb',
}


def extract_split_tables(wb) -> tuple:
    """
    把每个 sheet 单独写成一个 md 文件，返回 (索引文本, 文件列表)。
    调用方负责写文件；此函数只返回内容字典 {filename: content}。
    """
    meda_map = {}
    for row in wb['元数据'].iter_rows(min_row=3, values_only=True):
        if row[1]:
            meda_map[row[1]] = (row[2] or '', row[6] or '', row[7] or '')

    results   = {}   # fname -> md_content
    index_rows = []  # 用于生成索引

    for sname, fname in SHEET_FILENAMES.items():
        if sname not in wb.sheetnames:
            continue
        ws     = wb[sname]
        tables = {}
        cur_name = cur_cn = None

        for row in ws.iter_rows(values_only=True):
            if not row or all(v is None for v in row[:4]):
                continue
            if row[0] == '表英文名':
                cur_name = row[1]
                cur_cn   = (row[5] if len(row) > 5 else '') or ''
                if cur_name:
                    tables[cur_name] = {'cn': cur_cn, 'fields': []}
            elif cur_name and row[0] is None and row[1] and row[1] != '字段名':
                fc   = row[1]
                meta = meda_map.get(fc, ('', '', ''))
                tables[cur_name]['fields'].append({
                    'code': fc, 'name': meta[0], 'type': meta[1],
                    'nullable': '可空' if row[3] != 'N' else '非空',
                    'pk': 'PK' if row[2] == 'Y' else '', 'memo': meta[2],
                })

        if not tables:
            continue

        tnames = ', '.join(list(tables.keys())[:5])
        more   = f'...等{len(tables)}张' if len(tables) > 5 else f'共{len(tables)}张'
        index_rows.append(f"| {fname}.md | {sname} | {more} | {tnames}{'...' if len(tables)>5 else ''} |")

        lines = [
            f"# {sname} 表字段说明",
            "",
            f"共 {len(tables)} 张表。",
            "测试关注：[PK] 字段唯一性、[非空] 字段 NULL 处理、字段类型边界。",
            "",
        ]
        for tname, tinfo in tables.items():
            lines.append(f"## {tname}（{tinfo['cn']}）")
            if not tinfo['fields']:
                lines.append("_（字段信息待补充）_\n")
                continue
            lines.append("| 字段名 | 中文名 | 类型 | 约束 | 备注 |")
            lines.append("|--------|--------|------|------|------|")
            for f in tinfo['fields']:
                marks = []
                if f['pk']:                 marks.append('PK')
                if f['nullable'] == '非空': marks.append('非空')
                lines.append(
                    f"| {f['code']} | {f['name'] or '-'} | "
                    f"{f['type'] or '-'} | {' '.join(marks) or '-'} | {f['memo'] or '-'} |"
                )
            lines.append("")

        results[fname + ".md"] = "\n".join(lines)

    index = "\n".join([
        "# 表字段说明索引",
        "",
        "根据需求涉及的模块，读取对应文件。不要一次读取所有文件。",
        "",
        "| 文件名 | 模块 | 规模 | 包含的核心表（示例）|",
        "|--------|------|------|---------------------|",
    ] + index_rows + [""])

    return index, results


def main_with_split(xlsx_path):
    """扩展 main：在原有4个文件基础上，额外生成按模块拆分的表字段文件。"""
    from pathlib import Path
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)

    tables_dir = KB_DIR / "tables"
    tables_dir.mkdir(exist_ok=True)

    print("\n  生成按模块拆分的表字段说明...")
    index_text, file_contents = extract_split_tables(wb)

    # 写索引文件
    idx_path = tables_dir / "00_索引.md"
    idx_path.write_text(index_text, encoding="utf-8")
    print(f"    ✓  00_索引.md")

    # 写每个模块文件
    for fname, content in file_contents.items():
        out = tables_dir / fname
        out.write_text(content, encoding="utf-8")
        size = out.stat().st_size // 1024
        lines = len(content.splitlines())
        print(f"    ✓  {fname}  ({lines}行, {size}KB)")

    print(f"\n  表字段文件写入: {tables_dir}")
    print(f"  共 {len(file_contents)} 个模块文件 + 1 个索引文件")
