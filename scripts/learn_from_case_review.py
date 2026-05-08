#!/usr/bin/env python3
"""
learn_from_case_review.py - 从人工评审后的用例 Excel 学习生成规则

用法:
    python scripts/learn_from_case_review.py output/xxx/testcases.xlsx

人工在 Excel 中填写：
    是否纳入用例库 = 是/否
    未纳入原因 = 不纳入时填写

脚本会输出学习报告，并把可复用规则写入长期记忆 testpoint_hints。
"""

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

from dotenv import load_dotenv

WORKDIR = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
SKILLS_DIR = WORKDIR / "skills"
OUTPUT_DIR = WORKDIR / "output"
load_dotenv(WORKDIR / ".env", override=True)


def read_skill(name: str) -> str:
    path = SKILLS_DIR / name / f"{name}.md"
    if not path.exists():
        path = SKILLS_DIR / name / "SKILL.md"
    return path.read_text(encoding="utf-8") if path.exists() else ""


def normalize_header(value) -> str:
    return str(value or "").strip().replace("\n", "")


def read_reviewed_cases(xlsx_path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("错误: 需要 openpyxl，请先安装 pip install -r requirements.txt")
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["测试用例"] if "测试用例" in wb.sheetnames else wb.active
    headers = [normalize_header(c.value) for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        if any(item.values()):
            rows.append(item)
    return rows


def build_summary(cases: list[dict]) -> dict:
    included = []
    excluded = []
    for case in cases:
        mark = str(case.get("是否纳入用例库") or "").strip()
        if mark == "是":
            included.append(case)
        elif mark == "否":
            excluded.append(case)
    return {"included": included, "excluded": excluded}


def extract_json(text: str) -> dict:
    text = text.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    try:
        return json.loads(text)
    except Exception:
        start = text.find("{")
        if start >= 0:
            try:
                return json.loads(text[start:])
            except Exception:
                pass
    return {}


def call_llm(skill: str, included: list[dict], excluded: list[dict]) -> dict:
    try:
        import anthropic
    except ImportError:
        print("错误: 需要 anthropic，请先安装 pip install -r requirements.txt")
        sys.exit(1)

    model = os.environ.get("ANTHROPIC_MODEL") or os.environ.get("DEFAULT_LLM_MODEL") or os.environ.get("MODEL_ID", "claude-sonnet-4-6")
    client = anthropic.Anthropic()

    def slim(case: dict) -> dict:
        return {
            "case_id": case.get("用例ID", ""),
            "testpoint_id": case.get("测试点ID", ""),
            "module": case.get("功能模块", ""),
            "title": case.get("用例标题", ""),
            "source": case.get("来源", ""),
            "priority": case.get("优先级", ""),
            "test_data": case.get("测试数据", ""),
            "steps": case.get("操作步骤", ""),
            "expected": case.get("预期结果", ""),
            "include": case.get("是否纳入用例库", ""),
            "exclusion_reason": case.get("未纳入原因", ""),
            "remarks": case.get("备注", ""),
        }

    prompt = (
        f"{skill}\n\n"
        "## 被纳入用例样本\n"
        f"{json.dumps([slim(c) for c in included[:30]], ensure_ascii=False, indent=2)}\n\n"
        "## 未纳入用例样本\n"
        f"{json.dumps([slim(c) for c in excluded[:80]], ensure_ascii=False, indent=2)}\n\n"
        "请根据上述人工评审结果进行学习总结。"
    )
    response = client.messages.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=5000,
    )
    text = "".join(b.text for b in response.content if hasattr(b, "text"))
    return extract_json(text)


def save_memory(hints: list[str], req_stem: str):
    sys.path.insert(0, str(SCRIPT_DIR))
    from memory_store import MemoryStore

    memory = MemoryStore(req_stem)
    for hint in hints:
        memory.save_testpoint_hint(f"用例人工评审经验: {hint}")


def write_report(result: dict, xlsx_path: Path) -> Path:
    out_dir = OUTPUT_DIR / "case_review_learning"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{xlsx_path.stem}_{int(time.time())}.md"

    lines = [
        f"# 用例评审学习报告 - {xlsx_path.name}",
        "",
        f"## 总结",
        result.get("summary", ""),
        "",
        "## 未纳入原因分组",
    ]
    for group in result.get("not_included_reason_groups", []):
        lines.append(f"### {group.get('reason_type', '未分类')}（{group.get('count', 0)}条）")
        lines.append(f"- 生成改进规则: {group.get('generation_rule', '')}")
        examples = group.get("examples", [])
        if examples:
            lines.append("- 示例:")
            lines.extend(f"  - {e}" for e in examples)
        lines.append("")

    lines.append("## 正向模式")
    lines.extend(f"- {x}" for x in result.get("positive_patterns", []))
    lines.append("")
    lines.append("## 写入长期记忆的规则")
    lines.extend(f"- {x}" for x in result.get("memory_hints", []))
    lines.append("")
    lines.append("## 提示词优化建议")
    lines.extend(f"- {x}" for x in result.get("prompt_suggestions", []))

    out_path.write_text("\n".join(lines), encoding="utf-8")
    return out_path


def main():
    parser = argparse.ArgumentParser(description="从人工评审后的用例 Excel 学习生成规则")
    parser.add_argument("xlsx", help="已人工填写是否纳入用例库的 testcases.xlsx")
    parser.add_argument("--no-memory", action="store_true", help="只生成报告，不写入长期记忆")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).resolve()
    if not xlsx_path.exists():
        print(f"错误: 文件不存在 {xlsx_path}")
        sys.exit(1)

    cases = read_reviewed_cases(xlsx_path)
    summary = build_summary(cases)
    included = summary["included"]
    excluded = summary["excluded"]
    print(f"读取用例 {len(cases)} 条：纳入 {len(included)} 条，未纳入 {len(excluded)} 条")

    if not excluded:
        print("没有发现“是否纳入用例库=否”的用例，无需学习。")
        return

    skill = read_skill("case-review-learning")
    result = call_llm(skill, included, excluded)
    if not result:
        print("学习总结失败：模型未返回合法 JSON")
        sys.exit(1)

    report = write_report(result, xlsx_path)
    print(f"学习报告: {report.relative_to(WORKDIR)}")

    hints = [h for h in result.get("memory_hints", []) if isinstance(h, str) and h.strip()]
    if hints and not args.no_memory:
        save_memory(hints, xlsx_path.stem)
        print(f"已写入长期记忆 {len(hints)} 条")


if __name__ == "__main__":
    main()
