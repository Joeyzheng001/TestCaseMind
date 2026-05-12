#!/usr/bin/env python3
"""
agent.py - 测试用例生成 Agent (v2)

用法:
    python scripts/agent.py requirements.md
    python scripts/agent.py requirements.md --kb        # 启用知识库
    python scripts/agent.py requirements.md --skip-review
    python scripts/agent.py requirements.md --no-cases  # 只生成测试点，不展开用例

输出:
    output/testpoints_<name>_<ts>.json    测试点
    output/testcases_<name>_<ts>.json     测试用例（JSON）
    output/testcases_<name>_<ts>.xlsx     测试用例（Excel）
    output/testpoints_<name>_<ts>.xmind  测试点思维导图

Harness: s05 Skills + s04 Subagent + s06 Context Compact
"""

import argparse
import threading as _threading
import json
import os
import re
import subprocess
import sys
import time
from contextlib import contextmanager
from pathlib import Path

from anthropic import Anthropic
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
from task_store import TaskStore
from memory_store import MemoryStore
from kb_rag import KBRetriever
from memory_rag import MemoryRAG

WORKDIR = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
# 项目运行以当前仓库 .env 为准，避免 shell 中残留的旧 API token 覆盖新配置。
load_dotenv(WORKDIR / ".env", override=True)
client = Anthropic()
MODEL = os.environ.get("ANTHROPIC_MODEL") or os.environ.get("DEFAULT_LLM_MODEL") or os.environ.get("MODEL_ID", "claude-sonnet-4-6")

KB_DIR     = WORKDIR / "knowledge_base"
SKILLS_DIR = WORKDIR / "skills"
OUTPUT_DIR = WORKDIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
SECTIONS_DIR = WORKDIR / ".sections"
SECTIONS_DIR.mkdir(exist_ok=True)
LOG_DIR = WORKDIR / "logs"
LOG_DIR.mkdir(exist_ok=True)
RUN_LOG_PATH: Path | None = None
TIMINGS: dict[str, float] = {}
_RAG_FAILURES: set[str] = set()
_RAG_RETRIEVERS: dict[tuple[str, str], object] = {}


class TeeStream:
    """把 stdout/stderr 同时写到控制台和本地日志文件。"""

    def __init__(self, *streams):
        self.streams = streams

    def write(self, data):
        for stream in self.streams:
            stream.write(data)
            stream.flush()

    def flush(self):
        for stream in self.streams:
            stream.flush()

    def isatty(self):
        return any(getattr(stream, "isatty", lambda: False)() for stream in self.streams)


def setup_run_logging(stem: str, ts: int) -> Path:
    """初始化本次运行日志，完整保留控制台输出，便于排查。"""
    global RUN_LOG_PATH
    safe_stem = re.sub(r"[^a-zA-Z0-9._\-\u4e00-\u9fff]+", "_", stem).strip("_") or "run"
    RUN_LOG_PATH = LOG_DIR / f"{safe_stem}_{ts}.log"
    log_fp = open(RUN_LOG_PATH, "a", encoding="utf-8", buffering=1)
    sys.stdout = TeeStream(sys.stdout, log_fp)
    sys.stderr = TeeStream(sys.stderr, log_fp)
    return RUN_LOG_PATH


@contextmanager
def timed_stage(name: str, label: str | None = None):
    """记录阶段耗时并打印，便于定位性能瓶颈。"""
    start = time.perf_counter()
    try:
        yield
    finally:
        elapsed = time.perf_counter() - start
        TIMINGS[name] = TIMINGS.get(name, 0.0) + elapsed
        shown = label or name
        print(f"  [耗时] {shown}: {elapsed:.1f}s")


def get_cached_retriever(kb_dir: Path, index_dir: Path | None = None):
    """同一次运行复用 RAG retriever；失败后不反复初始化。"""
    idx = index_dir or (WORKDIR / ".kb_index")
    key = (str(kb_dir.resolve()), str(idx.resolve()))
    failure_key = "|".join(key)
    if failure_key in _RAG_FAILURES:
        raise RuntimeError("本次运行中该 RAG 索引已初始化失败，跳过重复尝试")
    if key not in _RAG_RETRIEVERS:
        _RAG_RETRIEVERS[key] = KBRetriever(kb_dir=kb_dir, index_dir=idx)
    return _RAG_RETRIEVERS[key]

def get_run_dir(stem: str, ts: int) -> Path:
    """每次运行单独一个目录：output/<需求文件名>/<时间戳>/"""
    run_dir = OUTPUT_DIR / stem / str(ts)
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


def write_manifest(
    run_dir: Path,
    *,
    status: str,
    req_path: Path,
    model: str,
    use_kb: bool,
    no_cases: bool,
    section: str,
    review: dict,
    testpoints: list,
    testcases: list | None = None,
    files: dict | None = None,
    task_summary: str = "",
    warnings: list | None = None,
    timings: dict | None = None,
) -> Path:
    """写入标准运行清单，供 MCP/外部工具稳定读取产物与统计信息。"""
    files = files or {}
    warnings = warnings or []
    testcases = testcases or []

    by_source = {
        "REQ": sum(1 for t in testpoints if get_source(t) == "REQ"),
        "KB": sum(1 for t in testpoints if get_source(t) == "KB"),
        "RISK": sum(1 for t in testpoints if get_source(t) == "RISK"),
    }

    def _rel(p):
        if not p:
            return ""
        path = Path(p)
        try:
            return str(path.relative_to(WORKDIR))
        except Exception:
            return str(path)

    manifest = {
        "schema_version": "1.0",
        "status": status,
        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "model": model,
        "requirement": {
            "path": _rel(req_path),
            "name": req_path.name,
            "section": section or "",
        },
        "options": {
            "knowledge_base": use_kb,
            "no_cases": no_cases,
        },
        "review": {
            "score": review.get("score", "N/A") if isinstance(review, dict) else "N/A",
            "summary": review.get("summary", "") if isinstance(review, dict) else "",
            "risk_count": len(review.get("risk_flags", [])) if isinstance(review, dict) else 0,
            "testable_feature_count": len(review.get("testable_features", [])) if isinstance(review, dict) else 0,
        },
        "testpoints": {
            "total": len(testpoints),
            "by_source": by_source,
        },
        "testcases": {
            "total": len(testcases),
        },
        "files": {k: _rel(v) for k, v in files.items() if v},
        "timings": {
            k: round(v, 2)
            for k, v in (timings or TIMINGS).items()
        },
        "task_summary": task_summary,
        "warnings": warnings,
    }

    manifest_path = run_dir / "manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return manifest_path

# ── s06: Context Compact ───────────────────────────────────────────────────
COMPACT_THRESHOLD = 30000
KEEP_RECENT = 3
PRESERVE_TOOLS = {"read_file"}

# ── Rate limiter: 防止并行批次同时打 API 导致 529 ─────────────────────────
_RATE_LOCK = _threading.Lock()
_LAST_API_CALL = 0.0
_MIN_API_INTERVAL = 0.8   # 同一时刻只允许一个 API 调用


def _wait_rate_limit():
    global _LAST_API_CALL
    with _RATE_LOCK:
        now = time.time()
        gap = now - _LAST_API_CALL
        if gap < _MIN_API_INTERVAL:
            time.sleep(_MIN_API_INTERVAL - gap)
        _LAST_API_CALL = time.time()


def estimate_tokens(messages: list) -> int:
    return len(str(messages)) // 4


def micro_compact(messages: list) -> None:
    tool_results = []
    for msg in messages:
        if msg["role"] == "user" and isinstance(msg.get("content"), list):
            for part in msg["content"]:
                if isinstance(part, dict) and part.get("type") == "tool_result":
                    tool_results.append(part)
    if len(tool_results) <= KEEP_RECENT:
        return
    tool_name_map = {}
    for msg in messages:
        if msg["role"] == "assistant":
            content = msg.get("content", [])
            if isinstance(content, list):
                for block in content:
                    if hasattr(block, "type") and block.type == "tool_use":
                        tool_name_map[block.id] = block.name
    for result in tool_results[:-KEEP_RECENT]:
        if not isinstance(result.get("content"), str) or len(result["content"]) <= 100:
            continue
        tool_name = tool_name_map.get(result.get("tool_use_id", ""), "unknown")
        if tool_name in PRESERVE_TOOLS:
            continue
        result["content"] = f"[Previous: used {tool_name}]"


def auto_compact(messages: list, label: str = "") -> list:
    print(f"  [compact{' '+label if label else ''}] 压缩上下文...")
    conversation_text = json.dumps(messages, default=str, ensure_ascii=False)[-60000:]
    response = client.messages.create(
        model=MODEL,
        messages=[{"role": "user", "content":
            "请用中文简洁总结以下对话，保留：1)已完成工作 2)当前状态 3)关键结论。\n\n"
            + conversation_text}],
        max_tokens=2000,
    )
    summary = next((b.text for b in response.content if hasattr(b, "text")), "无摘要")
    return [{"role": "user", "content": f"[上下文已压缩]\n\n{summary}"}]


# ── s05: Skill Loader ──────────────────────────────────────────────────────
def load_skill(name: str) -> str:
    skill_file = find_skill_file(name)
    if not skill_file.exists():
        return f"[Skill '{name}' 不存在]"
    text = skill_file.read_text(encoding="utf-8")
    match = re.match(r"^---\n(.*?)\n---\n(.*)", text, re.DOTALL)
    body = match.group(2).strip() if match else text
    return f'<skill name="{name}">\n{body}\n</skill>'


def find_skill_file(name: str) -> Path:
    """优先读取命名后的技能文件，兼容旧版 SKILL.md。"""
    skill_dir = SKILLS_DIR / name
    named_file = skill_dir / f"{name}.md"
    if named_file.exists():
        return named_file
    return skill_dir / "SKILL.md"


# ── 工具实现 ───────────────────────────────────────────────────────────────
def safe_path(p: str) -> Path:
    path = (WORKDIR / p).resolve()
    if not path.is_relative_to(WORKDIR):
        raise ValueError(f"路径越界: {p}")
    return path


def run_read(path: str, limit: int = None) -> str:
    try:
        lines = safe_path(path).read_text(encoding="utf-8").splitlines()
        if limit and limit < len(lines):
            lines = lines[:limit] + [f"...（省略 {len(lines)-limit} 行）"]
        return "\n".join(lines)[:50000]
    except Exception as e:
        return f"Error: {e}"


def run_bash(command: str) -> str:
    blocked = ["rm -rf /", "sudo", "shutdown", "> /dev/"]
    if any(b in command for b in blocked):
        return "Error: 危险命令被拦截"
    try:
        r = subprocess.run(command, shell=True, cwd=WORKDIR,
                           capture_output=True, text=True, timeout=30)
        out = (r.stdout + r.stderr).strip()
        return out[:20000] if out else "(无输出)"
    except subprocess.TimeoutExpired:
        return "Error: 超时"


def run_write(path: str, content: str) -> str:
    try:
        fp = safe_path(path)
        fp.parent.mkdir(parents=True, exist_ok=True)
        fp.write_text(content, encoding="utf-8")
        return f"已写入 {len(content)} 字节 → {fp}"
    except Exception as e:
        return f"Error: {e}"


CHILD_TOOLS = [
    {"name": "read_file",
     "description": "读取文件内容",
     "input_schema": {"type": "object",
                      "properties": {"path": {"type": "string"},
                                     "limit": {"type": "integer"}},
                      "required": ["path"]}},
    {"name": "bash",
     "description": "运行 shell 命令",
     "input_schema": {"type": "object",
                      "properties": {"command": {"type": "string"}},
                      "required": ["command"]}},
    {"name": "write_file",
     "description": "写入文件",
     "input_schema": {"type": "object",
                      "properties": {"path": {"type": "string"},
                                     "content": {"type": "string"}},
                      "required": ["path", "content"]}},
    {"name": "load_skill",
     "description": "加载指定技能的完整知识",
     "input_schema": {"type": "object",
                      "properties": {"name": {"type": "string"}},
                      "required": ["name"]}},
    {"name": "todo_write",
     "description": "记录执行计划，开始工作前必须先调用此工具列出步骤",
     "input_schema": {"type": "object",
                      "properties": {
                          "todos": {"type": "array",
                                    "items": {"type": "string"},
                                    "description": "计划步骤列表，按执行顺序排列"}},
                      "required": ["todos"]}},
]

def run_todo_write(todos: list) -> str:
    """s03: 打印执行计划，给用户可见的进度反馈。"""
    import re as _re
    lines = ["📋 执行计划:"]
    for i, todo in enumerate(todos, 1):
        clean_todo = _re.sub(r"^\s*\d+[.、)]\s*", "", str(todo)).strip()
        lines.append(f"  {i}. {clean_todo}")
    plan = "\n".join(lines)
    print(f"\n{plan}\n")
    return plan


def describe_tool_call(name: str, args: dict) -> str:
    """把子代理工具调用转换成人能读懂的日志。"""
    args = args or {}
    if name == "todo_write":
        return "制定执行计划"
    if name == "load_skill":
        return f"加载技能: {args.get('name', '')}"
    if name == "read_file":
        path = args.get("path", "")
        limit = args.get("limit")
        suffix = f"，最多读取 {limit} 字符" if limit else ""
        return f"读取需求文档/上下文文件: {path}{suffix}"
    if name == "write_file":
        path = args.get("path", "")
        content_len = len(str(args.get("content", "")))
        return f"写入阶段结果: {path}（{content_len} 字符）"
    if name == "bash":
        cmd = args.get("command", "")
        return f"执行辅助命令: {cmd}"
    return f"调用工具: {name}"


CHILD_HANDLERS = {
    "read_file":  lambda **kw: run_read(kw["path"], kw.get("limit")),
    "bash":       lambda **kw: run_bash(kw["command"]),
    "write_file": lambda **kw: run_write(kw["path"], kw["content"]),
    "load_skill": lambda **kw: load_skill(kw["name"]),
    "todo_write": lambda **kw: run_todo_write(kw["todos"]),
}


# ── JSON 提取工具 ──────────────────────────────────────────────────────────
def extract_json(text: str, fallback, expect_list: bool = False):
    """从模型输出中提取 JSON，兼容代码块、说明文字、截断输出。"""
    import re as _re
    text = text.strip()
    # 去掉 ```json ... ``` 包裹
    text = _re.sub(r"^```json\s*", "", text)
    text = _re.sub(r"\s*```\s*$", "", text)
    text = text.strip()

    # 直接解析
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # 找第一个完整 JSON 块
    start_chars = ["[", "{"] if expect_list else ["{", "["]
    for ch in start_chars:
        idx = text.find(ch)
        if idx == -1:
            continue
        end_ch = "]" if ch == "[" else "}"
        depth = 0
        for i, c in enumerate(text[idx:], idx):
            if c == ch:   depth += 1
            elif c == end_ch:
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[idx:i+1])
                    except json.JSONDecodeError:
                        break
        # 若输出被截断（depth>0），尝试补全后解析
        if depth > 0 and ch == "[":
            truncated = text[idx:].rstrip().rstrip(",")
            # 逐步去掉末尾不完整的对象，直到能解析
            for end in range(len(truncated)-1, idx, -1):
                if truncated[end] == "}":
                    candidate = truncated[:end+1] + "]"
                    try:
                        result = json.loads(candidate)
                        print(f"  [warn] JSON 被截断，成功恢复 {len(result)} 条记录")
                        return result
                    except json.JSONDecodeError:
                        continue
        break

    print(f"  [warn] JSON 解析失败，原始输出前300字符:\n{text[:300]}")
    return fallback


def _retry_json_fix(raw_text: str, system: str = "", max_tokens: int = 4000,
                    expect_list: bool = False, label: str = "",
                    char_limit: int = 4000) -> list | dict:
    """调 API 把非标准 JSON 输出修复为合法 JSON，返回解析后的对象。

    用于在模型输出无法直接解析时做一次格式修复重试。
    返回 list（expect_list=True）或 dict，失败返回空 list/dict。
    """
    if not raw_text or len(raw_text) < 50:
        return [] if expect_list else {}
    print(f"  [重试] 非 JSON 输出，尝试格式修复{label}...")
    try:
        resp = client.messages.create(
            model=MODEL,
            system=system or (
                "你是 JSON 格式化工具。只输出合法 JSON 数组，不输出任何其他内容。"
            ),
            messages=[{"role": "user", "content": raw_text[:char_limit]}],
            max_tokens=max_tokens,
        )
        fix_text = "".join(b.text for b in resp.content if hasattr(b, "text"))
        fallback = [] if expect_list else {}
        data = extract_json(fix_text, fallback=fallback, expect_list=expect_list)
        if data:
            count = len(data) if isinstance(data, list) else ""
            print(f"  [重试] 格式修复成功{label} {count}")
        return data
    except Exception as e:
        print(f"  [重试] 格式修复失败: {e}")
        return [] if expect_list else {}


# ── s04: Subagent ──────────────────────────────────────────────────────────
def run_subagent(system: str, prompt: str, label: str = "") -> str:
    messages = [{"role": "user", "content": prompt}]
    print(f"\n  [{label}] 子代理启动...")
    for _ in range(40):
        micro_compact(messages)
        if estimate_tokens(messages) > COMPACT_THRESHOLD:
            messages[:] = auto_compact(messages, label)
        # 遇到 529/429 过载自动重试（最多4次，间隔递增）
        for _retry in range(4):
            try:
                _wait_rate_limit()   # 并行批次间错开，避免同时打 API
                response = client.messages.create(
                    model=MODEL, system=system, messages=messages,
                    tools=CHILD_TOOLS, max_tokens=8000,
                )
                break
            except Exception as e:
                err_str = str(e).lower()
                if "529" in str(e) or "529" in err_str or "overloaded" in err_str or "529" in err_str:
                    wait = (_retry + 1) * 20
                    print(f"  [529] API 过载，{wait}s 后重试 ({_retry+1}/4)...")
                    import time as _time; _time.sleep(wait)
                    if _retry == 3:
                        # s11: 失败不崩溃，返回错误标记
                        return f"__ERROR__: {e}"
                elif "rate_limit" in err_str or "429" in str(e):
                    wait = (_retry + 1) * 30
                    print(f"  [429] 限速，{wait}s 后重试...")
                    import time as _time; _time.sleep(wait)
                    if _retry == 3:
                        return f"__ERROR__: {e}"
                else:
                    # 其他错误直接返回，不重试
                    return f"__ERROR__: {e}"
        messages.append({"role": "assistant", "content": response.content})
        if response.stop_reason != "tool_use":
            break
        results = []
        for block in response.content:
            if block.type == "tool_use":
                handler = CHILD_HANDLERS.get(block.name)
                try:
                    output = handler(**block.input) if handler else f"未知工具: {block.name}"
                except Exception as e:
                    output = f"Error: {e}"
                print(f"    → {describe_tool_call(block.name, block.input)}")
                results.append({
                    "type": "tool_result",
                    "tool_use_id": block.id,
                    "content": str(output)[:30000],
                })
        messages.append({"role": "user", "content": results})
    return "".join(b.text for b in response.content if hasattr(b, "text")) or "(无输出)"


# ── 阶段一：需求评审 ───────────────────────────────────────────────────────
def stage1_review(req_path: Path, memory=None) -> dict:
    system = (
        "你是一名资深测试工程师，负责需求文档评审。"
        "执行步骤（严格按顺序，不得跳过）：\n"
        "1. 用 todo_write 列出执行计划\n"
        "2. 用 load_skill 加载 requirement-review 技能\n"
        "3. 用 read_file 读取需求文档\n"
        "4. 用 write_file 把评审结果 JSON 写入 output/_review_tmp.json\n"
        "【绝对禁止】：不得使用 bash 工具执行任何命令。\n"
        "【写入格式】：write_file 的 content 必须是合法 JSON 对象，"
        "以 { 开头以 } 结尾，不含任何其他文字、代码块或 markdown 标记。"
    )
    # 用向量检索找相关历史经验（比全量注入更精准）
    req_preview = req_path.read_text(encoding="utf-8")[:500] if req_path.exists() else ""
    if memory:
        try:
            from memory_rag import MemoryRAG as _MR
            _mr = _MR()
            mem_ctx = _mr.search(f"{req_path.stem} {req_preview}", top_k=5)
        except Exception:
            mem_ctx = memory.get_context_for_review()
    else:
        mem_ctx = ""
    prompt = (
        f"需求文档路径: {req_path.relative_to(WORKDIR)}\n\n"
        + (f"【历史经验参考】\n{mem_ctx}\n\n" if mem_ctx else "")
        + "请开始需求评审。"
    )
    result = run_subagent(system, prompt, label="需求评审")
    if result.startswith("__ERROR__"):
        print(f"  [s11] 需求评审失败: {result}，使用空评审结果继续")
        return {"testable_features": [], "risk_flags": [], "score": 0}

    # 优先从临时文件读取（子代理用 write_file 写入）
    review_tmp = OUTPUT_DIR / "_review_tmp.json"
    if review_tmp.exists():
        try:
            raw = review_tmp.read_text(encoding="utf-8")
            review_tmp.unlink(missing_ok=True)
            data = extract_json(raw, fallback={})
            if data and isinstance(data, dict):
                return data
        except Exception:
            pass

    # 降级：从返回文本解析
    data = extract_json(result, fallback={"testable_features": [], "risk_flags": []})

    # 如果解析失败，尝试格式修复
    if not data or not isinstance(data, dict):
        data = _retry_json_fix(
            result, system="把以下内容转换为合法 JSON 对象，只输出 JSON，不要其他文字：",
            max_tokens=2000, expect_list=False, char_limit=3000,
        )

    return data if isinstance(data, dict) else {"testable_features": [], "risk_flags": []}


# ── 阶段二：测试点生成 ─────────────────────────────────────────────────────
def _batch_gen_tp_section(title: str, content: str, review: dict,
                           start_id: int, skill_text: str = "") -> list:
    """单次 API 直调：输入一个小章节，输出该章节的测试点 JSON 数组。无工具调用。"""
    review_summary = review.get("summary", "")
    testable = review.get("testable_features", [])
    testable_str = "\n".join(f"- {f}" for f in testable[:20]) if testable else ""

    prompt = (
        f"{skill_text}\n\n"
        f"## 需求评审摘要\n{review_summary}\n\n"
        f"## 可测功能点（来自评审）\n{testable_str}\n\n"
        f"## 当前章节: {title}\n\n"
        f"{content}\n\n"
        f"---\n"
        f"请为以上「{title}」章节生成测试点。\n"
        f"- testpoint_id 从 TP-{start_id+1:03d} 开始递增\n"
        f"- source 固定填 REQ\n"
        f"- 每条测试点只表达一个清晰、单一、可验证的目标，不要合并多个独立路径\n"
        f"- 测试点不是测试用例；同一验证目标下的枚举值、边界值、正反路径、异常路径不要拆成重复测试点，应写入 case_split_hints\n"
        f"- functional_module 必须具体到业务场景，如「流通类型-流通股-期初计算」而非笼统的「因子参数」\n"
        f"- test_scenario 必须含具体测试数据，如「期初持仓1000股，买入500股，卖出200股，验证计算公式」\n"
        f"- source_ref 必须指出需求章节或原文依据\n"
        f"- preconditions 写明具体数据：产品代码、证券代码、持仓数量、表字段值等\n"
        f"- test_steps 是字符串数组，每步可执行：[\"步骤1: 在DWD_AST_PD_HLDP_INFO插入产品P001持仓1000股\", \"步骤2: ...\"]\n"
        f"- expected_result 是字符串数组，必须含：①具体计算过程和数值 ②断言条件，如 [\"期初1000+买入500-卖出200=1300\", \"因子值=1300\", \"单位=股\"]\n"
        f"- case_split_hints 是字符串数组，必须列出后续应展开的用例维度，如正常值、等于边界、触发边界、空值、重复数据、页面/导出一致性\n"
        f"- 金额/现金/数量/比例类因子必须先覆盖正常计算主路径；涉及来源表时，在 case_split_hints 中列出单笔取值、多笔同类型汇总、无数据、不同产品/日期隔离\n"
        f"- 若文档语句无法判断是唯一记录直接取值还是多笔汇总，必须同时保留两种用例拆分提示，并在 remarks 标记待澄清\n"
        f"- 公式类必须覆盖正常值、等于边界、触发边界、超出边界；业务边界不要误写成参数错误\n"
        f"- priority: P0=核心公式/主流程, P1=参数组合/边界, P2=异常/特殊场景\n\n"
        f"只输出 JSON 数组，以 [ 开头以 ] 结尾，不要任何其他文字、markdown 代码块或说明。"
    )

    for attempt in range(3):
        try:
            _wait_rate_limit()
            with client.messages.stream(
                model=MODEL, messages=[{"role": "user", "content": prompt}],
                max_tokens=16000,
            ) as stream:
                text = stream.get_final_text().strip()
            if not text:
                if attempt < 2:
                    import time as _t; _t.sleep(2)
                    continue
                print(f"    [{title}] 模型返回空响应（已重试{attempt+1}次）")
                return []
            data = extract_json(text, fallback=[], expect_list=True)
            if isinstance(data, list) and data:
                result = []
                for i, tp in enumerate(data):
                    result.append(normalize_testpoint(tp, start_id + i))
                return result
            elif attempt < 2:
                import time as _t; _t.sleep(1)
                continue
            else:
                preview = text[:200].replace("\n", "\\n")
                print(f"    [{title}] JSON 解析失败: {preview}")
                return []
        except Exception as e:
            if attempt < 2:
                import time as _t; _t.sleep(2)
                continue
            print(f"    [{title}] API 异常: {e}")
            return []
    return []


def _split_doc_by_sections(doc_text: str) -> list:
    """按 ### / #### 标题拆分文档为多个章节，返回 [(标题, 内容), ...]"""
    import re as _re_sec
    lines = doc_text.splitlines()
    sections = []
    cur_title = ""
    cur_lines = []

    for line in lines:
        m = _re_sec.match(r'^(#{1,4})\s+(.+)$', line)
        if m:
            level = len(m.group(1))
            title = m.group(2).strip()
            if level <= 4 and not title.startswith("目录"):
                # 遇到新章节头，保存上一个
                if cur_lines and cur_title:
                    sections.append((cur_title, "\n".join(cur_lines)))
                cur_title = title
                cur_lines = [line]
                continue
        cur_lines.append(line)

    if cur_lines and cur_title:
        sections.append((cur_title, "\n".join(cur_lines)))
    return sections


def stage2_testpoints(req_path: Path, review: dict, use_kb: bool, memory=None) -> list:
    """
    两阶段测试点生成：
    阶段A：分批直调 API，每章节一次请求（无工具调用，避免多轮上下文膨胀）
    阶段B：读知识库 + 因子设计，补充 KB/RISK 测试点（按需）
    """
    # ── 加载 testpoint-gen 技能文本 ──────────────────────────────────────
    skill_text = ""
    skill_path = find_skill_file("testpoint-gen")
    if skill_path.exists():
        try:
            skill_text = skill_path.read_text(encoding="utf-8")
        except Exception:
            pass

    # ── 阶段 A：拆分文档，分批生成 REQ 测试点 ─────────────────────────────
    print(f"\n  [测试点-需求文档] 分批模式启动...")
    doc_text = req_path.read_text(encoding="utf-8")
    sections = _split_doc_by_sections(doc_text)
    print(f"  文档拆分为 {len(sections)} 个章节: {[t for t, _ in sections]}")

    eligible = []
    for title, content in sections:
        if len(content.strip()) < 80:
            print(f"    [{title}] 跳过（内容过短）")
            continue
        eligible.append((title, content))

    req_tps = []
    if eligible:
        max_workers = min(3, len(eligible))
        print(f"  阶段A并行生成: {len(eligible)} 个章节，并发 {max_workers}，完成后按原章节顺序合并")

        def run_one(item):
            idx, (title, content) = item
            print(f"    [{title}] 生成中（{len(content)} 字）...", flush=True)
            started = time.perf_counter()
            batch = _batch_gen_tp_section(title, content, review, 0, skill_text)
            elapsed = time.perf_counter() - started
            return idx, title, batch, elapsed

        results = {}
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(run_one, item): item[0]
                for item in enumerate(eligible)
            }
            for future in as_completed(futures):
                try:
                    idx, title, batch, elapsed = future.result()
                    results[idx] = batch
                    print(f"    [{title}] 完成 {len(batch)} 条，用时 {elapsed:.1f}s")
                except Exception as e:
                    idx = futures[future]
                    title = eligible[idx][0]
                    results[idx] = []
                    print(f"    [{title}] 异常，跳过: {e}")

        for idx in range(len(eligible)):
            for tp in results.get(idx, []):
                tp["testpoint_id"] = f"TP-{len(req_tps) + 1:03d}"
                req_tps.append(tp)

    print(f"  阶段A完成: {len(req_tps)} 条 REQ 测试点")

    if not req_tps:
        print(f"  [阶段A] 未生成 REQ 测试点，跳过此章节继续")
        return []

    if not use_kb or not KB_DIR.exists():
        return req_tps

    # ── 阶段 B：RAG 检索知识库 → KB/RISK 补充（语义检索，精准无截断）──────────
    print(f"\n  [测试点-知识库补充] RAG 语义检索...", flush=True)

    try:
        retriever  = get_cached_retriever(KB_DIR)
        # 用需求文档内容 + 评审结果做检索查询
        req_text   = req_path.read_text(encoding="utf-8")
        review_str = json.dumps(review, ensure_ascii=False)
        query      = f"{req_path.stem}\n{req_text[:1000]}\n{review_str[:500]}"
        with timed_stage("rag.kb_search", "知识库 RAG 检索"):
            kb_context = retriever.search_for_requirement(query, top_k=12)
    except Exception as e:
        _RAG_FAILURES.add("|".join((str(KB_DIR.resolve()), str((WORKDIR / ".kb_index").resolve()))))
        print(f"  [RAG] 检索失败: {e}，跳过知识库补充")
        return req_tps

    # 额外：用 RAG 从因子设计文档里检索相关段落
    # 用相关度阈值+token预算双重控制，自适应决定注入多少内容
    design_context = ""
    design_dir = KB_DIR / "design"
    if design_dir.exists() and list(design_dir.glob("*.md")):
        try:
            design_retriever = get_cached_retriever(design_dir, WORKDIR / ".design_index")
            design_query = f"{req_path.stem}\n{req_text[:800]}"
            # 先取较多候选，再按阈值和预算过滤，最终注入的是真正相关的段落
            with timed_stage("rag.design_search", "设计文档 RAG 检索"):
                candidates = design_retriever.search(design_query, top_k=20)

            SCORE_THRESHOLD = 0.60   # 低于此相关度的段落不采用
            TOKEN_BUDGET    = 4000   # 最多注入的字符数

            selected = []
            total_chars = 0
            for hit in candidates:
                if hit["score"] < SCORE_THRESHOLD:
                    break  # 结果已按相关度降序，后面的更低可以直接停
                if total_chars + len(hit["content"]) > TOKEN_BUDGET:
                    break
                selected.append(hit)
                total_chars += len(hit["content"])

            if selected:
                lines = ["【因子设计文档相关内容（RAG检索）】\n"]
                prev_source = None
                for hit in selected:
                    if hit["source"] != prev_source:
                        lines.append(f"\n--- {hit['source']} (相关度:{hit['score']:.2f}) ---")
                        prev_source = hit["source"]
                    lines.append(hit["content"])
                design_context = "\n".join(lines)
                unique_files = {h["source"] for h in selected}
                print(f"  [设计文档] 纳入 {len(selected)} 段（阈值≥{SCORE_THRESHOLD}，"
                      f"共{total_chars}字），来自 {len(unique_files)} 个文件", flush=True)
                for f in unique_files:
                    print(f"    ✓ {f}", flush=True)
            else:
                print(f"  [设计文档] 无相关内容（最高相关度: "
                      f"{candidates[0]['score']:.2f} < {SCORE_THRESHOLD}）" if candidates else
                      f"  [设计文档] 无候选内容", flush=True)
        except Exception as e:
            _RAG_FAILURES.add("|".join((str(design_dir.resolve()), str((WORKDIR / ".design_index").resolve()))))
            print(f"  [warn] 设计文档RAG检索失败: {e}", flush=True)

    if not kb_context and not design_context:
        print(f"  阶段B完成: 0 条（知识库无相关内容）")
        return req_tps

    # 直接调一次 API，用检索结果生成 KB/RISK 测试点
    offset   = len(req_tps)
    system_b = "你是一名资深测试工程师，专门生成 KB 和 RISK 来源的测试点。只输出 JSON 数组，不要其他文字。"

    # 合并 RAG 结果和设计文档内容
    full_context = kb_context
    if design_context:
        full_context = design_context + "\n\n" + kb_context if kb_context else design_context

    prompt_b = (
        f"需求文档: {req_path.name}\n\n"
        f"{full_context}\n\n"
        f"基于以上知识库内容，为需求文档生成 KB 和 RISK 来源的测试点：\n"
        f"- KB 测试点：针对知识库中的枚举值、字段约束、数据表取值逻辑、开发设计实现逻辑生成，"
        f"source_ref 填写知识库来源文件名或段落\n"
        f"- RISK 测试点：并发竞争、数据精度丢失、外部依赖失败、数据同步延迟、主备数据源切换等，至少3条\n"
        f"- 每条测试点只表达一个清晰验证目标，不要合并多个独立路径\n"
        f"- 同一验证目标下的枚举值、边界值、正反路径、异常路径写入 case_split_hints，供用例阶段展开\n"
        f"- test_steps 和 expected_result 必须是字符串数组，且包含可执行数据和可验证断言\n"
        f"- testpoint_id 从 TP-{offset+1:03d} 开始递增\n"
        f"- source 字段只能填 KB 或 RISK\n\n"
        "输出纯 JSON 数组，格式：\n"
        '[{"testpoint_id":"TP-xxx","functional_module":"xxx","test_scenario":"xxx",'
        '"source":"KB","source_ref":"来源文件名","preconditions":"xxx","test_steps":["步骤1"],'
        '"expected_result":["预期1"],"case_split_hints":["正常路径","边界路径","异常路径"],'
        '"priority":"P1","remarks":""}]'
    )

    try:
        response = client.messages.create(
            model=MODEL,
            system=system_b,
            messages=[{"role": "user", "content": prompt_b}],
            max_tokens=4000,
        )
        result_b = "".join(b.text for b in response.content if hasattr(b, "text"))
        data     = extract_json(result_b, fallback=[], expect_list=True)
        kb_tps   = []
        if isinstance(data, list):
            kb_tps = [normalize_testpoint(tp, offset + i) for i, tp in enumerate(data)]
        print(f"  阶段B完成: {len(kb_tps)} 条 KB/RISK 测试点")
    except Exception as e:
        print(f"  [s11] 阶段B失败: {e}")
        kb_tps = []

    return req_tps + kb_tps


# ── 阶段三：测试用例生成（分批处理，每批10条）──────────────────────────────
BATCH_SIZE = 10

def stage3_testcases_batch(batch: list, batch_no: int, case_id_start: int,
                           memory_context: str = "") -> list:
    """分批直调 API：传入测试点，直接返回用例 JSON，无工具调用。"""
    # 加载 testcase-gen 技能文本
    skill_text = ""
    skill_path = find_skill_file("testcase-gen")
    if skill_path.exists():
        try:
            skill_text = skill_path.read_text(encoding="utf-8")
        except Exception:
            pass

    batch_json = json.dumps(batch, ensure_ascii=False, indent=2)
    prompt = (
        f"{skill_text}\n\n"
        f"## 测试点列表（共 {len(batch)} 条）\n\n"
        f"```json\n{batch_json}\n```\n\n"
        + (f"## 历史人工评审经验\n{memory_context}\n\n" if memory_context else "")
        +
        f"---\n"
        f"将以上测试点展开为完整测试用例，严格遵循以下规则：\n"
        f"- case_id 从 TC-{case_id_start:03d} 开始递增\n"
        f"- 测试点和测试用例不是一对一关系：一个测试点可生成 1 条或多条用例\n"
        f"- 如果测试点包含 case_split_hints，必须优先逐项展开；每个有效拆分项至少形成 1 条用例\n"
        f"- 每个独立路径、条件分支、枚举值、边界值、异常路径都应单独形成用例，覆盖完整为止\n"
        f"- 金额/现金/数量/比例类测试点必须至少生成正常计算主路径用例；涉及来源表时优先覆盖单笔取值、多笔同类型汇总、无数据、不同产品/日期隔离\n"
        f"- 不要把多个有意义的路径合并进一条用例，也不要为了凑数生成重复用例\n"
        f"- 只有单一路径、无枚举、无边界、无依赖状态、无输出渠道差异的原子测试点，才允许只生成 1 条用例\n"
        f"- 若本批输出用例数等于测试点数，必须重新检查并补拆枚举、边界、空值、重复数据、主备数据源、展示/导出/接口一致性\n"
        f"- 每条用例包含: case_id, testpoint_id, functional_module, case_title, "
        f"source, priority, preconditions, test_data, steps, expected_result, "
        f"include_in_case_library, exclusion_reason, remarks\n"
        f"- include_in_case_library 固定输出空字符串，留给人工评审填写“是/否”\n"
        f"- exclusion_reason 固定输出空字符串，人工标记“否”时填写未纳入原因\n"
        f"- test_data: 给出具体的测试输入值（如 期初持仓=10000股, 买入成交=5000股, 卖出=2000股）\n"
        f"- steps: 字符串数组，每步包含具体操作和数据，如 [\"在DWD_AST_PD_HLDP_INFO插入产品P001, 证券600001, 持仓10000股\", ...]\n"
        f"- expected_result: 字符串数组，必须包含：①具体计算数值和公式 ②可验证的断言，如 [\"计算: 期初10000 + 买入5000 - 卖出2000 = 13000\", \"因子值=13000\", \"单位=股\"]\n"
        f"- preconditions: 写明产品/证券/表名/字段等具体信息\n"
        f"- case_title 必须区分等价类/边界值的不同取值（如「流通股类型=流通股」vs「流通股类型=非流通股」）\n\n"
        f"只输出 JSON 数组，以 [ 开头以 ] 结尾，不要任何其他文字、markdown 代码块或说明。"
    )

    for attempt in range(3):
        try:
            _wait_rate_limit()
            with client.messages.stream(
                model=MODEL, messages=[{"role": "user", "content": prompt}],
                max_tokens=32000,
            ) as stream:
                text = stream.get_final_text().strip()
            if not text:
                if attempt < 2:
                    import time as _t; _t.sleep(2)
                    continue
                print(f"  [batch{batch_no}] 模型返回空响应")
                return []
            data = extract_json(text, fallback=[], expect_list=True)
            if isinstance(data, list) and data:
                offset = (batch_no - 1) * BATCH_SIZE
                result = [normalize_testcase(c, offset + i + 1) for i, c in enumerate(data)]
                return result
            elif attempt < 2:
                import time as _t; _t.sleep(1)
                continue
            else:
                preview = text[:200].replace("\n", "\\n")
                print(f"  [batch{batch_no}] JSON 解析失败: {preview}")
                return []
        except Exception as e:
            if attempt < 2:
                import time as _t; _t.sleep(2)
                continue
            print(f"  [batch{batch_no}] API 异常: {e}")
            return []
    return []


def stage3_testcases(testpoints: list, req_path: Path, memory=None) -> list:
    """分批调子代理，并行处理所有批次。"""
    batches = [testpoints[i:i+BATCH_SIZE] for i in range(0, len(testpoints), BATCH_SIZE)]
    total   = len(batches)
    print(f"  共 {len(testpoints)} 条测试点，分 {total} 批并行处理（每批 {BATCH_SIZE} 条）")

    # 并行数不超过批次数，也不超过3（限速锁 + 降低并发避免 API 过载）
    max_workers = min(total, 3)
    results     = {}   # batch_no -> cases
    memory_context = memory.get_context_for_testpoints() if memory else ""

    def run_batch(args):
        batch_no, batch = args
        # case_id 按批次固定偏移，不依赖其他批次完成顺序
        case_id_start = (batch_no - 1) * BATCH_SIZE + 1
        cases = stage3_testcases_batch(batch, batch_no, case_id_start, memory_context)
        return batch_no, cases

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(run_batch, (i, batch)): i
            for i, batch in enumerate(batches, 1)
        }
        for future in as_completed(futures):
            try:
                batch_no, cases = future.result()
                results[batch_no] = cases
                print(f"  batch{batch_no}/{total} 完成，本批生成 {len(cases)} 条用例")
            except Exception as e:
                batch_no = futures[future]
                print(f"  [s11] batch{batch_no} 异常，跳过: {e}")
                results[batch_no] = []

    # 按批次顺序合并（并行完成顺序不定）
    all_cases = []
    for i in range(1, total + 1):
        all_cases.extend(results.get(i, []))

    # 分批并行时，每批只能预估起始编号；如果前序批次展开了更多用例，
    # 后续批次的 case_id 可能撞号。合并后统一重排，保证交付物编号稳定唯一。
    for idx, case in enumerate(all_cases, 1):
        case["case_id"] = f"TC-{idx:03d}"

    return all_cases


# ── 字段标准化 ────────────────────────────────────────────────────────────────
def normalize_testcase(case: dict, idx: int) -> dict:
    """
    把模型可能输出的各种字段名统一映射到标准字段。
    兼容: case_name/title/test_name → case_title
          pre_condition/precondition → preconditions
          step/test_steps/procedure → steps
          expected/result → expected_result
          module/test_module → functional_module
    """
    # 字段别名映射表
    alias = {
        "case_title":        ["case_name", "title", "test_name", "用例标题", "case_description"],
        "functional_module": ["module", "test_module", "feature", "功能模块"],
        "preconditions":     ["precondition", "pre_condition", "prerequisite", "前置条件"],
        "test_data":         ["input_data", "test_input", "data", "测试数据"],
        "steps":             ["step", "test_steps", "procedure", "操作步骤", "test_procedure"],
        "expected_result":   ["expected", "result", "expect", "预期结果", "expected_output"],
        "source":            ["test_source", "来源"],
        "priority":          ["level", "test_priority", "优先级"],
        "include_in_case_library": ["是否纳入用例库", "include", "included", "case_library"],
        "exclusion_reason":  ["未纳入原因", "exclude_reason", "not_include_reason", "reason"],
        "remarks":           ["remark", "note", "comment", "备注"],
    }

    normalized = dict(case)  # 先复制原始数据

    # 应用别名映射（只在标准字段不存在时才映射）
    for std_key, aliases in alias.items():
        if std_key not in normalized or not normalized[std_key]:
            for a in aliases:
                if a in normalized and normalized[a]:
                    normalized[std_key] = normalized[a]
                    break

    # 确保所有标准字段都存在
    defaults = {
        "case_id":           f"TC-{idx:03d}",
        "testpoint_id":      "",
        "functional_module": "",
        "case_title":        "",
        "source":            "REQ",
        "priority":          "P1",
        "preconditions":     "",
        "test_data":         "",
        "steps":             "",
        "expected_result":   "",
        "include_in_case_library": "",
        "exclusion_reason":  "",
        "actual_result":     "",
        "status":            "",
        "remarks":           "",
    }
    for k, v in defaults.items():
        if k not in normalized or normalized[k] is None:
            normalized[k] = v

    # steps 如果是列表，转成换行字符串
    if isinstance(normalized.get("steps"), list):
        normalized["steps"] = "\n".join(
            f"{i+1}. {s}" for i, s in enumerate(normalized["steps"])
        )

    return normalized


# ── 字段标准化：测试点 ──────────────────────────────────────────────────────
def normalize_testpoint(tp: dict, idx: int = 0) -> dict:
    """
    统一测试点字段名，兼容模型各种输出格式。
    支持字段名：id/title/level/desc/expected 等非标准字段。
    """
    n = dict(tp)

    # ── testpoint_id ──────────────────────────────────────────────────────
    if not n.get("testpoint_id"):
        n["testpoint_id"] = (n.get("id") or n.get("tp_id") or
                             n.get("case_id") or f"TP-{idx+1:03d}")

    # ── test_scenario（标题）─────────────────────────────────────────────
    if not n.get("test_scenario"):
        n["test_scenario"] = (n.get("title") or n.get("name") or
                              n.get("case_title") or n.get("scenario") or "")

    # ── priority（优先级）注意：必须先检查 level，再用默认值 ─────────────
    # 字段可能叫 level / test_priority / 优先级，值可能是 P0/P1/P2
    raw_priority = (n.get("priority") or n.get("level") or
                    n.get("test_priority") or n.get("优先级") or "P1")
    # 标准化：只接受 P0/P1/P2
    if raw_priority not in ("P0", "P1", "P2"):
        raw_priority = "P1"
    n["priority"] = raw_priority

    # ── functional_module（功能模块）─────────────────────────────────────
    # 优先从显式字段取，没有则从 title 推断关键词分组
    if not n.get("functional_module"):
        mod = (n.get("module") or n.get("feature") or
               n.get("category") or n.get("functional_area") or "")
        if not mod:
            # 从 title 提取：取"-"前的部分作为分组名
            title = n.get("test_scenario", "")
            if "-" in title:
                mod = title.split("-")[0].strip()
            elif "验证" in title or "计算" in title:
                mod = title[:8].strip()
            else:
                mod = "功能验证"
        n["functional_module"] = mod

    # ── source ────────────────────────────────────────────────────────────
    src = n.get("source") or n.get("test_source") or n.get("来源") or "REQ"
    if src not in ("REQ", "KB", "RISK"):
        src = "REQ"
    n["source"] = src

    # ── expected_result ───────────────────────────────────────────────────
    if not n.get("expected_result"):
        n["expected_result"] = (n.get("expected") or n.get("expect") or
                                n.get("expected_output") or "")

    # ── preconditions ─────────────────────────────────────────────────────
    if not n.get("preconditions"):
        n["preconditions"] = (n.get("precondition") or n.get("pre_condition") or
                              n.get("desc") or n.get("description") or "")

    # ── source_ref ────────────────────────────────────────────────────────
    if not n.get("source_ref"):
        n["source_ref"] = n.get("source_reference") or n.get("ref") or ""

    # ── remarks ───────────────────────────────────────────────────────────
    if not n.get("remarks"):
        n["remarks"] = n.get("remark") or n.get("note") or n.get("comment") or ""

    # ── case_split_hints ──────────────────────────────────────────────────
    if not n.get("case_split_hints"):
        hints = n.get("split_hints") or n.get("case_dimensions") or n.get("用例拆分提示")
        if hints:
            n["case_split_hints"] = hints if isinstance(hints, list) else [str(hints)]
        else:
            n["case_split_hints"] = ["原子验证点: 单一路径，无额外等价类或边界"]

    return n


# ── 输出：Excel ────────────────────────────────────────────────────────────
def export_excel(testcases: list, out_path: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  [skip] Excel 输出需要 openpyxl: pip install openpyxl")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    columns = [
        ("用例ID",      "case_id",          12),
        ("测试点ID",    "testpoint_id",      12),
        ("功能模块",    "functional_module", 18),
        ("用例标题",    "case_title",        35),
        ("来源",        "source",             8),
        ("优先级",      "priority",           8),
        ("前置条件",    "preconditions",     25),
        ("测试数据",    "test_data",         20),
        ("操作步骤",    "steps",             40),
        ("预期结果",    "expected_result",   35),
        ("是否纳入用例库", "include_in_case_library", 16),
        ("未纳入原因",  "exclusion_reason",  28),
        ("实际结果",    "actual_result",     25),
        ("执行状态",    "status",            10),
        ("备注",        "remarks",           20),
    ]

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="2B5FA8")
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    source_colors = {"REQ": "DDEEFF", "KB": "FFF9DD", "RISK": "FFE8E8"}

    for col_idx, (col_name, _, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = col_width
    ws.row_dimensions[1].height = 22

    priority_colors = {"P0": "FF4444", "P1": "FF8800", "P2": "888888"}

    for row_idx, case in enumerate(testcases, 2):
        case     = normalize_testcase(case, row_idx - 1)
        source   = case.get("source", "REQ")
        row_fill = PatternFill("solid", fgColor=source_colors.get(source, "FFFFFF"))

        for col_idx, (_, field_key, _) in enumerate(columns, 1):
            value = case.get(field_key, "")
            if isinstance(value, list):
                value = "\n".join(str(v) for v in value)
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.fill   = row_fill
            if field_key in ("case_id", "testpoint_id", "source", "priority",
                             "include_in_case_library", "status"):
                cell.alignment = center
                if field_key == "priority":
                    cell.font = Font(color=priority_colors.get(value, "000000"), bold=True)
            else:
                cell.alignment = left_wrap
        ws.row_dimensions[row_idx].height = 45

    ws.freeze_panes = "A2"
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        include_col = next(i for i, (_, key, _) in enumerate(columns, 1)
                           if key == "include_in_case_library")
        col_letter = ws.cell(row=1, column=include_col).column_letter
        dv = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}1048576")
    except Exception:
        pass

    legend_ws = wb.create_sheet("图例说明")
    for r, (a, b) in enumerate([
        ("颜色", "含义"), ("蓝色底", "REQ — 来自需求文档"),
        ("黄色底", "KB  — 来自知识库补充"), ("红色底", "RISK — 风险推断"),
        ("", ""), ("优先级", "说明"), ("P0", "核心必测"),
        ("P1", "重要应测"), ("P2", "边缘可测"),
        ("", ""), ("是否纳入用例库", "人工评审填写：是/否"),
        ("未纳入原因", "当是否纳入用例库=否时填写，用于后续学习"),
    ], 1):
        legend_ws.cell(row=r, column=1, value=a)
        legend_ws.cell(row=r, column=2, value=b)

    wb.save(out_path)
    return True


def _review_issues(review: dict) -> list:
    """兼容不同评审 schema，提取 issues 列表。"""
    if not isinstance(review, dict):
        return []
    issues = review.get("issues")
    if isinstance(issues, list):
        return issues

    collected = []
    for key in ("completeness_issues", "clarity_issues", "consistency_issues",
                "performance_issues", "integration_issues", "risk_flags"):
        values = review.get(key, [])
        if isinstance(values, list):
            for item in values:
                if isinstance(item, dict):
                    collected.append(item)
                else:
                    collected.append({"type": key, "question": str(item)})
    return collected


def _mindmap_to_markdown(node: dict, level: int = 1) -> list:
    """把 requirement-review 输出的树状 mindmap 转成 XMind 可导入 Markdown。"""
    if not isinstance(node, dict):
        return []
    title = node.get("主题") or node.get("title") or node.get("name") or "未命名节点"
    lines = [f"{'#' * max(1, min(level, 6))} {title}"]
    children = node.get("子主题") or node.get("children") or []
    if isinstance(children, list):
        for child in children:
            lines.extend(_mindmap_to_markdown(child, level + 1))
    return lines


def export_review_markdown(review: dict, req_name: str, out_path: Path) -> bool:
    """导出可读需求评审报告。"""
    if not isinstance(review, dict):
        review = {}

    lines = [
        f"# 需求评审报告 - {req_name}",
        "",
        "## 质量结论",
        f"- 质量等级: {review.get('quality_grade') or review.get('grade') or review.get('quality_assessment', {}).get('grade') or 'N/A'}",
        f"- 质量分: {review.get('score', 'N/A')}",
        f"- 概要理由: {review.get('summary') or review.get('quality_assessment', {}).get('summary_reason') or ''}",
        "",
    ]

    dimension_scores = review.get("dimension_scores") or {}
    if isinstance(dimension_scores, dict) and dimension_scores:
        lines.extend(["## 维度评分", "", "| 维度 | 权重 | 得分 | 理由 |", "| --- | ---: | ---: | --- |"])
        for name, info in dimension_scores.items():
            if isinstance(info, dict):
                lines.append(
                    f"| {name} | {info.get('weight', '')} | {info.get('score', '')} | "
                    f"{str(info.get('reason', '')).replace('|', '/')} |"
                )
        lines.append("")

    issues = _review_issues(review)
    lines.extend(["## 问题清单", ""])
    if issues:
        for i, item in enumerate(issues, 1):
            if not isinstance(item, dict):
                item = {"question": str(item)}
            title = item.get("question") or item.get("desc") or item.get("description") or item.get("title") or ""
            lines.append(f"### ISSUE-{i:03d} {title}")
            for label, key in (
                ("类型", "type"), ("优先级", "priority"), ("位置", "location"),
                ("原因", "reason"), ("影响", "impact"), ("建议", "suggestion"),
            ):
                value = item.get(key)
                if value:
                    lines.append(f"- {label}: {value}")
            lines.append("")
    else:
        lines.append("暂无问题。")
        lines.append("")

    mindmap = review.get("mindmap")
    if isinstance(mindmap, dict):
        lines.append("## 测试思维导图")
        lines.append("")
        lines.extend(_mindmap_to_markdown(mindmap, 3))
        lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    return True


def export_review_issues_excel(review: dict, out_path: Path) -> bool:
    """导出需求评审问题清单 Excel。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  [skip] 评审问题 Excel 输出需要 openpyxl: pip install openpyxl")
        return False

    issues = _review_issues(review)
    wb = Workbook()
    ws = wb.active
    ws.title = "评审问题"

    columns = [
        ("问题ID", "issue_id", 12),
        ("问题类型", "type", 14),
        ("优先级", "priority", 10),
        ("位置", "location", 24),
        ("问题", "question", 42),
        ("原因", "reason", 42),
        ("影响", "impact", 42),
        ("建议", "suggestion", 36),
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="7A4E00")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (name, _, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, item in enumerate(issues, 2):
        if not isinstance(item, dict):
            item = {"question": str(item)}
        row = {
            "issue_id": f"ISSUE-{row_idx - 1:03d}",
            "type": item.get("type", ""),
            "priority": item.get("priority", ""),
            "location": item.get("location", ""),
            "question": item.get("question") or item.get("desc") or item.get("description") or item.get("title", ""),
            "reason": item.get("reason", ""),
            "impact": item.get("impact", ""),
            "suggestion": item.get("suggestion", ""),
        }
        for col_idx, (_, key, _) in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            cell.border = border
            cell.alignment = center if key in ("issue_id", "type", "priority") else left
        ws.row_dimensions[row_idx].height = 54

    ws.freeze_panes = "A2"
    wb.save(out_path)
    return True


def export_review_mindmap(review: dict, req_name: str, out_path: Path) -> bool:
    """导出评审 mindmap Markdown，可导入 XMind。"""
    mindmap = review.get("mindmap") if isinstance(review, dict) else None
    if isinstance(mindmap, dict):
        lines = _mindmap_to_markdown(mindmap, 1)
    else:
        lines = [f"# {req_name}", "## 需求评审", "### 未生成 mindmap"]
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return True


def export_markdown_xmind(testpoints: list, review: dict, req_name: str, out_path: Path) -> bool:
    """
    生成可导入 XMind 的 Markdown 文件。
    XMind 导入步骤: 文件 → 导入 → Markdown
    层级结构:
        # 根节点（需求名）
        ## 风险项
        ### 风险1
        ## 功能模块
        ### [来源][优先级] 测试点标题
        #### 前置条件 / 预期结果
    """
    lines = []
    # 先标准化所有测试点字段
    testpoints = [normalize_testpoint(tp, i) for i, tp in enumerate(testpoints)]

    lines.append(f"# {req_name}")
    lines.append(f"## 概览")
    lines.append(f"### 评审分: {review.get('score', 'N/A')}")
    lines.append(f"### 测试点总数: {len(testpoints)}")
    req_c  = sum(1 for t in testpoints if t.get("source") == "REQ")
    kb_c   = sum(1 for t in testpoints if t.get("source") == "KB")
    risk_c = sum(1 for t in testpoints if t.get("source") == "RISK")
    lines.append(f"### REQ需求直出: {req_c} | KB知识库: {kb_c} | RISK风险: {risk_c}")

    # 风险项
    risks = review.get("risk_flags", [])
    if risks:
        lines.append(f"## ⚠ 风险项 ({len(risks)}条)")
        for r in risks:
            lines.append(f"### [{r.get('type','?')}] {r.get('desc','')}")

    # 按功能模块分组（归一化模块名，避免同一模块因名称细微差异被分开）
    def normalize_module(name: str) -> str:
        """模块名归一化：去掉括号内容、版本号、多余空格"""
        import re as _re
        name = name.strip()
        name = _re.sub(r'[（(][^）)]*[）)]', '', name)  # 去掉括号
        name = _re.sub(r'V[\d.]+', '', name)             # 去掉版本号
        name = _re.sub(r'[\s_-]+', '', name)             # 去掉空格/下划线
        return name.strip() or "未分类"

    # 先按归一化名分组，保留原始名（取第一个出现的）
    modules: dict = {}          # 归一化名 → [tp, ...]
    mod_display: dict = {}      # 归一化名 → 显示名
    for tp in testpoints:
        raw_mod  = tp.get("functional_module") or tp.get("feature") or "未分类"
        norm_mod = normalize_module(raw_mod)
        if norm_mod not in mod_display:
            mod_display[norm_mod] = raw_mod
        modules.setdefault(norm_mod, []).append(tp)

    # 按 P0 数量降序排列模块（重要模块在前）
    modules = dict(sorted(
        modules.items(),
        key=lambda x: (
            -sum(1 for t in x[1] if t.get("priority") == "P0"),
            -len(x[1])
        )
    ))

    source_icon = {"REQ": "🔵", "KB": "🟡", "RISK": "🔴"}

    for norm_mod, tps in modules.items():
        mod_name = mod_display.get(norm_mod, norm_mod)
        mc = sum(1 for t in tps if t.get("source")=="REQ")
        kc = sum(1 for t in tps if t.get("source")=="KB")
        rc = sum(1 for t in tps if t.get("source")=="RISK")
        p0 = sum(1 for t in tps if t.get("priority")=="P0")
        lines.append(f"## {mod_name} ({len(tps)}条)")
        lines.append(f"### 统计: REQ={mc} KB={kc} RISK={rc} | P0={p0}")

        # 同模块内按来源排序：REQ → KB → RISK，相同来源内 P0 优先
        tps_sorted = sorted(tps, key=lambda t: (
            {"REQ": 0, "KB": 1, "RISK": 2}.get(t.get("source", "REQ"), 3),
            {"P0": 0, "P1": 1, "P2": 2}.get(t.get("priority", "P1"), 3)
        ))
        for tp in tps_sorted:
            src   = tp.get("source", "REQ")
            pri   = tp.get("priority", "P1")
            icon  = source_icon.get(src, "⚪")
            title = tp.get("test_scenario") or tp.get("title") or tp.get("case_title", "")
            lines.append(f"### {icon}[{src}][{pri}] {title}")

            # 子节点放关键信息
            if tp.get("preconditions"):
                lines.append(f"#### 前置: {tp['preconditions']}")
            if tp.get("expected_result"):
                lines.append(f"#### 预期: {tp['expected_result']}")
            if tp.get("source_ref"):
                lines.append(f"#### 来源: {tp['source_ref']}")
            if tp.get("remarks"):
                lines.append(f"#### 备注: {tp['remarks']}")

    try:
        out_path.write_text("\n".join(lines), encoding="utf-8")
        return True
    except Exception as e:
        print(f"  [warn] Markdown 生成失败: {e}")
        return False


# ── 统计辅助 ──────────────────────────────────────────────────────────────
def get_source(tp: dict) -> str:
    s = tp.get("source", "")
    if s in ("REQ", "KB", "RISK"):
        return s
    kb = tp.get("kb_source", "") + tp.get("source_ref", "")
    if "knowledge_base" in kb:
        return "KB"
    remarks = tp.get("remarks", "")
    if "风险" in remarks:
        return "RISK"
    return "REQ"


# ── 主流程 ─────────────────────────────────────────────────────────────────
def _extract_section(req_path: Path, keyword: str) -> str:
    """
    从需求文档中提取包含指定关键词的章节。
    支持 Markdown # ## ### 标题层级。
    """
    try:
        text = req_path.read_text(encoding="utf-8")
    except Exception:
        return ""

    # 过滤删除线内容（~~删除的需求~~）
    import re as _re_st
    text = _re_st.sub(r'~~.+?~~', '', text)

    lines = text.splitlines()
    result = []
    in_section = False
    section_level = 0

    for line in lines:
        # 检测标题行
        if line.startswith("#"):
            level = len(line) - len(line.lstrip("#"))
            title = line.lstrip("#").strip()

            if keyword in title:
                # 找到目标章节
                in_section    = True
                section_level = level
                result.append(line)
            elif in_section:
                # 遇到同级或更高级标题，章节结束
                if level <= section_level:
                    in_section = False
                else:
                    result.append(line)
        elif in_section:
            result.append(line)

    # 如果没匹配到标题，尝试去掉编号前缀后重试
    # 例如: "5.1现货持仓数量" → "现货持仓数量" 匹配 "## 现货持仓数量"
    if not result:
        import re as _re_num
        stripped = _re_num.sub(r'^[\d\.\s]+', '', keyword)
        if stripped and stripped != keyword:
            for line in lines:
                if line.startswith("#"):
                    level = len(line) - len(line.lstrip("#"))
                    title = line.lstrip("#").strip()
                    if stripped in title:
                        in_section    = True
                        section_level = level
                        result.append(line)
                    elif in_section:
                        if level <= section_level:
                            in_section = False
                        else:
                            result.append(line)
                elif in_section:
                    result.append(line)

    # 如果还没匹配到，尝试全文关键词段落匹配（分别用原关键词和去编号关键词）
    if not result:
        for kw in (keyword, stripped if stripped != keyword else keyword):
            paras = text.split("\n\n")
            for para in paras:
                if kw in para:
                    result.append(para)
            if result:
                break

    return "\n".join(result).strip()


def _normalize_section_text(text: str) -> str:
    """章节匹配归一化：去编号、空白和常见近义差异。"""
    import re as _re
    value = _re.sub(r'^[\d.\s、_-]+', '', text or "")
    value = _re.sub(r'\s+', '', value)
    replacements = {
        "退补款": "退款",
        "退补金": "退款",
        "退补": "退款",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    return value.lower()


def _list_section_titles(req_path: Path) -> list[str]:
    """列出 Markdown 标题，供章节找不到时提示候选。"""
    try:
        text = req_path.read_text(encoding="utf-8")
    except Exception:
        return []
    titles = []
    for line in text.splitlines():
        if line.startswith("#"):
            title = line.lstrip("#").strip()
            if title and not title.startswith("目录"):
                titles.append(title)
    return titles


def _suggest_sections(req_path: Path, keyword: str, limit: int = 5) -> list[str]:
    """返回相似章节候选，仅用于提示用户人工确认，不自动选择。"""
    from difflib import SequenceMatcher
    titles = _list_section_titles(req_path)
    norm_kw = _normalize_section_text(keyword)
    scored = []
    for title in titles:
        norm_title = _normalize_section_text(title)
        if not norm_title:
            continue
        score = SequenceMatcher(None, norm_kw, norm_title).ratio()
        if norm_kw and (norm_kw in norm_title or norm_title in norm_kw):
            score = max(score, 0.92)
        scored.append((score, title))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [title for score, title in scored[:limit] if score >= 0.35]


def _load_section_filter_keywords() -> list:
    """
    从 config/section_filter.json 加载章节过滤黑名单关键词。
    文件不存在或读取失败时返回内置默认列表。
    """
    import json as _json
    config_path = WORKDIR / "config" / "section_filter.json"
    try:
        if config_path.exists():
            data = _json.loads(config_path.read_text(encoding="utf-8"))
            keywords = data.get("non_core_keywords", [])
            if keywords:
                return keywords
    except Exception:
        pass
    # 内置默认（保持向后兼容）
    return [
        "文档范围", "适用范围", "参考文档", "参考资料", "相关文档",
        "修订记录", "版本历史", "变更历史", "文档说明", "目录",
        "监管条文", "监管解读", "法规", "条文解读",
        "客户规则", "规则示例", "示例", "案例",
        "背景", "概述", "简介", "介绍", "说明",
        "术语", "词汇", "缩略语", "定义",
    ]


def _split_sections(req_path: Path, min_lines: int = 5,
                    memory=None) -> list:
    """
    自动按 Markdown 二级标题(##)拆分需求文档为多个章节。
    三层过滤机制：
      第一层：黑名单关键词快速过滤
      第二层：长期记忆（历史学习结果）
      第三层：模型判断兜底（结果自动写回记忆）
    返回 [(章节标题, 章节内容)] 列表。
    """
    # ── 第一层：黑名单（从配置文件加载，文件不存在则用内置默认）──────────────
    NON_CORE_KW = _load_section_filter_keywords()

    def blacklist_hit(title: str) -> bool:
        clean = title.strip("*~_ ")
        return any(kw in clean for kw in NON_CORE_KW)

    # ── 解析所有章节 ──────────────────────────────────────────────────────────
    try:
        text = req_path.read_text(encoding="utf-8")
    except Exception:
        return []

    # 过滤删除线内容（~~废弃的需求~~）
    import re as _re_st2
    # 多行删除线也处理
    text = _re_st2.sub(r'~~.+?~~', '', text, flags=_re_st2.DOTALL)

    lines     = text.splitlines()
    raw_sections = []
    cur_title = None
    cur_lines = []

    for line in lines:
        if line.startswith("## ") and not line.startswith("### "):
            if cur_title and len(cur_lines) >= min_lines:
                raw_sections.append((cur_title, "\n".join(cur_lines)))
            raw_title = line.lstrip("# ").strip()
            # 标题本身是删除线（~~章节名~~）则跳过整个章节
            import re as _re_title
            if _re_title.fullmatch(r'~~.+~~', raw_title):
                print(f"  [删除线] 跳过已废弃章节: {raw_title}")
                cur_title = None
                cur_lines = []
            else:
                cur_title = raw_title
                cur_lines = [line]
        elif line.startswith("# ") and not line.startswith("## "):
            pass
        elif cur_title:
            cur_lines.append(line)
    if cur_title and len(cur_lines) >= min_lines:
        raw_sections.append((cur_title, "\n".join(cur_lines)))

    if not raw_sections:
        return []

    # ── 第一层：黑名单过滤 ────────────────────────────────────────────────────
    after_blacklist = []
    blacklist_skipped = []
    for title, content in raw_sections:
        if blacklist_hit(title):
            print(f"  [黑名单] 跳过: {title}")
            blacklist_skipped.append(title)
        else:
            after_blacklist.append((title, content))

    if not after_blacklist:
        return []

    # ── 第二层：长期记忆 ──────────────────────────────────────────────────────
    memory_skipped = []
    memory_kept    = []
    after_memory   = []

    if memory:
        patterns = memory.get_section_filter_patterns()
        mem_skip = set(patterns.get("skip", []))
        mem_keep = set(patterns.get("keep", []))

        for title, content in after_blacklist:
            clean = title.strip("*~_ ")
            if clean in mem_skip:
                print(f"  [记忆] 跳过（历史学习）: {title}")
                memory_skipped.append(title)
            elif clean in mem_keep:
                print(f"  [记忆] 保留（历史学习）: {title}")
                memory_kept.append(title)
                after_memory.append((title, content))
            else:
                after_memory.append((title, content))  # 未知，交给模型判断
    else:
        after_memory = after_blacklist

    # ── 第三层：模型判断（只对记忆未覆盖的章节）────────────────────────────────
    model_skipped = []
    final_sections = []

    unknown = [(t, c) for t, c in after_memory if t not in memory_kept]

    if unknown and len(unknown) > 0:
        titles_list = "\n".join(
            f"{i+1}. {t}" for i, (t, _) in enumerate(unknown)
        )
        print(f"  [模型判断] 对 {len(unknown)} 个未知章节进行判断...")
        try:
            resp = client.messages.create(
                model=MODEL,
                system="你是一名测试架构师，判断需求文档的章节是否属于核心需求内容。只输出 JSON，不要其他文字。",
                messages=[{"role": "user", "content": (
                    f"以下是需求文档的章节列表，请判断每个章节是否属于「核心需求内容」"
                    f"（即包含功能需求、计算逻辑、数据规则、枚举值定义等可以生成测试点的内容）：\n\n"
                    f"{titles_list}\n\n"
                    "输出 JSON，格式：\n"
                    '{"core": [1, 2, 5], "skip": [3, 4]}\n'
                    "（core 和 skip 各填章节序号列表）"
                )}],
                max_tokens=500,
            )
            result_text = "".join(b.text for b in resp.content if hasattr(b, "text"))
            import json as _json, re as _re
            result_text = _re.sub(r"```.*?```", "", result_text, flags=_re.DOTALL).strip()
            # 找到 JSON 对象起始位置
            json_start = result_text.find("{")
            if json_start >= 0:
                result_text = result_text[json_start:]
            try:
                judgment = _json.loads(result_text)
            except Exception:
                # 用正则兜底提取数字
                core_nums = [int(x) for x in _re.findall(r'"core"\s*:\s*\[([^\]]*)\]', result_text)]
                skip_nums = [int(x) for x in _re.findall(r'"skip"\s*:\s*\[([^\]]*)\]', result_text)]
                all_nums  = [int(x) for x in _re.findall(r'\d+', result_text)]
                judgment  = {"core": list(range(1, len(unknown)+1)), "skip": []}
            core_idxs = {i - 1 for i in judgment.get("core", [])}
            skip_idxs = {i - 1 for i in judgment.get("skip", [])}

            for i, (title, content) in enumerate(unknown):
                if i in skip_idxs:
                    print(f"  [模型] 跳过: {title}")
                    model_skipped.append(title)
                else:
                    final_sections.append((title, content))

        except Exception as e:
            print(f"  [warn] 模型判断失败: {e}，保留所有未知章节")
            final_sections.extend(unknown)
    else:
        final_sections.extend(unknown)

    # 加入记忆已确认保留的章节
    for title, content in after_memory:
        if title in memory_kept and (title, content) not in final_sections:
            final_sections.append((title, content))

    # ── 写回记忆（第三层学习结果）────────────────────────────────────────────
    if memory and model_skipped:
        all_skipped = blacklist_skipped + memory_skipped + model_skipped
        all_kept    = [t for t, _ in final_sections]
        memory.save_section_filter_result(
            skipped=[t for t in all_skipped if t not in blacklist_skipped],
            kept=all_kept
        )

    return final_sections


def main():
    parser = argparse.ArgumentParser(description="测试用例生成 Agent v2")
    parser.add_argument("requirement",    help="需求文档路径（.md / .txt）")
    parser.add_argument("--kb",           action="store_true", help="启用知识库检索")
    parser.add_argument("--skip-review",  action="store_true", help="跳过需求评审")
    parser.add_argument("--no-cases",     action="store_true", help="只生成测试点，不展开用例")
    parser.add_argument("--resume",       action="store_true", help="续跑：自动找最近未完成的任务")
    parser.add_argument("--section",      type=str, default="", help="只针对指定章节生成，如 --section 结算风控金")
    args = parser.parse_args()

    req_path = Path(args.requirement).resolve()
    if not req_path.exists():
        print(f"错误: 找不到需求文档 {req_path}")
        sys.exit(1)

    ts = int(time.time())
    log_path = setup_run_logging(req_path.stem, ts)
    print(f"  [日志] 本次执行日志: {log_path.relative_to(WORKDIR)}")

    # 自动处理 docx：转换为 md 并放入 knowledge_base/
    if req_path.suffix.lower() in (".docx", ".doc"):
        with timed_stage("input.docx_convert", "Word 转 Markdown"):
            md_path = KB_DIR / (req_path.stem + ".md")
            if not md_path.exists():
                print(f"  检测到 .docx，自动转换为 Markdown...")
                import subprocess as _sp
                # 优先用 docx2md.py（表格格式更干净）
                docx2md_script = SCRIPT_DIR / "docx2md.py"
                if docx2md_script.exists():
                    r = _sp.run(
                        [sys.executable, str(docx2md_script), str(req_path), "-o", str(md_path)],
                        capture_output=True, text=True, timeout=120
                    )
                    if r.returncode != 0 or not md_path.exists():
                        # 降级 pandoc
                        r = _sp.run(
                            ["pandoc", str(req_path), "-t", "markdown", "-o", str(md_path)],
                            capture_output=True, text=True
                        )
                else:
                    r = _sp.run(
                        ["pandoc", str(req_path), "-t", "markdown", "-o", str(md_path)],
                        capture_output=True, text=True
                    )
                if r.returncode != 0 or not md_path.exists():
                    print(f"  [错误] 转换失败: {r.stderr[:200]}")
                    sys.exit(1)
                print(f"  转换完成: {md_path.name}")
            else:
                print(f"  使用已有转换版本: {md_path.name}")
        req_path = md_path
    elif not req_path.is_relative_to(WORKDIR):
        # 不在 WORKDIR 内的文件复制到 knowledge_base/
        dst = KB_DIR / req_path.name
        if not dst.exists():
            import shutil as _shutil
            _shutil.copy2(req_path, dst)
            print(f"  文件已复制到 knowledge_base/{req_path.name}")
        req_path = dst

    stem     = req_path.stem
    req_name = req_path.name

    # 自动清理旧任务文件（只保留最近 30 个）
    tasks_dir = WORKDIR / ".tasks"
    if tasks_dir.exists():
        task_files = sorted(tasks_dir.glob("*.json"), key=lambda f: f.stat().st_mtime)
        to_delete  = task_files[:-30] if len(task_files) > 30 else []
        for f in to_delete:
            try: f.unlink()
            except Exception: pass
        if to_delete:
            print(f"  [清理] 删除 {len(to_delete)} 个旧任务文件")

    # 章节过滤：如果指定了 --section，提取对应章节内容写入临时文件
    section_keyword = args.section.strip()
    if section_keyword:
        print(f"  [章节过滤] 只处理包含「{section_keyword}」的章节")
        section_text = _extract_section(req_path, section_keyword)
        if section_text:
            # 写入临时文件，后续所有阶段读这个文件
            section_path = SECTIONS_DIR / f"_section_{stem}_{ts}.md"
            section_path.write_text(section_text, encoding="utf-8")
            print(f"  [章节过滤] 提取到 {len(section_text.splitlines())} 行内容，继续生成")
            req_path = section_path
            stem     = f"{stem}_{section_keyword}"
            req_name = section_path.name
        else:
            print(f"  [章节过滤] 未找到包含「{section_keyword}」的章节，已停止。")
            suggestions = _suggest_sections(req_path, section_keyword)
            if suggestions:
                print("  [章节过滤] 可能相关的章节如下，请人工确认后用准确章节名重新运行：")
                for i, title in enumerate(suggestions, 1):
                    print(f"    {i}. {title}")
                print("  示例：python scripts/agent.py \"需求文档.docx\" --kb --section \"ETF退款金\"")
            else:
                print("  [章节过滤] 未找到相似章节。请检查章节名称、空格、编号或转换后的 Markdown 标题。")
            return

    RUN_DIR  = get_run_dir(stem, ts)   # 本次运行的输出目录

    print(f"\n{'='*52}")
    print(f"  测试 Agent v2 启动")
    print(f"  需求文档: {req_name}")
    print(f"  知识库:   {'启用' if args.kb else '关闭'}")
    print(f"  生成用例: {'否' if args.no_cases else '是'}")
    print(f"{'='*52}\n")

    # s07: 初始化任务存储，支持续跑
    if args.resume:
        task = TaskStore.find_latest(stem)
        if task:
            print(f"  [s07] 续跑模式: {task.summary()}")
        else:
            print(f"  [s07] 未找到可续跑的任务，重新开始")
            task = TaskStore(stem, ts)
    else:
        task = TaskStore(stem, ts)
    print(f"  [s07] 任务文件: {task.path.name}")

    # s09: 初始化记忆系统
    memory     = MemoryStore(stem)
    memory_rag = MemoryRAG()   # 向量化长期记忆检索
    lt_counts  = {k: len(v) for k, v in memory._lt.items() if isinstance(v, list)}
    print(f"  [s09] 长期记忆: {lt_counts}（向量检索已就绪）\n")

    # ① 需求评审
    if task.is_done("review"):
        review = task.get_result("review")
        print(f"  [s07] 跳过需求评审（已完成）")
    elif args.skip_review:
        review = {"testable_features": [], "risk_flags": [], "score": 0}
        task.done("review", review)
        print("  [跳过] 需求评审")
    else:
        task.start("review")
        try:
            with timed_stage("stage.review", "阶段一 需求评审"):
                review = stage1_review(req_path, memory=memory)
            task.done("review", review)
            memory.save_after_review(review)   # s09: 保存评审经验
            memory_rag.invalidate()               # 触发记忆索引重建
        except Exception as e:
            task.fail("review", str(e))
            review = {"testable_features": [], "risk_flags": [], "score": 0}
            print(f"  [s11] 评审异常，使用空结果继续: {e}")
        score    = review.get("score", "N/A")
        features = review.get("testable_features", [])
        risks    = review.get("risk_flags", [])
        print(f"\n  评审完成 → 质量分: {score}, 功能点: {len(features)}, 风险项: {len(risks)}")
        if isinstance(score, int) and score < 60:
            print(f"  [警告] 需求质量较低（{score}/100），建议完善后再生成测试点")

    # ② 测试点生成
    if task.is_done("testpoints"):
        testpoints = task.get_result("testpoints")
        print(f"  [s07] 跳过测试点生成（已完成，共 {len(testpoints)} 条）")
    else:
        task.start("testpoints")
        try:
            with timed_stage("stage.testpoints", "阶段二 测试点生成"):
                # 无 --section 时，自动识别章节逐章处理后合并
                if not section_keyword:
                    sections = _split_sections(req_path, memory=memory)
                else:
                    sections = []

                if len(sections) >= 2:
                    print(f"  [自动章节] 识别到 {len(sections)} 个章节，逐章处理后合并")
                    for i, (title, _) in enumerate(sections):
                        print(f"    {i+1}. {title}")

                    all_tps = []
                    for i, (title, content) in enumerate(sections):
                        s_offset = len(all_tps)
                        print(f"\n  [章节 {i+1}/{len(sections)}] {title}")
                        s_path = OUTPUT_DIR / f"_sec_{int(time.time()*1000)%1000000}.md"
                        # 检查章节内容是否有实质内容（至少有10行非空行且含中文）
                        content_lines = [l for l in content.splitlines() if l.strip()]
                        cn_chars = sum(1 for c in content if '\u4e00' <= c <= '\u9fff')
                        print(f"    内容: {len(content_lines)}行, {cn_chars}个中文字符", flush=True)
                        # 跳过条件：非空行少于8行，或中文字符少于100个，或没有具体数字/字段名
                        has_specific_content = any(
                            any(kw in line for kw in ["字段", "表", "取值", "逻辑", "计算", "=", "：", "规则"])
                            for line in content_lines
                        )
                        if len(content_lines) < 8 or cn_chars < 100 or not has_specific_content:
                            print(f"    [跳过] 章节内容不足（行数{len(content_lines)}, 中文{cn_chars}, 有具体内容:{has_specific_content}）")
                            continue
                        s_path.write_text(f"# {title}\n\n{content}", encoding="utf-8")
                        try:
                            tps = stage2_testpoints(s_path, review, args.kb, memory=memory)
                            for j, tp in enumerate(tps):
                                tp["testpoint_id"] = f"TP-{s_offset + j + 1:03d}"
                                tp.setdefault("section", title)
                            all_tps.extend(tps)
                            print(f"    → {len(tps)} 条测试点")
                        finally:
                            s_path.unlink(missing_ok=True)

                    testpoints = all_tps
                    print(f"\n  [合并] 共 {len(testpoints)} 条测试点，来自 {len(sections)} 个章节")
                else:
                    testpoints = stage2_testpoints(req_path, review, args.kb, memory=memory)

            task.done("testpoints", testpoints)
            memory.save_after_testpoints(testpoints, review)  # s09: 保存测试点经验
            memory_rag.invalidate()                               # 触发记忆索引重建
        except Exception as e:
            task.fail("testpoints", str(e))
            testpoints = []
            print(f"  [s11] 测试点生成异常: {e}")

    # 标准化为扁平列表
    flat_tps = []
    if testpoints and isinstance(testpoints[0], dict):
        if "testpoints" in testpoints[0]:
            for module in testpoints:
                flat_tps.extend(module.get("testpoints", []))
        else:
            flat_tps = testpoints

    req_count  = sum(1 for t in flat_tps if get_source(t) == "REQ")
    kb_count   = sum(1 for t in flat_tps if get_source(t) == "KB")
    risk_count = sum(1 for t in flat_tps if get_source(t) == "RISK")

    # 用模型名+时间戳生成文件后缀，确保同模型多次运行也不重名
    _model_tag = re.sub(r"[^a-zA-Z0-9._-]", "-", MODEL.split("/")[-1])[:16].strip("-")
    _sfx = f"_{_model_tag}_{RUN_DIR.name}"

    # 保存需求评审 JSON，作为机器可读底层产物；同时导出可读文档。
    review_out = RUN_DIR / f"review{_sfx}.json"
    review_out.write_text(json.dumps({
        "meta": {
            "requirement": str(req_path),
            "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "model": MODEL,
            "section": section_keyword,
        },
        "review": review,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    review_md_out = RUN_DIR / f"review_report{_sfx}.md"
    review_mindmap_out = RUN_DIR / f"review_mindmap{_sfx}.md"
    review_issues_out = RUN_DIR / f"review_issues{_sfx}.xlsx"

    review_md_ok = export_review_markdown(review, req_name, review_md_out)
    review_mindmap_ok = export_review_mindmap(review, req_name, review_mindmap_out)
    try:
        review_issues_ok = export_review_issues_excel(review, review_issues_out)
    except Exception as e:
        print(f"  [warn] 评审问题 Excel 生成失败: {e}")
        review_issues_ok = False

    # 保存测试点 JSON
    tp_out = RUN_DIR / f"testpoints{_sfx}.json"
    tp_out.write_text(json.dumps({
        "meta": {
            "requirement": str(req_path),
            "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "total": len(flat_tps),
            "by_source": {"REQ": req_count, "KB": kb_count, "RISK": risk_count},
        },
        "review": review,
        "testpoints": flat_tps,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    # 生成 Markdown（供导入 XMind）
    md_out   = RUN_DIR / f"testpoints_xmind{_sfx}.md"
    xmind_ok = export_markdown_xmind(flat_tps, review, req_name, md_out)

    print(f"\n{'='*52}")
    print(f"  ② 测试点生成完成")
    print(f"     总数: {len(flat_tps)}  🔵REQ={req_count}  🟡KB={kb_count}  🔴RISK={risk_count}")
    print(f"  输出目录: {RUN_DIR.relative_to(WORKDIR)}")
    if review_md_ok:
        print(f"     评审报告: {review_md_out.name}")
    if review_issues_ok:
        print(f"     问题清单: {review_issues_out.name}")
    if review_mindmap_ok:
        print(f"     评审思维导图: {review_mindmap_out.name}")
    print(f"     测试点JSON: {tp_out.name}")
    if xmind_ok:
        print(f"     测试点思维导图: {md_out.name}")
    print(f"{'='*52}")

    manifest_files = {
        "run_log": RUN_LOG_PATH,
        "review_json": review_out,
        "review_report_md": review_md_out if review_md_ok else "",
        "review_issues_xlsx": review_issues_out if review_issues_ok else "",
        "review_mindmap_md": review_mindmap_out if review_mindmap_ok else "",
        "testpoints_json": tp_out,
        "testpoints_xmind_md": md_out if xmind_ok else "",
    }
    write_manifest(
        RUN_DIR,
        status="testpoints_done",
        req_path=req_path,
        model=MODEL,
        use_kb=args.kb,
        no_cases=args.no_cases,
        section=section_keyword,
        review=review,
        testpoints=flat_tps,
        files=manifest_files,
        task_summary=task.summary(),
    )

    if args.no_cases:
        write_manifest(
            RUN_DIR,
            status="done",
            req_path=req_path,
            model=MODEL,
            use_kb=args.kb,
            no_cases=args.no_cases,
            section=section_keyword,
            review=review,
            testpoints=flat_tps,
            files=manifest_files,
            task_summary=task.summary(),
        )
        return

    # ③ 测试用例生成（测试点为空则跳过）
    if not flat_tps:
        print("\n  [跳过] 测试点为空，跳过用例生成")
        write_manifest(
            RUN_DIR,
            status="no_testpoints",
            req_path=req_path,
            model=MODEL,
            use_kb=args.kb,
            no_cases=args.no_cases,
            section=section_keyword,
            review=review,
            testpoints=flat_tps,
            files=manifest_files,
            task_summary=task.summary(),
            warnings=["测试点为空，跳过用例生成"],
        )
        return

    if task.is_done("testcases"):
        testcases = task.get_result("testcases")
        print(f"  [s07] 跳过用例生成（已完成，共 {len(testcases)} 条）")
    else:
        task.start("testcases")
        try:
            with timed_stage("stage.testcases", "阶段三 测试用例生成"):
                testcases = stage3_testcases(flat_tps, req_path, memory=memory)
            task.done("testcases", testcases)
        except Exception as e:
            task.fail("testcases", str(e))
            testcases = []
            print(f"  [s11] 用例生成异常: {e}")

    with timed_stage("stage.export", "阶段四 本地导出"):
        tc_out   = RUN_DIR / f"testcases{_sfx}.json"
        xlsx_out = RUN_DIR / f"testcases{_sfx}.xlsx"

        tc_out.write_text(json.dumps(testcases, ensure_ascii=False, indent=2), encoding="utf-8")
        try:
            xlsx_ok = export_excel(testcases, xlsx_out)
        except Exception as e:
            print(f"  [warn] Excel 生成失败: {e}")
            import traceback; traceback.print_exc()
            xlsx_ok = False

        # ④ 测分文档生成（本地，零 token）
        from gen_report import generate_report
        report_out = RUN_DIR / f"report{_sfx}.md"
        try:
            report_md = generate_report(
                {"meta": {"requirement": str(req_path), "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                          "total": len(flat_tps),
                          "by_source": {"REQ": req_count, "KB": kb_count, "RISK": risk_count}},
                 "review": review, "testpoints": flat_tps},
                testcases,
                report_out
            )
            report_out.write_text(report_md, encoding="utf-8")
            report_ok = True
        except Exception as e:
            print(f"  [warn] 测分文档生成失败: {e}")
            report_ok = False

    task.done("export", {"testcases": str(tc_out.name), "excel": str(xlsx_out.name),
                         "report": str(report_out.name) if report_ok else ""})

    manifest_files.update({
        "testcases_json": tc_out,
        "testcases_xlsx": xlsx_out if xlsx_ok else "",
        "report_md": report_out if report_ok else "",
    })
    manifest_path = write_manifest(
        RUN_DIR,
        status="done",
        req_path=req_path,
        model=MODEL,
        use_kb=args.kb,
        no_cases=args.no_cases,
        section=section_keyword,
        review=review,
        testpoints=flat_tps,
        testcases=testcases,
        files=manifest_files,
        task_summary=task.summary(),
        warnings=[
            msg for msg, ok in (
                ("Excel 生成失败", xlsx_ok),
                ("测分文档生成失败", report_ok),
            )
            if not ok
        ],
    )

    print(f"\n{'='*52}")
    print(f"  ③ 测试用例生成完成")
    print(f"     总数: {len(testcases)} 条")
    print(f"     JSON:  {tc_out.name}")
    if xlsx_ok:
        print(f"     Excel: {xlsx_out.name}")
    if report_ok:
        print(f"     测分:  {report_out.name}")
    print(f"     Manifest: {manifest_path.name}")
    print(f"\n  [s07] 最终任务状态: {task.summary()}")
    print(f"{'='*52}\n")


if __name__ == "__main__":
    main()
