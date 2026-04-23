#!/usr/bin/env python3
"""
regen_excel.py - 从已有测试用例 JSON 重新生成 Excel 文件

纯本地脚本，不调用任何 API，不消耗 token。
适合：调整 Excel 格式、列顺序、颜色、新增列时直接重跑。

用法:
    python regen_excel.py output/testcases_xxx.json
    python regen_excel.py output/testcases_xxx.json --out output/新格式.xlsx
    python regen_excel.py output/testpoints_xxx.json --from-testpoints  # 直接从测试点生成简版 Excel
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要 openpyxl，请先安装: pip install openpyxl")
    sys.exit(1)


# ── 颜色配置（在这里修改颜色，不需要动主逻辑）──────────────────────────────
SOURCE_COLORS = {
    "REQ":  "DDEEFF",   # 蓝色底 — 需求直出
    "KB":   "FFF9DD",   # 黄色底 — 知识库补充
    "RISK": "FFE8E8",   # 红色底 — 风险推断
}
PRIORITY_COLORS = {
    "P0": "CC0000",   # 深红
    "P1": "E07000",   # 橙色
    "P2": "888888",   # 灰色
}
HEADER_FILL  = "2B5FA8"   # 表头深蓝
HEADER_COLOR = "FFFFFF"   # 表头文字白色
ROW_HEIGHT   = 48         # 数据行高（像素）
HEADER_HEIGHT = 24        # 表头行高


# ── 列定义（调整列顺序/宽度只需改这里）─────────────────────────────────────
TESTCASE_COLUMNS = [
    # (列名,          字段key,          列宽,   对齐)
    ("用例ID",        "case_id",         10,    "center"),
    ("测试点ID",      "testpoint_id",    10,    "center"),
    ("功能模块",      "functional_module",16,   "left"),
    ("用例标题",      "case_title",      32,    "left"),
    ("来源",          "source",           7,    "center"),
    ("优先级",        "priority",         7,    "center"),
    ("前置条件",      "preconditions",   22,    "left"),
    ("测试数据",      "test_data",       18,    "left"),
    ("操作步骤",      "steps",           38,    "left"),
    ("预期结果",      "expected_result", 32,    "left"),
    ("实际结果",      "actual_result",   22,    "left"),
    ("执行状态",      "status",           9,    "center"),
    ("备注",          "remarks",         18,    "left"),
]

TESTPOINT_COLUMNS = [
    ("测试点ID",      "testpoint_id",    10,    "center"),
    ("功能模块",      "functional_module",18,   "left"),
    ("测试场景",      "test_scenario",   35,    "left"),
    ("来源",          "source",           7,    "center"),
    ("优先级",        "priority",         7,    "center"),
    ("前置条件",      "preconditions",   22,    "left"),
    ("预期结果",      "expected_result", 32,    "left"),
    ("来源引用",      "source_ref",      22,    "left"),
    ("备注",          "remarks",         15,    "left"),
]


# ── 字段标准化（兼容各种字段名）─────────────────────────────────────────────
def normalize_testcase(case: dict, idx: int) -> dict:
    n = dict(case)
    alias = {
        "case_title":        ["case_name", "title", "test_name", "name"],
        "functional_module": ["module", "test_module", "feature", "功能模块"],
        "preconditions":     ["precondition", "pre_condition", "prerequisite"],
        "test_data":         ["input_data", "test_input", "data"],
        "steps":             ["step", "test_steps", "procedure", "操作步骤"],
        "expected_result":   ["expected", "result", "expect", "expected_output"],
        "source":            ["test_source", "来源"],
        "priority":          ["level", "test_priority", "优先级"],
        "remarks":           ["remark", "note", "comment"],
    }
    for std_key, aliases in alias.items():
        if not n.get(std_key):
            for a in aliases:
                if n.get(a):
                    n[std_key] = n[a]
                    break
    defaults = {
        "case_id":           f"TC-{idx:03d}",
        "testpoint_id":      "",
        "functional_module": "",
        "case_title":        "",
        "source":            "REQ",
        "priority":          "P1",
        "preconditions":     "",
        "test_data":         "",
        "steps":             "",
        "expected_result":   "",
        "actual_result":     "",
        "status":            "",
        "remarks":           "",
    }
    for k, v in defaults.items():
        if k not in n or n[k] is None:
            n[k] = v
    # source 合法性
    if n["source"] not in ("REQ", "KB", "RISK"):
        n["source"] = "REQ"
    # priority 合法性
    if n["priority"] not in ("P0", "P1", "P2"):
        n["priority"] = "P1"
    # steps 列表转字符串
    if isinstance(n.get("steps"), list):
        n["steps"] = "\n".join(f"{i+1}. {s}" for i, s in enumerate(n["steps"]))
    return n


def normalize_testpoint(tp: dict, idx: int) -> dict:
    n = dict(tp)
    if not n.get("testpoint_id"):
        n["testpoint_id"] = n.get("id") or n.get("tp_id") or f"TP-{idx+1:03d}"
    if not n.get("test_scenario"):
        n["test_scenario"] = n.get("title") or n.get("name") or n.get("case_title") or ""
    raw_pri = n.get("priority") or n.get("level") or "P1"
    n["priority"] = raw_pri if raw_pri in ("P0", "P1", "P2") else "P1"
    if not n.get("functional_module"):
        mod = n.get("module") or n.get("feature") or ""
        if not mod:
            title = n.get("test_scenario", "")
            mod = title.split("-")[0].strip() if "-" in title else "功能验证"
        n["functional_module"] = mod
    src = n.get("source") or "REQ"
    n["source"] = src if src in ("REQ", "KB", "RISK") else "REQ"
    if not n.get("expected_result"):
        n["expected_result"] = n.get("expected") or n.get("expect") or ""
    if not n.get("preconditions"):
        n["preconditions"] = n.get("precondition") or n.get("desc") or ""
    if not n.get("source_ref"):
        n["source_ref"] = n.get("source_reference") or n.get("ref") or ""
    if not n.get("remarks"):
        n["remarks"] = n.get("remark") or n.get("note") or ""
    return n


# ── Excel 生成核心 ─────────────────────────────────────────────────────────
def build_excel(rows: list, columns: list, title: str, out_path: Path,
                legend_rows: list = None) -> bool:
    wb  = Workbook()
    ws  = wb.active
    ws.title = title

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 表头
    for col_idx, (col_name, _, col_width, align) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color=HEADER_COLOR, size=11)
        cell.fill      = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[1].height = HEADER_HEIGHT

    # 数据行
    for row_idx, row_data in enumerate(rows, 2):
        source   = row_data.get("source", "REQ")
        row_fill = PatternFill("solid", fgColor=SOURCE_COLORS.get(source, "FFFFFF"))

        for col_idx, (_, field_key, _, align) in enumerate(columns, 1):
            value = row_data.get(field_key, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = border
            cell.fill      = row_fill
            cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)

            # 优先级特殊颜色
            if field_key == "priority":
                cell.font = Font(color=PRIORITY_COLORS.get(value, "000000"), bold=True, size=11)
            else:
                cell.font = Font(size=11)

        ws.row_dimensions[row_idx].height = ROW_HEIGHT

    # 冻结首行
    ws.freeze_panes = "A2"

    # 图例 sheet
    legend_ws = wb.create_sheet("图例说明")
    legend_data = [
        ("颜色",   "含义"),
        ("蓝色底", "REQ — 来自需求文档"),
        ("黄色底", "KB  — 来自知识库补充"),
        ("红色底", "RISK — 风险推断"),
        ("",       ""),
        ("优先级", "说明"),
        ("P0",     "核心必测，上线前必须通过"),
        ("P1",     "重要应测，正式测试必须覆盖"),
        ("P2",     "边缘可测，时间充裕时覆盖"),
    ]
    if legend_rows:
        legend_data += [("", "")] + legend_rows
    for r, (a, b) in enumerate(legend_data, 1):
        legend_ws.cell(row=r, column=1, value=a)
        legend_ws.cell(row=r, column=2, value=b)
    legend_ws.column_dimensions["A"].width = 12
    legend_ws.column_dimensions["B"].width = 30

    wb.save(out_path)
    return True


# ── 统计辅助 ───────────────────────────────────────────────────────────────
def print_stats(rows: list, label: str):
    req_c  = sum(1 for r in rows if r.get("source") == "REQ")
    kb_c   = sum(1 for r in rows if r.get("source") == "KB")
    risk_c = sum(1 for r in rows if r.get("source") == "RISK")
    p0_c   = sum(1 for r in rows if r.get("priority") == "P0")
    p1_c   = sum(1 for r in rows if r.get("priority") == "P1")
    p2_c   = sum(1 for r in rows if r.get("priority") == "P2")
    print(f"  {label}总数: {len(rows)}")
    print(f"  来源: 🔵REQ={req_c}  🟡KB={kb_c}  🔴RISK={risk_c}")
    print(f"  优先级: P0={p0_c}  P1={p1_c}  P2={p2_c}")


# ── 主函数 ─────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="从测试 JSON 重新生成 Excel（不消耗 token）")
    parser.add_argument("json_file", help="测试用例或测试点 JSON 文件路径")
    parser.add_argument("--out", help="输出 Excel 路径（默认同目录同名 .xlsx）")
    parser.add_argument("--from-testpoints", action="store_true",
                        help="输入文件是测试点 JSON（而非测试用例）")
    args = parser.parse_args()

    json_path = Path(args.json_file).expanduser()
    if not json_path.exists():
        print(f"错误: 找不到文件 {json_path}")
        sys.exit(1)

    data = json.loads(json_path.read_text(encoding="utf-8"))

    if args.from_testpoints:
        # 从测试点 JSON 生成简版 Excel
        raw_tps  = data.get("testpoints", [])
        review   = data.get("review", {})
        req_name = Path(data.get("meta", {}).get("requirement", "unknown")).name
        rows     = [normalize_testpoint(tp, i) for i, tp in enumerate(raw_tps)]
        columns  = TESTPOINT_COLUMNS
        title    = "测试点"
        legend   = [("评审分", str(review.get("score", "N/A"))),
                    ("需求文档", req_name)]
        print(f"\n需求文档: {req_name}")
        print(f"评审分: {review.get('score', 'N/A')}")
        print_stats(rows, "测试点")
    else:
        # 从测试用例 JSON 生成完整 Excel
        raw_cases = data if isinstance(data, list) else data.get("testcases", data)
        if isinstance(raw_cases, dict):
            # 可能是包裹在 meta/testcases 结构里
            raw_cases = raw_cases.get("testcases", [])
        rows    = [normalize_testcase(c, i+1) for i, c in enumerate(raw_cases)]
        columns = TESTCASE_COLUMNS
        title   = "测试用例"
        legend  = None
        print(f"\n测试用例文件: {json_path.name}")
        print_stats(rows, "测试用例")

    if not rows:
        print("错误: 数据为空，请检查 JSON 文件内容")
        sys.exit(1)

    # 输出路径
    out_path = Path(args.out).expanduser() if args.out else json_path.with_suffix(".xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"\n生成 Excel: {out_path.name} ...", end=" ", flush=True)
    ok = build_excel(rows, columns, title, out_path, legend)

    if ok:
        size_kb = out_path.stat().st_size // 1024
        print(f"✓  ({size_kb} KB)")
        print(f"\n输出文件: {out_path}")
    else:
        print("✗ 生成失败")
        sys.exit(1)


if __name__ == "__main__":
    main()
