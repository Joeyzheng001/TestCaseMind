#!/usr/bin/env python3
"""
review_cases.py - 测试用例评审与优化

输入人工编写的测试用例（XMind MD 格式）和需求文档，
自动完成两项工作：
1. 评审报告：覆盖率分析 + 用例质量评分
2. 优化版用例：补充遗漏场景，修正质量问题，生成新的 MD + Excel

用法:
    python review_cases.py 需求文档.md 人工用例.md
    python review_cases.py 需求文档.md 人工用例.md --out output/review/
    python review_cases.py 需求文档.md 人工用例.md --no-optimize  # 只出报告不优化

输出:
    output/review_<需求名>_<时间戳>/
        ├── review_report.md      ← 评审报告（覆盖率+质量评分）
        ├── optimized_cases.md    ← 优化版用例（XMind 格式）
        └── optimized_cases.xlsx  ← 优化版用例（Excel）
"""

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

from dotenv import load_dotenv

AGENT_DIR = Path(__file__).parent
load_dotenv(AGENT_DIR / ".env", override=True)

try:
    import anthropic
except ImportError:
    print("错误: 请先安装 pip install anthropic")
    sys.exit(1)

client = anthropic.Anthropic()
MODEL  = os.environ.get("MODEL_ID", "claude-sonnet-4-6")

OUTPUT_DIR = AGENT_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


# ── 工具函数 ──────────────────────────────────────────────────────────────────
def read_file(path: Path, limit: int = 0) -> str:
    text = path.read_text(encoding="utf-8")
    if limit and len(text) > limit:
        text = text[:limit] + f"\n...(内容已截断，共 {len(text)} 字符)"
    return text


def call_llm(system: str, prompt: str, max_tokens: int = 4000) -> str:
    response = client.messages.create(
        model=MODEL,
        system=system,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=max_tokens,
    )
    return "".join(b.text for b in response.content if hasattr(b, "text"))


def extract_json(text: str) -> dict | list:
    """从文本中提取 JSON。"""
    text = text.strip()
    # 去掉代码块标记
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    try:
        return json.loads(text)
    except Exception:
        # 尝试找到 JSON 起始位置
        for start_char in ["{", "["]:
            idx = text.find(start_char)
            if idx >= 0:
                try:
                    return json.loads(text[idx:])
                except Exception:
                    pass
    return {}


def parse_xmind_md(md_text: str) -> list:
    """
    解析 XMind 导入格式的 Markdown，提取用例列表。
    格式示例：
      ## 模块名
      ### 用例标题
      #### 步骤/预期
    """
    cases = []
    current_module = ""
    current_case = None

    for line in md_text.splitlines():
        line = line.strip()
        if not line:
            continue

        if line.startswith("## ") and not line.startswith("### "):
            current_module = line[3:].strip()
        elif line.startswith("### "):
            if current_case:
                cases.append(current_case)
            title = line[4:].strip()
            # 去掉优先级标记如 [P0] [REQ]
            title_clean = re.sub(r"\[.*?\]", "", title).strip()
            current_case = {
                "module":   current_module,
                "title":    title_clean,
                "raw":      title,
                "steps":    "",
                "expected": "",
                "priority": _extract_priority(title),
            }
        elif line.startswith("#### ") and current_case:
            content = line[5:].strip()
            if any(k in content for k in ["步骤", "操作", "执行"]):
                current_case["steps"] = content
            elif any(k in content for k in ["预期", "期望", "结果"]):
                current_case["expected"] = content
            else:
                # 追加到步骤
                if current_case["steps"]:
                    current_case["steps"] += "\n" + content
                else:
                    current_case["steps"] = content

    if current_case:
        cases.append(current_case)

    return cases


def _extract_priority(text: str) -> str:
    m = re.search(r"\[P([012])\]", text, re.IGNORECASE)
    return f"P{m.group(1)}" if m else "P1"


# ── 阶段一：覆盖率分析 ────────────────────────────────────────────────────────
def analyze_coverage(req_text: str, cases: list) -> dict:
    """
    分析人工用例对需求的覆盖率：
    - 哪些需求功能点已覆盖
    - 哪些场景被遗漏（边界/异常/枚举值）
    - KB/RISK 类场景是否覆盖
    """
    cases_summary = "\n".join(
        f"- [{c['priority']}][{c['module']}] {c['title']}"
        for c in cases
    )

    system = (
        "你是一名资深测试架构师，专门评审测试用例的覆盖完整性。"
        "只输出 JSON，不要有其他文字。"
    )
    prompt = f"""
需求文档内容：
{req_text[:4000]}

---
人工编写的测试用例（共 {len(cases)} 条）：
{cases_summary}

---
请分析测试用例对需求的覆盖情况，输出以下 JSON：
{{
  "covered_features": ["已覆盖的功能点列表"],
  "missing_scenarios": [
    {{
      "type": "boundary|enum|exception|risk|integration",
      "description": "遗漏场景的具体描述",
      "suggestion": "建议补充的测试用例标题"
    }}
  ],
  "coverage_score": 75,
  "coverage_comment": "覆盖率总体评价",
  "has_boundary_tests": true,
  "has_exception_tests": true,
  "has_enum_tests": false,
  "has_risk_tests": false
}}
"""
    result = call_llm(system, prompt)
    data = extract_json(result)
    return data if isinstance(data, dict) else {}


# ── 阶段二：用例质量评分 ──────────────────────────────────────────────────────
def evaluate_quality(cases: list) -> dict:
    """
    评估每条用例的质量：
    - 步骤是否清晰可执行
    - 预期结果是否明确
    - 测试数据是否具体
    - 标题是否准确描述场景
    """
    # 抽样评估（最多20条，避免 token 过多）
    sample = cases[:20]
    cases_json = json.dumps(
        [{"title": c["title"], "module": c["module"],
          "steps": c["steps"][:200], "expected": c["expected"][:200]}
         for c in sample],
        ensure_ascii=False, indent=2
    )

    system = (
        "你是一名资深测试工程师，专门评估测试用例的编写质量。"
        "只输出 JSON，不要有其他文字。"
    )
    prompt = f"""
以下是 {len(sample)} 条测试用例（共 {len(cases)} 条，此处抽样）：
{cases_json}

请从以下4个维度评分（每项0-25分），并给出具体问题：
{{
  "scores": {{
    "step_clarity": 20,
    "expected_clarity": 18,
    "data_specificity": 15,
    "title_accuracy": 22
  }},
  "total_score": 75,
  "grade": "良",
  "common_issues": [
    "步骤描述过于笼统，缺少具体数值",
    "预期结果只写了'成功'，未说明具体返回值"
  ],
  "good_practices": [
    "优先级标注清晰",
    "模块分类合理"
  ],
  "improvement_suggestions": [
    "在步骤中补充具体的测试数据（如输入值=100）",
    "预期结果中说明具体的返回值或状态码"
  ]
}}
"""
    result = call_llm(system, prompt)
    data = extract_json(result)
    return data if isinstance(data, dict) else {}


# ── 阶段三：生成优化版用例 ───────────────────────────────────────────────────
def generate_optimized_cases(req_text: str, cases: list,
                             coverage: dict, quality: dict) -> list:
    """
    基于评审结果，生成优化版用例：
    1. 修正原有用例的质量问题
    2. 补充遗漏的测试场景
    """
    missing = coverage.get("missing_scenarios", [])
    issues  = quality.get("common_issues", [])
    suggestions = quality.get("improvement_suggestions", [])

    missing_desc = "\n".join(
        f"- [{m['type']}] {m['description']} → 建议补充：{m['suggestion']}"
        for m in missing[:10]
    )
    issues_desc = "\n".join(f"- {i}" for i in issues)
    suggest_desc = "\n".join(f"- {s}" for s in suggestions)

    # 原有用例摘要
    existing_titles = "\n".join(
        f"- [{c['priority']}][{c['module']}] {c['title']}"
        for c in cases
    )

    system = (
        "你是一名资深测试工程师，专门优化测试用例。"
        "只输出 JSON 数组，不要有其他文字。"
    )
    prompt = f"""
需求文档（节选）：
{req_text[:2000]}

---
现有用例（{len(cases)} 条）：
{existing_titles}

---
评审发现的质量问题：
{issues_desc}

改进建议：
{suggest_desc}

遗漏的测试场景：
{missing_desc}

---
请输出优化后的完整用例列表 JSON 数组，包含：
1. 改进后的原有用例（修正质量问题）
2. 新增的遗漏场景用例

每条用例格式：
[
  {{
    "case_id": "TC-001",
    "module": "功能模块名",
    "title": "用例标题（清晰描述场景）",
    "priority": "P0",
    "is_new": false,
    "optimization_note": "改动说明（新增用例填写补充原因）",
    "preconditions": "前置条件（具体）",
    "steps": "1. 步骤一（含具体数值）\\n2. 步骤二",
    "expected": "预期结果（明确的返回值或状态）",
    "test_data": "具体测试数据"
  }}
]
"""
    result = call_llm(system, prompt, max_tokens=6000)
    data = extract_json(result)
    return data if isinstance(data, list) else []


# ── 报告生成 ──────────────────────────────────────────────────────────────────
def generate_review_report(req_name: str, cases: list,
                           coverage: dict, quality: dict,
                           optimized: list) -> str:
    """生成完整的评审报告 Markdown。"""
    ts       = time.strftime("%Y-%m-%d %H:%M:%S")
    cov_score = coverage.get("coverage_score", 0)
    qual_score = quality.get("total_score", 0)
    overall  = (cov_score + qual_score) / 2

    if overall >= 85:   grade = "🟢 优"
    elif overall >= 70: grade = "🟡 良"
    elif overall >= 60: grade = "🟠 中"
    else:               grade = "🔴 差"

    new_count = sum(1 for c in optimized if c.get("is_new"))
    opt_count = len(optimized) - new_count

    lines = [
        "# 测试用例评审报告",
        "",
        f"| 项目 | 内容 |",
        f"|------|------|",
        f"| 需求文档 | {req_name} |",
        f"| 评审时间 | {ts} |",
        f"| 原有用例 | {len(cases)} 条 |",
        f"| 综合评级 | {grade}（{overall:.0f}/100）|",
        "",
        "---",
        "",
        "## 一、覆盖率分析",
        "",
        f"**覆盖率得分：{cov_score}/100**",
        "",
        f"> {coverage.get('coverage_comment', '')}",
        "",
        "### 1.1 已覆盖功能点",
        "",
    ]

    for f in coverage.get("covered_features", []):
        lines.append(f"- ✅ {f}")
    lines.append("")

    # 覆盖维度
    lines += [
        "### 1.2 覆盖维度检查",
        "",
        f"| 维度 | 状态 |",
        f"|------|------|",
        f"| 边界值测试 | {'✅ 有' if coverage.get('has_boundary_tests') else '❌ 缺失'} |",
        f"| 异常流测试 | {'✅ 有' if coverage.get('has_exception_tests') else '❌ 缺失'} |",
        f"| 枚举值测试 | {'✅ 有' if coverage.get('has_enum_tests') else '❌ 缺失'} |",
        f"| 风险场景测试 | {'✅ 有' if coverage.get('has_risk_tests') else '❌ 缺失'} |",
        "",
    ]

    missing = coverage.get("missing_scenarios", [])
    if missing:
        lines += [
            "### 1.3 遗漏场景",
            "",
            f"共发现 **{len(missing)}** 个遗漏场景：",
            "",
        ]
        type_map = {
            "boundary":    "⚠ 边界值",
            "enum":        "⚠ 枚举值",
            "exception":   "⚠ 异常流",
            "risk":        "🔴 风险",
            "integration": "🔗 集成",
        }
        for m in missing:
            t = type_map.get(m.get("type", ""), "⚠ 其他")
            lines.append(f"- **{t}**：{m.get('description', '')}")
            lines.append(f"  → 建议补充：*{m.get('suggestion', '')}*")
        lines.append("")

    # 质量评分
    scores = quality.get("scores", {})
    lines += [
        "## 二、用例质量评估",
        "",
        f"**质量得分：{qual_score}/100（{quality.get('grade', '')}）**",
        "",
        "### 2.1 分项评分",
        "",
        "| 维度 | 得分 | 满分 |",
        "|------|------|------|",
        f"| 步骤清晰度 | {scores.get('step_clarity', 0)} | 25 |",
        f"| 预期结果明确性 | {scores.get('expected_clarity', 0)} | 25 |",
        f"| 测试数据具体性 | {scores.get('data_specificity', 0)} | 25 |",
        f"| 用例标题准确性 | {scores.get('title_accuracy', 0)} | 25 |",
        "",
    ]

    good = quality.get("good_practices", [])
    if good:
        lines += ["### 2.2 做得好的地方", ""]
        for g in good:
            lines.append(f"- ✅ {g}")
        lines.append("")

    issues = quality.get("common_issues", [])
    if issues:
        lines += ["### 2.3 常见问题", ""]
        for i in issues:
            lines.append(f"- ❌ {i}")
        lines.append("")

    suggestions = quality.get("improvement_suggestions", [])
    if suggestions:
        lines += ["### 2.4 改进建议", ""]
        for s in suggestions:
            lines.append(f"- 💡 {s}")
        lines.append("")

    # 优化结果
    lines += [
        "## 三、优化结果",
        "",
        f"| 项目 | 数量 |",
        f"|------|------|",
        f"| 优化后总用例数 | {len(optimized)} 条 |",
        f"| 改进原有用例 | {opt_count} 条 |",
        f"| 新增补充用例 | {new_count} 条 |",
        "",
        "优化版用例已输出为：",
        "- `optimized_cases.md` — 可导入 XMind",
        "- `optimized_cases.xlsx` — Excel 格式",
        "",
        "---",
        "",
        f"*本报告由 TestCaseMind 自动生成  ·  {ts}*",
    ]

    return "\n".join(lines)


def cases_to_xmind_md(cases: list, req_name: str) -> str:
    """把优化版用例转为 XMind MD 格式。"""
    lines = [f"# {req_name}（优化版）", ""]
    modules: dict = {}
    for c in cases:
        mod = c.get("module", "未分类")
        modules.setdefault(mod, []).append(c)

    for mod, mod_cases in modules.items():
        p0 = sum(1 for c in mod_cases if c.get("priority") == "P0")
        p1 = sum(1 for c in mod_cases if c.get("priority") == "P1")
        p2 = sum(1 for c in mod_cases if c.get("priority") == "P2")
        new = sum(1 for c in mod_cases if c.get("is_new"))
        lines.append(f"## {mod} ({len(mod_cases)}条{f', 含{new}条新增' if new else ''})")
        lines.append(f"### 统计: P0={p0} P1={p1} P2={p2}")
        for c in mod_cases:
            tag = "🆕" if c.get("is_new") else "✏"
            pri = c.get("priority", "P1")
            title = c.get("title", "")
            lines.append(f"### {tag}[{pri}] {title}")
            if c.get("optimization_note"):
                lines.append(f"#### 优化说明: {c['optimization_note']}")
            if c.get("preconditions"):
                lines.append(f"#### 前置条件: {c['preconditions']}")
            if c.get("steps"):
                lines.append(f"#### 操作步骤: {c['steps']}")
            if c.get("expected"):
                lines.append(f"#### 预期结果: {c['expected']}")
            if c.get("test_data"):
                lines.append(f"#### 测试数据: {c['test_data']}")
        lines.append("")

    return "\n".join(lines)


def cases_to_excel(cases: list, out_path: Path) -> bool:
    """把优化版用例写入 Excel。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "优化版测试用例"

        headers = ["用例ID", "模块", "用例标题", "优先级", "是否新增",
                   "优化说明", "前置条件", "操作步骤", "预期结果", "测试数据"]
        widths  = [10, 16, 35, 8, 8, 25, 20, 35, 30, 20]

        thin   = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", fgColor="2B5FA8")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border
            ws.column_dimensions[ws.cell(1, col).column_letter].width = w
        ws.row_dimensions[1].height = 22

        fill_new = PatternFill("solid", fgColor="E8F5E9")   # 绿色：新增
        fill_opt = PatternFill("solid", fgColor="FFF9C4")   # 黄色：优化

        for row_idx, c in enumerate(cases, 2):
            is_new = c.get("is_new", False)
            fill   = fill_new if is_new else fill_opt
            values = [
                c.get("case_id", f"TC-{row_idx-1:03d}"),
                c.get("module", ""),
                c.get("title", ""),
                c.get("priority", "P1"),
                "是" if is_new else "否",
                c.get("optimization_note", ""),
                c.get("preconditions", ""),
                c.get("steps", ""),
                c.get("expected", ""),
                c.get("test_data", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row_idx].height = 45

        ws.freeze_panes = "A2"
        wb.save(out_path)
        return True
    except Exception as e:
        print(f"  [warn] Excel 生成失败: {e}")
        return False


# ── 主函数 ────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="测试用例评审与优化")
    parser.add_argument("requirement", help="需求文档路径（.md / .txt）")
    parser.add_argument("testcases",  help="人工用例路径（XMind 导出的 .md）")
    parser.add_argument("--out",      help="输出目录（默认 output/review_xxx/）")
    parser.add_argument("--no-optimize", action="store_true",
                        help="只生成评审报告，不生成优化版用例")
    args = parser.parse_args()

    req_path = Path(args.requirement).expanduser()
    tc_path  = Path(args.testcases).expanduser()

    if not req_path.exists():
        print(f"错误: 找不到需求文档 {req_path}")
        sys.exit(1)
    if not tc_path.exists():
        print(f"错误: 找不到用例文档 {tc_path}")
        sys.exit(1)

    ts      = int(time.time())
    out_dir = Path(args.out).expanduser() if args.out else \
              OUTPUT_DIR / f"review_{req_path.stem}_{ts}"
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*52}")
    print(f"  测试用例评审与优化")
    print(f"  需求文档: {req_path.name}")
    print(f"  人工用例: {tc_path.name}")
    print(f"  输出目录: {out_dir.relative_to(AGENT_DIR)}")
    print(f"{'='*52}\n")

    # 读取输入
    req_text = read_file(req_path, limit=6000)
    tc_text  = read_file(tc_path)
    cases    = parse_xmind_md(tc_text)
    print(f"  解析用例: {len(cases)} 条\n")

    if not cases:
        print("错误: 未能从 Markdown 中解析到用例，请确认格式正确")
        sys.exit(1)

    # 阶段一：覆盖率分析
    print("  [1/3] 覆盖率分析...", flush=True)
    coverage = analyze_coverage(req_text, cases)
    cov_score = coverage.get("coverage_score", 0)
    missing   = len(coverage.get("missing_scenarios", []))
    print(f"    覆盖率得分: {cov_score}/100，发现 {missing} 个遗漏场景")

    # 阶段二：质量评估
    print("  [2/3] 用例质量评估...", flush=True)
    quality   = evaluate_quality(cases)
    qual_score = quality.get("total_score", 0)
    print(f"    质量得分: {qual_score}/100（{quality.get('grade', '')}）")

    # 阶段三：生成优化版用例
    optimized = []
    if not args.no_optimize:
        print("  [3/3] 生成优化版用例...", flush=True)
        optimized = generate_optimized_cases(req_text, cases, coverage, quality)
        new_count = sum(1 for c in optimized if c.get("is_new"))
        print(f"    优化版用例: {len(optimized)} 条（含 {new_count} 条新增）")
    else:
        print("  [3/3] 跳过优化（--no-optimize）")

    # 生成报告
    print("\n  写入输出文件...")
    report_md = generate_review_report(req_path.name, cases, coverage, quality, optimized)
    (out_dir / "review_report.md").write_text(report_md, encoding="utf-8")
    print(f"    ✓ review_report.md")

    if optimized:
        xmind_md = cases_to_xmind_md(optimized, req_path.stem)
        (out_dir / "optimized_cases.md").write_text(xmind_md, encoding="utf-8")
        print(f"    ✓ optimized_cases.md（可导入 XMind）")

        xlsx_ok = cases_to_excel(optimized, out_dir / "optimized_cases.xlsx")
        if xlsx_ok:
            print(f"    ✓ optimized_cases.xlsx")

    print(f"\n{'='*52}")
    print(f"  评审完成")
    print(f"  覆盖率: {cov_score}/100  质量: {qual_score}/100")
    print(f"  输出: {out_dir.relative_to(AGENT_DIR)}/")
    print(f"{'='*52}\n")


if __name__ == "__main__":
    main()
