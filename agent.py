#!/usr/bin/env python3
"""
agent.py - 测试用例生成 Agent (v2)

用法:
    python agent.py requirements.md
    python agent.py requirements.md --kb        # 启用知识库
    python agent.py requirements.md --skip-review
    python agent.py requirements.md --no-cases  # 只生成测试点，不展开用例

输出:
    output/testpoints_<name>_<ts>.json    测试点
    output/testcases_<name>_<ts>.json     测试用例（JSON）
    output/testcases_<name>_<ts>.xlsx     测试用例（Excel）
    output/testpoints_<name>_<ts>.xmind  测试点思维导图

Harness: s05 Skills + s04 Subagent + s06 Context Compact
"""

import argparse
import json
import os
import re
import subprocess
import sys
import time
from pathlib import Path

from anthropic import Anthropic
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
from task_store import TaskStore
from memory_store import MemoryStore
from kb_rag import KBRetriever
from memory_rag import MemoryRAG

load_dotenv(override=True)

WORKDIR = Path(__file__).parent
client = Anthropic()
MODEL = os.environ.get("ANTHROPIC_MODEL") or os.environ.get("DEFAULT_LLM_MODEL") or os.environ.get("MODEL_ID", "claude-sonnet-4-6")

KB_DIR     = WORKDIR / "knowledge_base"
SKILLS_DIR = WORKDIR / "skills"
OUTPUT_DIR = WORKDIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

def get_run_dir(stem: str, ts: int) -> Path:
    """每次运行单独一个目录：output/<需求文件名>/<时间戳>/"""
    run_dir = OUTPUT_DIR / stem / str(ts)
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir

# ── s06: Context Compact ───────────────────────────────────────────────────
COMPACT_THRESHOLD = 30000
KEEP_RECENT = 3
PRESERVE_TOOLS = {"read_file"}


def estimate_tokens(messages: list) -> int:
    return len(str(messages)) // 4


def micro_compact(messages: list) -> None:
    tool_results = []
    for msg in messages:
        if msg["role"] == "user" and isinstance(msg.get("content"), list):
            for part in msg["content"]:
                if isinstance(part, dict) and part.get("type") == "tool_result":
                    tool_results.append(part)
    if len(tool_results) <= KEEP_RECENT:
        return
    tool_name_map = {}
    for msg in messages:
        if msg["role"] == "assistant":
            content = msg.get("content", [])
            if isinstance(content, list):
                for block in content:
                    if hasattr(block, "type") and block.type == "tool_use":
                        tool_name_map[block.id] = block.name
    for result in tool_results[:-KEEP_RECENT]:
        if not isinstance(result.get("content"), str) or len(result["content"]) <= 100:
            continue
        tool_name = tool_name_map.get(result.get("tool_use_id", ""), "unknown")
        if tool_name in PRESERVE_TOOLS:
            continue
        result["content"] = f"[Previous: used {tool_name}]"


def auto_compact(messages: list, label: str = "") -> list:
    print(f"  [compact{' '+label if label else ''}] 压缩上下文...")
    conversation_text = json.dumps(messages, default=str, ensure_ascii=False)[-60000:]
    response = client.messages.create(
        model=MODEL,
        messages=[{"role": "user", "content":
            "请用中文简洁总结以下对话，保留：1)已完成工作 2)当前状态 3)关键结论。\n\n"
            + conversation_text}],
        max_tokens=2000,
    )
    summary = next((b.text for b in response.content if hasattr(b, "text")), "无摘要")
    return [{"role": "user", "content": f"[上下文已压缩]\n\n{summary}"}]


# ── s05: Skill Loader ──────────────────────────────────────────────────────
def load_skill(name: str) -> str:
    skill_file = SKILLS_DIR / name / "SKILL.md"
    if not skill_file.exists():
        return f"[Skill '{name}' 不存在]"
    text = skill_file.read_text(encoding="utf-8")
    match = re.match(r"^---\n(.*?)\n---\n(.*)", text, re.DOTALL)
    body = match.group(2).strip() if match else text
    return f'<skill name="{name}">\n{body}\n</skill>'


# ── 工具实现 ───────────────────────────────────────────────────────────────
def safe_path(p: str) -> Path:
    path = (WORKDIR / p).resolve()
    if not path.is_relative_to(WORKDIR):
        raise ValueError(f"路径越界: {p}")
    return path


def run_read(path: str, limit: int = None) -> str:
    try:
        lines = safe_path(path).read_text(encoding="utf-8").splitlines()
        if limit and limit < len(lines):
            lines = lines[:limit] + [f"...（省略 {len(lines)-limit} 行）"]
        return "\n".join(lines)[:50000]
    except Exception as e:
        return f"Error: {e}"


def run_bash(command: str) -> str:
    blocked = ["rm -rf /", "sudo", "shutdown", "> /dev/"]
    if any(b in command for b in blocked):
        return "Error: 危险命令被拦截"
    try:
        r = subprocess.run(command, shell=True, cwd=WORKDIR,
                           capture_output=True, text=True, timeout=30)
        out = (r.stdout + r.stderr).strip()
        return out[:20000] if out else "(无输出)"
    except subprocess.TimeoutExpired:
        return "Error: 超时"


def run_write(path: str, content: str) -> str:
    try:
        fp = safe_path(path)
        fp.parent.mkdir(parents=True, exist_ok=True)
        fp.write_text(content, encoding="utf-8")
        return f"已写入 {len(content)} 字节 → {fp}"
    except Exception as e:
        return f"Error: {e}"


CHILD_TOOLS = [
    {"name": "read_file",
     "description": "读取文件内容",
     "input_schema": {"type": "object",
                      "properties": {"path": {"type": "string"},
                                     "limit": {"type": "integer"}},
                      "required": ["path"]}},
    {"name": "bash",
     "description": "运行 shell 命令",
     "input_schema": {"type": "object",
                      "properties": {"command": {"type": "string"}},
                      "required": ["command"]}},
    {"name": "write_file",
     "description": "写入文件",
     "input_schema": {"type": "object",
                      "properties": {"path": {"type": "string"},
                                     "content": {"type": "string"}},
                      "required": ["path", "content"]}},
    {"name": "load_skill",
     "description": "加载指定技能的完整知识",
     "input_schema": {"type": "object",
                      "properties": {"name": {"type": "string"}},
                      "required": ["name"]}},
    {"name": "todo_write",
     "description": "记录执行计划，开始工作前必须先调用此工具列出步骤",
     "input_schema": {"type": "object",
                      "properties": {
                          "todos": {"type": "array",
                                    "items": {"type": "string"},
                                    "description": "计划步骤列表，按执行顺序排列"}},
                      "required": ["todos"]}},
]

def run_todo_write(todos: list) -> str:
    """s03: 打印执行计划，给用户可见的进度反馈。"""
    lines = ["📋 执行计划:"]
    for i, todo in enumerate(todos, 1):
        lines.append(f"  {i}. {todo}")
    plan = "\n".join(lines)
    print(f"\n{plan}\n")
    return plan


CHILD_HANDLERS = {
    "read_file":  lambda **kw: run_read(kw["path"], kw.get("limit")),
    "bash":       lambda **kw: run_bash(kw["command"]),
    "write_file": lambda **kw: run_write(kw["path"], kw["content"]),
    "load_skill": lambda **kw: load_skill(kw["name"]),
    "todo_write": lambda **kw: run_todo_write(kw["todos"]),
}


# ── JSON 提取工具 ──────────────────────────────────────────────────────────
def extract_json(text: str, fallback, expect_list: bool = False):
    """从模型输出中提取 JSON，兼容代码块、说明文字、截断输出。"""
    import re as _re
    text = text.strip()
    # 去掉 ```json ... ``` 包裹
    text = _re.sub(r"^```json\s*", "", text)
    text = _re.sub(r"\s*```\s*$", "", text)
    text = text.strip()

    # 直接解析
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # 找第一个完整 JSON 块
    start_chars = ["[", "{"] if expect_list else ["{", "["]
    for ch in start_chars:
        idx = text.find(ch)
        if idx == -1:
            continue
        end_ch = "]" if ch == "[" else "}"
        depth = 0
        for i, c in enumerate(text[idx:], idx):
            if c == ch:   depth += 1
            elif c == end_ch:
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[idx:i+1])
                    except json.JSONDecodeError:
                        break
        # 若输出被截断（depth>0），尝试补全后解析
        if depth > 0 and ch == "[":
            truncated = text[idx:].rstrip().rstrip(",")
            # 逐步去掉末尾不完整的对象，直到能解析
            for end in range(len(truncated)-1, idx, -1):
                if truncated[end] == "}":
                    candidate = truncated[:end+1] + "]"
                    try:
                        result = json.loads(candidate)
                        print(f"  [warn] JSON 被截断，成功恢复 {len(result)} 条记录")
                        return result
                    except json.JSONDecodeError:
                        continue
        break

    print(f"  [warn] JSON 解析失败，原始输出前300字符:\n{text[:300]}")
    return fallback


# ── s04: Subagent ──────────────────────────────────────────────────────────
def run_subagent(system: str, prompt: str, label: str = "") -> str:
    messages = [{"role": "user", "content": prompt}]
    print(f"\n  [{label}] 子代理启动...")
    for _ in range(40):
        micro_compact(messages)
        if estimate_tokens(messages) > COMPACT_THRESHOLD:
            messages[:] = auto_compact(messages, label)
        # 遇到 529/429 过载自动重试（最多4次，间隔递增）
        for _retry in range(4):
            try:
                response = client.messages.create(
                    model=MODEL, system=system, messages=messages,
                    tools=CHILD_TOOLS, max_tokens=8000,
                )
                break
            except Exception as e:
                err_str = str(e).lower()
                if "529" in str(e) or "529" in err_str or "overloaded" in err_str or "529" in err_str:
                    wait = (_retry + 1) * 20
                    print(f"  [529] API 过载，{wait}s 后重试 ({_retry+1}/4)...")
                    import time as _time; _time.sleep(wait)
                    if _retry == 3:
                        # s11: 失败不崩溃，返回错误标记
                        return f"__ERROR__: {e}"
                elif "rate_limit" in err_str or "429" in str(e):
                    wait = (_retry + 1) * 30
                    print(f"  [429] 限速，{wait}s 后重试...")
                    import time as _time; _time.sleep(wait)
                    if _retry == 3:
                        return f"__ERROR__: {e}"
                else:
                    # 其他错误直接返回，不重试
                    return f"__ERROR__: {e}"
        messages.append({"role": "assistant", "content": response.content})
        if response.stop_reason != "tool_use":
            break
        results = []
        for block in response.content:
            if block.type == "tool_use":
                handler = CHILD_HANDLERS.get(block.name)
                try:
                    output = handler(**block.input) if handler else f"未知工具: {block.name}"
                except Exception as e:
                    output = f"Error: {e}"
                print(f"    → {block.name}({list(block.input.keys())[0] if block.input else ''})")
                results.append({
                    "type": "tool_result",
                    "tool_use_id": block.id,
                    "content": str(output)[:30000],
                })
        messages.append({"role": "user", "content": results})
    return "".join(b.text for b in response.content if hasattr(b, "text")) or "(无输出)"


# ── 阶段一：需求评审 ───────────────────────────────────────────────────────
def stage1_review(req_path: Path, memory=None) -> dict:
    system = (
        "你是一名资深测试工程师，负责需求文档评审。"
        "执行步骤（严格按顺序，不得跳过）：\n"
        "1. 用 todo_write 列出执行计划\n"
        "2. 用 load_skill 加载 requirement-review 技能\n"
        "3. 用 read_file 读取需求文档\n"
        "4. 用 write_file 把评审结果 JSON 写入 output/_review_tmp.json\n"
        "【绝对禁止】：不得使用 bash 工具执行任何命令。\n"
        "【写入格式】：write_file 的 content 必须是合法 JSON 对象，"
        "以 { 开头以 } 结尾，不含任何其他文字、代码块或 markdown 标记。"
    )
    # 用向量检索找相关历史经验（比全量注入更精准）
    req_preview = req_path.read_text(encoding="utf-8")[:500] if req_path.exists() else ""
    if memory:
        try:
            from memory_rag import MemoryRAG as _MR
            _mr = _MR()
            mem_ctx = _mr.search(f"{req_path.stem} {req_preview}", top_k=5)
        except Exception:
            mem_ctx = memory.get_context_for_review()
    else:
        mem_ctx = ""
    prompt = (
        f"需求文档路径: {req_path.relative_to(WORKDIR)}\n\n"
        + (f"【历史经验参考】\n{mem_ctx}\n\n" if mem_ctx else "")
        + "请开始需求评审。"
    )
    result = run_subagent(system, prompt, label="需求评审")
    if result.startswith("__ERROR__"):
        print(f"  [s11] 需求评审失败: {result}，使用空评审结果继续")
        return {"testable_features": [], "risk_flags": [], "score": 0}

    # 优先从临时文件读取（子代理用 write_file 写入）
    review_tmp = OUTPUT_DIR / "_review_tmp.json"
    if review_tmp.exists():
        try:
            raw = review_tmp.read_text(encoding="utf-8")
            review_tmp.unlink(missing_ok=True)
            data = extract_json(raw, fallback={})
            if data and isinstance(data, dict):
                return data
        except Exception:
            pass

    # 降级：从返回文本解析
    data = extract_json(result, fallback={"testable_features": [], "risk_flags": []})

    # 如果解析失败，尝试格式修复
    if not data or not isinstance(data, dict):
        print("  [重试] 评审 JSON 解析失败，尝试格式修复...")
        try:
            fix_resp = client.messages.create(
                model=MODEL,
                system="把以下内容转换为合法 JSON 对象，只输出 JSON，不要其他文字：",
                messages=[{"role": "user", "content": result[:3000]}],
                max_tokens=2000,
            )
            fix_text = "".join(b.text for b in fix_resp.content if hasattr(b, "text"))
            data = extract_json(fix_text, fallback={"testable_features": [], "risk_flags": []})
            if data:
                print("  [重试] 格式修复成功")
        except Exception as e:
            print(f"  [重试] 格式修复失败: {e}")

    return data if isinstance(data, dict) else {"testable_features": [], "risk_flags": []}


# ── 阶段二：测试点生成 ─────────────────────────────────────────────────────
def stage2_testpoints(req_path: Path, review: dict, use_kb: bool, memory=None) -> list:
    """
    两阶段测试点生成：
    阶段A：只读需求文档，生成 REQ 来源测试点（轻量）
    阶段B：读知识库 + 因子设计，补充 KB/RISK 测试点（按需）
    """
    # ── 阶段 A：需求文档 → REQ 测试点 ──────────────────────────────────────
    system_a = (
        "你是一名资深测试工程师，专门生成测试点 JSON。"
        "执行步骤（严格按顺序）：\n"
        "1. 用 todo_write 列出步骤\n"
        "2. 用 load_skill 加载 testpoint-gen 技能\n"
        "3. 用 read_file 读取需求文档\n"
        "4. 用 write_file 把 JSON 数组写入指定文件\n"
        "【绝对禁止】：不得使用 bash 工具，不得在终端运行命令。\n"
        "【写入格式】：write_file 的 content 参数必须是合法 JSON 数组，"
        "以 [ 开头以 ] 结尾，不含任何其他文字、代码块标记或说明。"
    )
    tp_tmp_a = OUTPUT_DIR / "_testpoints_stage2_a.json"
    tp_tmp_a.unlink(missing_ok=True)

    # 向量检索历史测试点经验
    if memory:
        try:
            from memory_rag import MemoryRAG as _MR
            _mr = _MR()
            req_preview = req_path.read_text(encoding="utf-8")[:500] if req_path.exists() else ""
            mem_tp_ctx = _mr.search(f"{req_path.stem} {req_preview}", top_k=6)
        except Exception:
            mem_tp_ctx = memory.get_context_for_testpoints()
    else:
        mem_tp_ctx = ""
    prompt_a = (
        f"需求文档路径: {req_path.relative_to(WORKDIR)}\n\n"
        f"需求评审结果:\n{json.dumps(review, ensure_ascii=False, indent=2)}\n\n"
        + (f"【历史经验参考】\n{mem_tp_ctx}\n\n" if mem_tp_ctx else "")
        + "请仅基于需求文档生成 REQ 来源测试点。\n"
        "输出格式：纯 JSON 数组，每个元素包含 testpoint_id/functional_module/test_scenario/"
        "source/preconditions/test_steps/expected_result/priority/remarks 字段。\n"
        "source 字段固定填 REQ。\n"
        f"完成后必须用 write_file 把 JSON 数组（只有数组，不含其他文字）写入 {tp_tmp_a.relative_to(WORKDIR)}。"
    )
    result_a = run_subagent(system_a, prompt_a, label="测试点-需求文档")

    # 读回阶段 A 的结果
    req_tps = []
    if tp_tmp_a.exists():
        raw = tp_tmp_a.read_text(encoding="utf-8")
        tp_tmp_a.unlink(missing_ok=True)
        data = extract_json(raw, fallback=[], expect_list=True)
        if isinstance(data, list):
            req_tps = [normalize_testpoint(tp, i) for i, tp in enumerate(data)]

    # 从返回文本降级解析
    if not req_tps and not result_a.startswith("__ERROR__"):
        data = extract_json(result_a, fallback=[], expect_list=True)
        if isinstance(data, list):
            req_tps = [normalize_testpoint(tp, i) for i, tp in enumerate(data)]

    # 阶段A失败时自动重试：让模型把已有内容重新格式化为 JSON
    if not req_tps and not result_a.startswith("__ERROR__") and len(result_a) > 100:
        print(f"  [重试] 阶段A输出非 JSON，尝试格式修复...")
        fix_prompt = (
            "以下是测试点内容，请将其转换为合法 JSON 数组格式输出，"
            "不要有任何其他文字，直接以 [ 开头，以 ] 结尾：\n\n"
            + result_a[:3000]
        )
        try:
            fix_response = client.messages.create(
                model=MODEL,
                system="你是 JSON 格式化工具。只输出合法 JSON 数组，不输出任何其他内容。",
                messages=[{"role": "user", "content": fix_prompt}],
                max_tokens=4000,
            )
            fix_text = "".join(b.text for b in fix_response.content if hasattr(b, "text"))
            data = extract_json(fix_text, fallback=[], expect_list=True)
            if isinstance(data, list) and data:
                req_tps = [normalize_testpoint(tp, i) for i, tp in enumerate(data)]
                print(f"  [重试] 格式修复成功，恢复 {len(req_tps)} 条测试点")
        except Exception as e:
            print(f"  [重试] 格式修复失败: {e}")

    print(f"  阶段A完成: {len(req_tps)} 条 REQ 测试点")

    # REQ=0 时先尝试格式修复重试，修复失败才停止
    if not req_tps:
        if not result_a.startswith("__ERROR__") and len(result_a) > 50:
            print(f"  [重试] 阶段A输出非 JSON，尝试格式修复...")
            try:
                fix_resp = client.messages.create(
                    model=MODEL,
                    system="你是 JSON 格式化工具。把以下内容转换为合法 JSON 数组，只输出数组本身，以 [ 开头以 ] 结尾，不要其他文字。",
                    messages=[{"role": "user", "content": result_a[:4000]}],
                    max_tokens=4000,
                )
                fix_text = "".join(b.text for b in fix_resp.content if hasattr(b, "text"))
                data = extract_json(fix_text, fallback=[], expect_list=True)
                if isinstance(data, list) and data:
                    req_tps = [normalize_testpoint(tp, i) for i, tp in enumerate(data)]
                    print(f"  [重试] 格式修复成功，恢复 {len(req_tps)} 条测试点")
            except Exception as e:
                print(f"  [重试] 格式修复失败: {e}")

        if not req_tps:
            print(f"  [阶段A] 未生成 REQ 测试点，跳过此章节继续")
            return []

    if not use_kb or not KB_DIR.exists():
        return req_tps

    # ── 阶段 B：RAG 检索知识库 → KB/RISK 补充（语义检索，精准无截断）──────────
    print(f"\n  [测试点-知识库补充] RAG 语义检索...", flush=True)

    try:
        retriever  = KBRetriever(kb_dir=KB_DIR)
        # 用需求文档内容 + 评审结果做检索查询
        req_text   = req_path.read_text(encoding="utf-8")
        review_str = json.dumps(review, ensure_ascii=False)
        query      = f"{req_path.stem}\n{req_text[:1000]}\n{review_str[:500]}"
        kb_context = retriever.search_for_requirement(query, top_k=12)
    except Exception as e:
        print(f"  [RAG] 检索失败: {e}，跳过知识库补充")
        return req_tps

    # 额外：用 RAG 从因子设计文档里检索相关段落
    # 用相关度阈值+token预算双重控制，自适应决定注入多少内容
    design_context = ""
    design_dir = KB_DIR / "design"
    if design_dir.exists() and list(design_dir.glob("*.md")):
        try:
            design_retriever = KBRetriever(
                kb_dir=design_dir,
                index_dir=WORKDIR / ".design_index",
            )
            design_query = f"{req_path.stem}\n{req_text[:800]}"
            # 先取较多候选，再按阈值和预算过滤，最终注入的是真正相关的段落
            candidates   = design_retriever.search(design_query, top_k=20)

            SCORE_THRESHOLD = 0.60   # 低于此相关度的段落不采用
            TOKEN_BUDGET    = 4000   # 最多注入的字符数

            selected = []
            total_chars = 0
            for hit in candidates:
                if hit["score"] < SCORE_THRESHOLD:
                    break  # 结果已按相关度降序，后面的更低可以直接停
                if total_chars + len(hit["content"]) > TOKEN_BUDGET:
                    break
                selected.append(hit)
                total_chars += len(hit["content"])

            if selected:
                lines = ["【因子设计文档相关内容（RAG检索）】\n"]
                prev_source = None
                for hit in selected:
                    if hit["source"] != prev_source:
                        lines.append(f"\n--- {hit['source']} (相关度:{hit['score']:.2f}) ---")
                        prev_source = hit["source"]
                    lines.append(hit["content"])
                design_context = "\n".join(lines)
                unique_files = {h["source"] for h in selected}
                print(f"  [设计文档] 纳入 {len(selected)} 段（阈值≥{SCORE_THRESHOLD}，"
                      f"共{total_chars}字），来自 {len(unique_files)} 个文件", flush=True)
                for f in unique_files:
                    print(f"    ✓ {f}", flush=True)
            else:
                print(f"  [设计文档] 无相关内容（最高相关度: "
                      f"{candidates[0]['score']:.2f} < {SCORE_THRESHOLD}）" if candidates else
                      f"  [设计文档] 无候选内容", flush=True)
        except Exception as e:
            print(f"  [warn] 设计文档RAG检索失败: {e}", flush=True)

    if not kb_context and not design_context:
        print(f"  阶段B完成: 0 条（知识库无相关内容）")
        return req_tps

    # 直接调一次 API，用检索结果生成 KB/RISK 测试点
    offset   = len(req_tps)
    system_b = "你是一名资深测试工程师，专门生成 KB 和 RISK 来源的测试点。只输出 JSON 数组，不要其他文字。"

    # 合并 RAG 结果和设计文档内容
    full_context = kb_context
    if design_context:
        full_context = design_context + "\n\n" + kb_context if kb_context else design_context

    prompt_b = (
        f"需求文档: {req_path.name}\n\n"
        f"{full_context}\n\n"
        f"基于以上知识库内容，为需求文档生成 KB 和 RISK 来源的测试点：\n"
        f"- KB 测试点：针对知识库中的枚举值、字段约束、数据表取值逻辑，每个枚举值一条，"
        f"source_ref 填写知识库来源文件名\n"
        f"- RISK 测试点：并发竞争、数据精度丢失、外部依赖失败、数据同步延迟等，至少3条\n"
        f"- testpoint_id 从 TP-{offset+1:03d} 开始递增\n"
        f"- source 字段只能填 KB 或 RISK\n\n"
        "输出纯 JSON 数组，格式：\n"
        '[{"testpoint_id":"TP-xxx","functional_module":"xxx","test_scenario":"xxx",'
        '"source":"KB","source_ref":"来源文件名","preconditions":"xxx","test_steps":"xxx",'
        '"expected_result":"xxx","priority":"P1","remarks":""}]'
    )

    try:
        response = client.messages.create(
            model=MODEL,
            system=system_b,
            messages=[{"role": "user", "content": prompt_b}],
            max_tokens=4000,
        )
        result_b = "".join(b.text for b in response.content if hasattr(b, "text"))
        data     = extract_json(result_b, fallback=[], expect_list=True)
        kb_tps   = []
        if isinstance(data, list):
            kb_tps = [normalize_testpoint(tp, offset + i) for i, tp in enumerate(data)]
        print(f"  阶段B完成: {len(kb_tps)} 条 KB/RISK 测试点")
    except Exception as e:
        print(f"  [s11] 阶段B失败: {e}")
        kb_tps = []

    return req_tps + kb_tps


# ── 阶段三：测试用例生成（分批处理，每批10条）──────────────────────────────
BATCH_SIZE = 10

def stage3_testcases_batch(batch: list, batch_no: int, case_id_start: int) -> list:
    """处理单批测试点，返回用例列表。"""
    tp_file  = OUTPUT_DIR / f"_tp_batch_{batch_no}.json"
    out_file = OUTPUT_DIR / f"_tc_batch_{batch_no}.json"
    out_file.unlink(missing_ok=True)

    tp_file.write_text(json.dumps(batch, ensure_ascii=False, indent=2), encoding="utf-8")

    system = (
        "你是一名资深测试工程师，负责把测试点展开为完整测试用例。\n"
        "严格按以下步骤执行，每个步骤只执行一次，不得重复：\n"
        "1. todo_write：列出执行计划（只调一次）\n"
        "2. load_skill：加载 testcase-gen 技能（只调一次）\n"
        "3. read_file：读取测试点文件（只调一次）\n"
        "4. write_file：写入用例 JSON 数组（只调一次）\n"
        "【绝对禁止】：\n"
        "- 不得重复调用任何工具\n"
        "- 不得使用 bash 工具\n"
        "- write_file 的 content 必须是合法 JSON 数组，以 [ 开头以 ] 结尾\n"
        "- 不得在 JSON 前后添加任何文字说明"
    )
    prompt = (
        f"测试点文件: {tp_file.relative_to(WORKDIR)}（共 {len(batch)} 条测试点）\n"
        f"用例ID 从 TC-{case_id_start:03d} 开始递增。\n\n"
        "展开规则：每个独立路径/分支对应一条用例，覆盖完整为止。\n"
        f"生成完毕后用 write_file 写入 {out_file.relative_to(WORKDIR)}。"
    )
    result = run_subagent(system, prompt, label=f"用例生成 batch{batch_no}")
    tp_file.unlink(missing_ok=True)

    # s11: 失败跳过本批，不影响其他批次
    if result.startswith("__ERROR__"):
        print(f"  [s11] batch{batch_no} 失败，跳过: {result[:80]}")
        return []

    if out_file.exists():
        raw  = out_file.read_text(encoding="utf-8")
        out_file.unlink(missing_ok=True)
        data = extract_json(raw, fallback=[], expect_list=True)
        if not isinstance(data, list):
            return []
        # 标准化测试用例字段，case_id 从本批起始序号递增
        offset = (batch_no - 1) * BATCH_SIZE
        data = [normalize_testcase(c, offset + i + 1) for i, c in enumerate(data)]
        return data
    print(f"  [warn] batch{batch_no} 未写入文件，跳过")
    return []


def stage3_testcases(testpoints: list, req_path: Path) -> list:
    """分批调子代理，并行处理所有批次。"""
    batches = [testpoints[i:i+BATCH_SIZE] for i in range(0, len(testpoints), BATCH_SIZE)]
    total   = len(batches)
    print(f"  共 {len(testpoints)} 条测试点，分 {total} 批并行处理（每批 {BATCH_SIZE} 条）")

    # 并行数不超过批次数，也不超过4（避免 API 限速）
    max_workers = min(total, 4)
    results     = {}   # batch_no -> cases

    def run_batch(args):
        batch_no, batch = args
        # case_id 按批次固定偏移，不依赖其他批次完成顺序
        case_id_start = (batch_no - 1) * BATCH_SIZE + 1
        cases = stage3_testcases_batch(batch, batch_no, case_id_start)
        return batch_no, cases

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(run_batch, (i, batch)): i
            for i, batch in enumerate(batches, 1)
        }
        for future in as_completed(futures):
            try:
                batch_no, cases = future.result()
                results[batch_no] = cases
                print(f"  batch{batch_no}/{total} 完成，本批生成 {len(cases)} 条用例")
            except Exception as e:
                batch_no = futures[future]
                print(f"  [s11] batch{batch_no} 异常，跳过: {e}")
                results[batch_no] = []

    # 按批次顺序合并（并行完成顺序不定）
    all_cases = []
    for i in range(1, total + 1):
        all_cases.extend(results.get(i, []))

    return all_cases


# ── 字段标准化 ────────────────────────────────────────────────────────────────
def normalize_testcase(case: dict, idx: int) -> dict:
    """
    把模型可能输出的各种字段名统一映射到标准字段。
    兼容: case_name/title/test_name → case_title
          pre_condition/precondition → preconditions
          step/test_steps/procedure → steps
          expected/result → expected_result
          module/test_module → functional_module
    """
    # 字段别名映射表
    alias = {
        "case_title":        ["case_name", "title", "test_name", "用例标题", "case_description"],
        "functional_module": ["module", "test_module", "feature", "功能模块"],
        "preconditions":     ["precondition", "pre_condition", "prerequisite", "前置条件"],
        "test_data":         ["input_data", "test_input", "data", "测试数据"],
        "steps":             ["step", "test_steps", "procedure", "操作步骤", "test_procedure"],
        "expected_result":   ["expected", "result", "expect", "预期结果", "expected_output"],
        "source":            ["test_source", "来源"],
        "priority":          ["level", "test_priority", "优先级"],
        "remarks":           ["remark", "note", "comment", "备注"],
    }

    normalized = dict(case)  # 先复制原始数据

    # 应用别名映射（只在标准字段不存在时才映射）
    for std_key, aliases in alias.items():
        if std_key not in normalized or not normalized[std_key]:
            for a in aliases:
                if a in normalized and normalized[a]:
                    normalized[std_key] = normalized[a]
                    break

    # 确保所有标准字段都存在
    defaults = {
        "case_id":           f"TC-{idx:03d}",
        "testpoint_id":      "",
        "functional_module": "",
        "case_title":        "",
        "source":            "REQ",
        "priority":          "P1",
        "preconditions":     "",
        "test_data":         "",
        "steps":             "",
        "expected_result":   "",
        "actual_result":     "",
        "status":            "",
        "remarks":           "",
    }
    for k, v in defaults.items():
        if k not in normalized or normalized[k] is None:
            normalized[k] = v

    # steps 如果是列表，转成换行字符串
    if isinstance(normalized.get("steps"), list):
        normalized["steps"] = "\n".join(
            f"{i+1}. {s}" for i, s in enumerate(normalized["steps"])
        )

    return normalized


# ── 输出：Excel ────────────────────────────────────────────────────────────
def export_excel(testcases: list, out_path: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  [skip] Excel 输出需要 openpyxl: pip install openpyxl")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 列定义：(列名, 字段key, 宽度)
    columns = [
        ("用例ID",      "case_id",          12),
        ("测试点ID",    "testpoint_id",      12),
        ("功能模块",    "functional_module", 18),
        ("用例标题",    "case_title",        35),
        ("来源",        "source",             8),
        ("优先级",      "priority",           8),
        ("前置条件",    "preconditions",     25),
        ("测试数据",    "test_data",         20),
        ("操作步骤",    "steps",             40),
        ("预期结果",    "expected_result",   35),
        ("实际结果",    "actual_result",     25),
        ("执行状态",    "status",            10),
        ("备注",        "remarks",           20),
    ]

    # 样式
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="2B5FA8")
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 来源颜色
    source_colors = {"REQ": "DDEEFF", "KB": "FFF9DD", "RISK": "FFE8E8"}

    # 表头
    for col_idx, (col_name, _, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = col_width
    ws.row_dimensions[1].height = 22

    # 数据行
    priority_colors = {"P0": "FF4444", "P1": "FF8800", "P2": "888888"}

    for row_idx, case in enumerate(testcases, 2):
        case   = normalize_testcase(case, row_idx - 1)   # 标准化字段
        source = case.get("source", "REQ")
        row_fill = PatternFill("solid", fgColor=source_colors.get(source, "FFFFFF"))

        for col_idx, (_, field_key, _) in enumerate(columns, 1):
            value = case.get(field_key, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.fill   = row_fill

            if field_key in ("case_id", "testpoint_id", "source", "priority", "status"):
                cell.alignment = center
                if field_key == "priority":
                    cell.font = Font(color=priority_colors.get(value, "000000"), bold=True)
            else:
                cell.alignment = left_wrap

        ws.row_dimensions[row_idx].height = 45

    # 冻结首行
    ws.freeze_panes = "A2"

    # 图例 sheet
    legend_ws = wb.create_sheet("图例说明")
    legend_data = [
        ("颜色", "含义"),
        ("蓝色底", "REQ — 来自需求文档"),
        ("黄色底", "KB  — 来自知识库补充"),
        ("红色底", "RISK — 风险推断"),
        ("",       ""),
        ("优先级", "说明"),
        ("P0",     "核心必测"),
        ("P1",     "重要应测"),
        ("P2",     "边缘可测"),
    ]
    for r, (a, b) in enumerate(legend_data, 1):
        legend_ws.cell(row=r, column=1, value=a)
        legend_ws.cell(row=r, column=2, value=b)

    wb.save(out_path)
    return True


def export_excel(testcases: list, out_path: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  [skip] Excel 输出需要 openpyxl: pip install openpyxl")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    columns = [
        ("用例ID",      "case_id",          12),
        ("测试点ID",    "testpoint_id",      12),
        ("功能模块",    "functional_module", 18),
        ("用例标题",    "case_title",        35),
        ("来源",        "source",             8),
        ("优先级",      "priority",           8),
        ("前置条件",    "preconditions",     25),
        ("测试数据",    "test_data",         20),
        ("操作步骤",    "steps",             40),
        ("预期结果",    "expected_result",   35),
        ("实际结果",    "actual_result",     25),
        ("执行状态",    "status",            10),
        ("备注",        "remarks",           20),
    ]

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="2B5FA8")
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    source_colors = {"REQ": "DDEEFF", "KB": "FFF9DD", "RISK": "FFE8E8"}

    for col_idx, (col_name, _, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = col_width
    ws.row_dimensions[1].height = 22

    priority_colors = {"P0": "FF4444", "P1": "FF8800", "P2": "888888"}

    for row_idx, case in enumerate(testcases, 2):
        case     = normalize_testcase(case, row_idx - 1)
        source   = case.get("source", "REQ")
        row_fill = PatternFill("solid", fgColor=source_colors.get(source, "FFFFFF"))

        for col_idx, (_, field_key, _) in enumerate(columns, 1):
            value = case.get(field_key, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.fill   = row_fill
            if field_key in ("case_id", "testpoint_id", "source", "priority", "status"):
                cell.alignment = center
                if field_key == "priority":
                    cell.font = Font(color=priority_colors.get(value, "000000"), bold=True)
            else:
                cell.alignment = left_wrap
        ws.row_dimensions[row_idx].height = 45

    ws.freeze_panes = "A2"

    legend_ws = wb.create_sheet("图例说明")
    for r, (a, b) in enumerate([
        ("颜色", "含义"), ("蓝色底", "REQ — 来自需求文档"),
        ("黄色底", "KB  — 来自知识库补充"), ("红色底", "RISK — 风险推断"),
        ("", ""), ("优先级", "说明"), ("P0", "核心必测"),
        ("P1", "重要应测"), ("P2", "边缘可测"),
    ], 1):
        legend_ws.cell(row=r, column=1, value=a)
        legend_ws.cell(row=r, column=2, value=b)

    wb.save(out_path)
    return True
def normalize_testpoint(tp: dict, idx: int = 0) -> dict:
    """
    统一测试点字段名，兼容模型各种输出格式。
    支持字段名：id/title/level/desc/expected 等非标准字段。
    """
    n = dict(tp)

    # ── testpoint_id ──────────────────────────────────────────────────────
    if not n.get("testpoint_id"):
        n["testpoint_id"] = (n.get("id") or n.get("tp_id") or
                             n.get("case_id") or f"TP-{idx+1:03d}")

    # ── test_scenario（标题）─────────────────────────────────────────────
    if not n.get("test_scenario"):
        n["test_scenario"] = (n.get("title") or n.get("name") or
                              n.get("case_title") or n.get("scenario") or "")

    # ── priority（优先级）注意：必须先检查 level，再用默认值 ─────────────
    # 字段可能叫 level / test_priority / 优先级，值可能是 P0/P1/P2
    raw_priority = (n.get("priority") or n.get("level") or
                    n.get("test_priority") or n.get("优先级") or "P1")
    # 标准化：只接受 P0/P1/P2
    if raw_priority not in ("P0", "P1", "P2"):
        raw_priority = "P1"
    n["priority"] = raw_priority

    # ── functional_module（功能模块）─────────────────────────────────────
    # 优先从显式字段取，没有则从 title 推断关键词分组
    if not n.get("functional_module"):
        mod = (n.get("module") or n.get("feature") or
               n.get("category") or n.get("functional_area") or "")
        if not mod:
            # 从 title 提取：取"-"前的部分作为分组名
            title = n.get("test_scenario", "")
            if "-" in title:
                mod = title.split("-")[0].strip()
            elif "验证" in title or "计算" in title:
                mod = title[:8].strip()
            else:
                mod = "功能验证"
        n["functional_module"] = mod

    # ── source ────────────────────────────────────────────────────────────
    src = n.get("source") or n.get("test_source") or n.get("来源") or "REQ"
    if src not in ("REQ", "KB", "RISK"):
        src = "REQ"
    n["source"] = src

    # ── expected_result ───────────────────────────────────────────────────
    if not n.get("expected_result"):
        n["expected_result"] = (n.get("expected") or n.get("expect") or
                                n.get("expected_output") or "")

    # ── preconditions ─────────────────────────────────────────────────────
    if not n.get("preconditions"):
        n["preconditions"] = (n.get("precondition") or n.get("pre_condition") or
                              n.get("desc") or n.get("description") or "")

    # ── source_ref ────────────────────────────────────────────────────────
    if not n.get("source_ref"):
        n["source_ref"] = n.get("source_reference") or n.get("ref") or ""

    # ── remarks ───────────────────────────────────────────────────────────
    if not n.get("remarks"):
        n["remarks"] = n.get("remark") or n.get("note") or n.get("comment") or ""

    return n


def export_markdown_xmind(testpoints: list, review: dict, req_name: str, out_path: Path) -> bool:
    """
    生成可导入 XMind 的 Markdown 文件。
    XMind 导入步骤: 文件 → 导入 → Markdown
    层级结构:
        # 根节点（需求名）
        ## 风险项
        ### 风险1
        ## 功能模块
        ### [来源][优先级] 测试点标题
        #### 前置条件 / 预期结果
    """
    lines = []
    # 先标准化所有测试点字段
    testpoints = [normalize_testpoint(tp, i) for i, tp in enumerate(testpoints)]

    lines.append(f"# {req_name}")
    lines.append(f"## 概览")
    lines.append(f"### 评审分: {review.get('score', 'N/A')}")
    lines.append(f"### 测试点总数: {len(testpoints)}")
    req_c  = sum(1 for t in testpoints if t.get("source") == "REQ")
    kb_c   = sum(1 for t in testpoints if t.get("source") == "KB")
    risk_c = sum(1 for t in testpoints if t.get("source") == "RISK")
    lines.append(f"### REQ需求直出: {req_c} | KB知识库: {kb_c} | RISK风险: {risk_c}")

    # 风险项
    risks = review.get("risk_flags", [])
    if risks:
        lines.append(f"## ⚠ 风险项 ({len(risks)}条)")
        for r in risks:
            lines.append(f"### [{r.get('type','?')}] {r.get('desc','')}")

    # 按功能模块分组（归一化模块名，避免同一模块因名称细微差异被分开）
    def normalize_module(name: str) -> str:
        """模块名归一化：去掉括号内容、版本号、多余空格"""
        import re as _re
        name = name.strip()
        name = _re.sub(r'[（(][^）)]*[）)]', '', name)  # 去掉括号
        name = _re.sub(r'V[\d.]+', '', name)             # 去掉版本号
        name = _re.sub(r'[\s_-]+', '', name)             # 去掉空格/下划线
        return name.strip() or "未分类"

    # 先按归一化名分组，保留原始名（取第一个出现的）
    modules: dict = {}          # 归一化名 → [tp, ...]
    mod_display: dict = {}      # 归一化名 → 显示名
    for tp in testpoints:
        raw_mod  = tp.get("functional_module") or tp.get("feature") or "未分类"
        norm_mod = normalize_module(raw_mod)
        if norm_mod not in mod_display:
            mod_display[norm_mod] = raw_mod
        modules.setdefault(norm_mod, []).append(tp)

    # 按 P0 数量降序排列模块（重要模块在前）
    modules = dict(sorted(
        modules.items(),
        key=lambda x: (
            -sum(1 for t in x[1] if t.get("priority") == "P0"),
            -len(x[1])
        )
    ))

    source_icon = {"REQ": "🔵", "KB": "🟡", "RISK": "🔴"}

    for norm_mod, tps in modules.items():
        mod_name = mod_display.get(norm_mod, norm_mod)
        mc = sum(1 for t in tps if t.get("source")=="REQ")
        kc = sum(1 for t in tps if t.get("source")=="KB")
        rc = sum(1 for t in tps if t.get("source")=="RISK")
        p0 = sum(1 for t in tps if t.get("priority")=="P0")
        lines.append(f"## {mod_name} ({len(tps)}条)")
        lines.append(f"### 统计: REQ={mc} KB={kc} RISK={rc} | P0={p0}")

        # 同模块内按来源排序：REQ → KB → RISK，相同来源内 P0 优先
        tps_sorted = sorted(tps, key=lambda t: (
            {"REQ": 0, "KB": 1, "RISK": 2}.get(t.get("source", "REQ"), 3),
            {"P0": 0, "P1": 1, "P2": 2}.get(t.get("priority", "P1"), 3)
        ))
        for tp in tps_sorted:
            src   = tp.get("source", "REQ")
            pri   = tp.get("priority", "P1")
            icon  = source_icon.get(src, "⚪")
            title = tp.get("test_scenario") or tp.get("title") or tp.get("case_title", "")
            lines.append(f"### {icon}[{src}][{pri}] {title}")

            # 子节点放关键信息
            if tp.get("preconditions"):
                lines.append(f"#### 前置: {tp['preconditions']}")
            if tp.get("expected_result"):
                lines.append(f"#### 预期: {tp['expected_result']}")
            if tp.get("source_ref"):
                lines.append(f"#### 来源: {tp['source_ref']}")
            if tp.get("remarks"):
                lines.append(f"#### 备注: {tp['remarks']}")

    try:
        out_path.write_text("\n".join(lines), encoding="utf-8")
        return True
    except Exception as e:
        print(f"  [warn] Markdown 生成失败: {e}")
        return False


# ── 统计辅助 ──────────────────────────────────────────────────────────────
def get_source(tp: dict) -> str:
    s = tp.get("source", "")
    if s in ("REQ", "KB", "RISK"):
        return s
    kb = tp.get("kb_source", "") + tp.get("source_ref", "")
    if "knowledge_base" in kb:
        return "KB"
    remarks = tp.get("remarks", "")
    if "风险" in remarks:
        return "RISK"
    return "REQ"


# ── 主流程 ─────────────────────────────────────────────────────────────────
def _extract_section(req_path: Path, keyword: str) -> str:
    """
    从需求文档中提取包含指定关键词的章节。
    支持 Markdown # ## ### 标题层级。
    """
    try:
        text = req_path.read_text(encoding="utf-8")
    except Exception:
        return ""

    # 过滤删除线内容（~~删除的需求~~）
    import re as _re_st
    text = _re_st.sub(r'~~.+?~~', '', text)

    lines = text.splitlines()
    result = []
    in_section = False
    section_level = 0

    for line in lines:
        # 检测标题行
        if line.startswith("#"):
            level = len(line) - len(line.lstrip("#"))
            title = line.lstrip("#").strip()

            if keyword in title:
                # 找到目标章节
                in_section    = True
                section_level = level
                result.append(line)
            elif in_section:
                # 遇到同级或更高级标题，章节结束
                if level <= section_level:
                    in_section = False
                else:
                    result.append(line)
        elif in_section:
            result.append(line)

    # 如果没匹配到标题，尝试全文关键词段落匹配
    if not result:
        paras = text.split("\n\n")
        for para in paras:
            if keyword in para:
                result.append(para)

    return "\n".join(result).strip()


def _split_sections(req_path: Path, min_lines: int = 5,
                    memory=None) -> list:
    """
    自动按 Markdown 二级标题(##)拆分需求文档为多个章节。
    三层过滤机制：
      第一层：黑名单关键词快速过滤
      第二层：长期记忆（历史学习结果）
      第三层：模型判断兜底（结果自动写回记忆）
    返回 [(章节标题, 章节内容)] 列表。
    """
    # ── 第一层：黑名单 ────────────────────────────────────────────────────────
    NON_CORE_KW = [
        "文档范围", "适用范围", "参考文档", "参考资料", "相关文档",
        "修订记录", "版本历史", "变更历史", "文档说明", "目录",
        "监管条文", "监管解读", "法规", "条文解读",
        "客户规则", "规则示例", "示例", "案例",
        "背景", "概述", "简介", "介绍", "说明",
        "术语", "词汇", "缩略语", "定义",
    ]

    def blacklist_hit(title: str) -> bool:
        clean = title.strip("*~_ ")
        return any(kw in clean for kw in NON_CORE_KW)

    # ── 解析所有章节 ──────────────────────────────────────────────────────────
    try:
        text = req_path.read_text(encoding="utf-8")
    except Exception:
        return []

    # 过滤删除线内容（~~废弃的需求~~）
    import re as _re_st2
    # 多行删除线也处理
    text = _re_st2.sub(r'~~.+?~~', '', text, flags=_re_st2.DOTALL)

    lines     = text.splitlines()
    raw_sections = []
    cur_title = None
    cur_lines = []

    for line in lines:
        if line.startswith("## ") and not line.startswith("### "):
            if cur_title and len(cur_lines) >= min_lines:
                raw_sections.append((cur_title, "\n".join(cur_lines)))
            raw_title = line.lstrip("# ").strip()
            # 标题本身是删除线（~~章节名~~）则跳过整个章节
            import re as _re_title
            if _re_title.fullmatch(r'~~.+~~', raw_title):
                print(f"  [删除线] 跳过已废弃章节: {raw_title}")
                cur_title = None
                cur_lines = []
            else:
                cur_title = raw_title
                cur_lines = [line]
        elif line.startswith("# ") and not line.startswith("## "):
            pass
        elif cur_title:
            cur_lines.append(line)
    if cur_title and len(cur_lines) >= min_lines:
        raw_sections.append((cur_title, "\n".join(cur_lines)))

    if not raw_sections:
        return []

    # ── 第一层：黑名单过滤 ────────────────────────────────────────────────────
    after_blacklist = []
    blacklist_skipped = []
    for title, content in raw_sections:
        if blacklist_hit(title):
            print(f"  [黑名单] 跳过: {title}")
            blacklist_skipped.append(title)
        else:
            after_blacklist.append((title, content))

    if not after_blacklist:
        return []

    # ── 第二层：长期记忆 ──────────────────────────────────────────────────────
    memory_skipped = []
    memory_kept    = []
    after_memory   = []

    if memory:
        patterns = memory.get_section_filter_patterns()
        mem_skip = set(patterns.get("skip", []))
        mem_keep = set(patterns.get("keep", []))

        for title, content in after_blacklist:
            clean = title.strip("*~_ ")
            if clean in mem_skip:
                print(f"  [记忆] 跳过（历史学习）: {title}")
                memory_skipped.append(title)
            elif clean in mem_keep:
                print(f"  [记忆] 保留（历史学习）: {title}")
                memory_kept.append(title)
                after_memory.append((title, content))
            else:
                after_memory.append((title, content))  # 未知，交给模型判断
    else:
        after_memory = after_blacklist

    # ── 第三层：模型判断（只对记忆未覆盖的章节）────────────────────────────────
    model_skipped = []
    final_sections = []

    unknown = [(t, c) for t, c in after_memory if t not in memory_kept]

    if unknown and len(unknown) > 0:
        titles_list = "\n".join(
            f"{i+1}. {t}" for i, (t, _) in enumerate(unknown)
        )
        print(f"  [模型判断] 对 {len(unknown)} 个未知章节进行判断...")
        try:
            import anthropic as _ant
            import os as _os
            _client = _ant.Anthropic()
            resp = _client.messages.create(
                model=_os.environ.get("ANTHROPIC_MODEL") or _os.environ.get("DEFAULT_LLM_MODEL") or _os.environ.get("MODEL_ID", "claude-sonnet-4-6"),
                system="你是一名测试架构师，判断需求文档的章节是否属于核心需求内容。只输出 JSON，不要其他文字。",
                messages=[{"role": "user", "content": (
                    f"以下是需求文档的章节列表，请判断每个章节是否属于「核心需求内容」"
                    f"（即包含功能需求、计算逻辑、数据规则、枚举值定义等可以生成测试点的内容）：\n\n"
                    f"{titles_list}\n\n"
                    "输出 JSON，格式：\n"
                    '{"core": [1, 2, 5], "skip": [3, 4]}\n'
                    "（core 和 skip 各填章节序号列表）"
                )}],
                max_tokens=500,
            )
            result_text = "".join(b.text for b in resp.content if hasattr(b, "text"))
            import json as _json, re as _re
            result_text = _re.sub(r"```.*?```", "", result_text, flags=_re.DOTALL).strip()
            # 找到 JSON 对象起始位置
            json_start = result_text.find("{")
            if json_start >= 0:
                result_text = result_text[json_start:]
            try:
                judgment = _json.loads(result_text)
            except Exception:
                # 用正则兜底提取数字
                core_nums = [int(x) for x in _re.findall(r'"core"\s*:\s*\[([^\]]*)\]', result_text)]
                skip_nums = [int(x) for x in _re.findall(r'"skip"\s*:\s*\[([^\]]*)\]', result_text)]
                all_nums  = [int(x) for x in _re.findall(r'\d+', result_text)]
                judgment  = {"core": list(range(1, len(unknown)+1)), "skip": []}
            core_idxs = {i - 1 for i in judgment.get("core", [])}
            skip_idxs = {i - 1 for i in judgment.get("skip", [])}

            for i, (title, content) in enumerate(unknown):
                if i in skip_idxs:
                    print(f"  [模型] 跳过: {title}")
                    model_skipped.append(title)
                else:
                    final_sections.append((title, content))

        except Exception as e:
            print(f"  [warn] 模型判断失败: {e}，保留所有未知章节")
            final_sections.extend(unknown)
    else:
        final_sections.extend(unknown)

    # 加入记忆已确认保留的章节
    for title, content in after_memory:
        if title in memory_kept and (title, content) not in final_sections:
            final_sections.append((title, content))

    # ── 写回记忆（第三层学习结果）────────────────────────────────────────────
    if memory and model_skipped:
        all_skipped = blacklist_skipped + memory_skipped + model_skipped
        all_kept    = [t for t, _ in final_sections]
        memory.save_section_filter_result(
            skipped=[t for t in all_skipped if t not in blacklist_skipped],
            kept=all_kept
        )

    return final_sections


def main():
    parser = argparse.ArgumentParser(description="测试用例生成 Agent v2")
    parser.add_argument("requirement",    help="需求文档路径（.md / .txt）")
    parser.add_argument("--kb",           action="store_true", help="启用知识库检索")
    parser.add_argument("--skip-review",  action="store_true", help="跳过需求评审")
    parser.add_argument("--no-cases",     action="store_true", help="只生成测试点，不展开用例")
    parser.add_argument("--resume",       action="store_true", help="续跑：自动找最近未完成的任务")
    parser.add_argument("--section",      type=str, default="", help="只针对指定章节生成，如 --section 结算风控金")
    args = parser.parse_args()

    req_path = Path(args.requirement).resolve()
    if not req_path.exists():
        print(f"错误: 找不到需求文档 {req_path}")
        sys.exit(1)

    # 自动处理 docx：转换为 md 并放入 knowledge_base/
    if req_path.suffix.lower() in (".docx", ".doc"):
        md_path = KB_DIR / (req_path.stem + ".md")
        if not md_path.exists():
            print(f"  检测到 .docx，自动转换为 Markdown...")
            import subprocess as _sp
            # 优先用 docx2md.py（表格格式更干净）
            docx2md_script = WORKDIR / "docx2md.py"
            if docx2md_script.exists():
                r = _sp.run(
                    [sys.executable, str(docx2md_script), str(req_path), "-o", str(md_path)],
                    capture_output=True, text=True, timeout=120
                )
                if r.returncode != 0 or not md_path.exists():
                    # 降级 pandoc
                    r = _sp.run(
                        ["pandoc", str(req_path), "-t", "markdown", "-o", str(md_path)],
                        capture_output=True, text=True
                    )
            else:
                r = _sp.run(
                    ["pandoc", str(req_path), "-t", "markdown", "-o", str(md_path)],
                    capture_output=True, text=True
                )
            if r.returncode != 0 or not md_path.exists():
                print(f"  [错误] 转换失败: {r.stderr[:200]}")
                sys.exit(1)
            print(f"  转换完成: {md_path.name}")
        else:
            print(f"  使用已有转换版本: {md_path.name}")
        req_path = md_path
    elif not req_path.is_relative_to(WORKDIR):
        # 不在 WORKDIR 内的文件复制到 knowledge_base/
        dst = KB_DIR / req_path.name
        if not dst.exists():
            import shutil as _shutil
            _shutil.copy2(req_path, dst)
            print(f"  文件已复制到 knowledge_base/{req_path.name}")
        req_path = dst

    ts       = int(time.time())
    stem     = req_path.stem
    req_name = req_path.name

    # 自动清理旧任务文件（只保留最近 30 个）
    tasks_dir = WORKDIR / ".tasks"
    if tasks_dir.exists():
        task_files = sorted(tasks_dir.glob("*.json"), key=lambda f: f.stat().st_mtime)
        to_delete  = task_files[:-30] if len(task_files) > 30 else []
        for f in to_delete:
            try: f.unlink()
            except Exception: pass
        if to_delete:
            print(f"  [清理] 删除 {len(to_delete)} 个旧任务文件")

    # 章节过滤：如果指定了 --section，提取对应章节内容写入临时文件
    section_keyword = args.section.strip()
    if section_keyword:
        print(f"  [章节过滤] 只处理包含「{section_keyword}」的章节")
        section_text = _extract_section(req_path, section_keyword)
        if section_text:
            # 写入临时文件，后续所有阶段读这个文件
            section_path = WORKDIR / f"_section_{stem}_{ts}.md"
            section_path.write_text(section_text, encoding="utf-8")
            print(f"  [章节过滤] 提取到 {len(section_text.splitlines())} 行内容，继续生成")
            req_path = section_path
            stem     = f"{stem}_{section_keyword}"
            req_name = section_path.name
        else:
            print(f"  [章节过滤] 未找到包含「{section_keyword}」的章节，使用完整文档")

    RUN_DIR  = get_run_dir(stem, ts)   # 本次运行的输出目录

    print(f"\n{'='*52}")
    print(f"  测试 Agent v2 启动")
    print(f"  需求文档: {req_name}")
    print(f"  知识库:   {'启用' if args.kb else '关闭'}")
    print(f"  生成用例: {'否' if args.no_cases else '是'}")
    print(f"{'='*52}\n")

    # s07: 初始化任务存储，支持续跑
    if args.resume:
        task = TaskStore.find_latest(stem)
        if task:
            print(f"  [s07] 续跑模式: {task.summary()}")
        else:
            print(f"  [s07] 未找到可续跑的任务，重新开始")
            task = TaskStore(stem, ts)
    else:
        task = TaskStore(stem, ts)
    print(f"  [s07] 任务文件: {task.path.name}")

    # s09: 初始化记忆系统
    memory     = MemoryStore(stem)
    memory_rag = MemoryRAG()   # 向量化长期记忆检索
    lt_counts  = {k: len(v) for k, v in memory._lt.items() if isinstance(v, list)}
    print(f"  [s09] 长期记忆: {lt_counts}（向量检索已就绪）\n")

    # ① 需求评审
    if task.is_done("review"):
        review = task.get_result("review")
        print(f"  [s07] 跳过需求评审（已完成）")
    elif args.skip_review:
        review = {"testable_features": [], "risk_flags": [], "score": 0}
        task.done("review", review)
        print("  [跳过] 需求评审")
    else:
        task.start("review")
        try:
            review = stage1_review(req_path, memory=memory)
            task.done("review", review)
            memory.save_after_review(review)   # s09: 保存评审经验
            memory_rag.invalidate()               # 触发记忆索引重建
        except Exception as e:
            task.fail("review", str(e))
            review = {"testable_features": [], "risk_flags": [], "score": 0}
            print(f"  [s11] 评审异常，使用空结果继续: {e}")
        score    = review.get("score", "N/A")
        features = review.get("testable_features", [])
        risks    = review.get("risk_flags", [])
        print(f"\n  评审完成 → 质量分: {score}, 功能点: {len(features)}, 风险项: {len(risks)}")
        if isinstance(score, int) and score < 60:
            print(f"  [警告] 需求质量较低（{score}/100），建议完善后再生成测试点")

    # ② 测试点生成
    if task.is_done("testpoints"):
        testpoints = task.get_result("testpoints")
        print(f"  [s07] 跳过测试点生成（已完成，共 {len(testpoints)} 条）")
    else:
        task.start("testpoints")
        try:
            # 无 --section 时，自动识别章节逐章处理后合并
            if not section_keyword:
                sections = _split_sections(req_path, memory=memory)
            else:
                sections = []

            if len(sections) >= 2:
                print(f"  [自动章节] 识别到 {len(sections)} 个章节，逐章处理后合并")
                for i, (title, _) in enumerate(sections):
                    print(f"    {i+1}. {title}")

                all_tps = []
                for i, (title, content) in enumerate(sections):
                    s_offset = len(all_tps)
                    print(f"\n  [章节 {i+1}/{len(sections)}] {title}")
                    s_path = OUTPUT_DIR / f"_sec_{int(time.time()*1000)%1000000}.md"
                    # 检查章节内容是否有实质内容（至少有10行非空行且含中文）
                    content_lines = [l for l in content.splitlines() if l.strip()]
                    cn_chars = sum(1 for c in content if '\u4e00' <= c <= '\u9fff')
                    print(f"    内容: {len(content_lines)}行, {cn_chars}个中文字符", flush=True)
                    # 跳过条件：非空行少于8行，或中文字符少于100个，或没有具体数字/字段名
                    has_specific_content = any(
                        any(kw in line for kw in ["字段", "表", "取值", "逻辑", "计算", "=", "：", "规则"])
                        for line in content_lines
                    )
                    if len(content_lines) < 8 or cn_chars < 100 or not has_specific_content:
                        print(f"    [跳过] 章节内容不足（行数{len(content_lines)}, 中文{cn_chars}, 有具体内容:{has_specific_content}）")
                        continue
                    s_path.write_text(f"# {title}\n\n{content}", encoding="utf-8")
                    try:
                        tps = stage2_testpoints(s_path, review, args.kb, memory=memory)
                        for j, tp in enumerate(tps):
                            tp["testpoint_id"] = f"TP-{s_offset + j + 1:03d}"
                            tp.setdefault("section", title)
                        all_tps.extend(tps)
                        print(f"    → {len(tps)} 条测试点")
                    finally:
                        s_path.unlink(missing_ok=True)

                testpoints = all_tps
                print(f"\n  [合并] 共 {len(testpoints)} 条测试点，来自 {len(sections)} 个章节")
            else:
                testpoints = stage2_testpoints(req_path, review, args.kb, memory=memory)

            task.done("testpoints", testpoints)
            memory.save_after_testpoints(testpoints, review)  # s09: 保存测试点经验
            memory_rag.invalidate()                               # 触发记忆索引重建
        except Exception as e:
            task.fail("testpoints", str(e))
            testpoints = []
            print(f"  [s11] 测试点生成异常: {e}")

    # 标准化为扁平列表
    flat_tps = []
    if testpoints and isinstance(testpoints[0], dict):
        if "testpoints" in testpoints[0]:
            for module in testpoints:
                flat_tps.extend(module.get("testpoints", []))
        else:
            flat_tps = testpoints

    req_count  = sum(1 for t in flat_tps if get_source(t) == "REQ")
    kb_count   = sum(1 for t in flat_tps if get_source(t) == "KB")
    risk_count = sum(1 for t in flat_tps if get_source(t) == "RISK")

    # 用模型名+时间戳生成文件后缀，确保同模型多次运行也不重名
    _model_tag = re.sub(r"[^a-zA-Z0-9._-]", "-", MODEL.split("/")[-1])[:16].strip("-")
    _sfx = f"_{_model_tag}_{RUN_DIR.name}"

    # 保存测试点 JSON
    tp_out = RUN_DIR / f"testpoints{_sfx}.json"
    tp_out.write_text(json.dumps({
        "meta": {
            "requirement": str(req_path),
            "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "total": len(flat_tps),
            "by_source": {"REQ": req_count, "KB": kb_count, "RISK": risk_count},
        },
        "review": review,
        "testpoints": flat_tps,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    # 生成 Markdown（供导入 XMind）
    md_out   = RUN_DIR / f"testpoints_xmind{_sfx}.md"
    xmind_ok = export_markdown_xmind(flat_tps, review, req_name, md_out)

    print(f"\n{'='*52}")
    print(f"  ② 测试点生成完成")
    print(f"     总数: {len(flat_tps)}  🔵REQ={req_count}  🟡KB={kb_count}  🔴RISK={risk_count}")
    print(f"  输出目录: {RUN_DIR.relative_to(WORKDIR)}")
    print(f"     JSON:  {tp_out.name}")
    if xmind_ok:
        print(f"     Markdown(→XMind): {md_out.name}")
    print(f"{'='*52}")

    if args.no_cases:
        return

    # ③ 测试用例生成（测试点为空则跳过）
    if not flat_tps:
        print("\n  [跳过] 测试点为空，跳过用例生成")
        return

    if task.is_done("testcases"):
        testcases = task.get_result("testcases")
        print(f"  [s07] 跳过用例生成（已完成，共 {len(testcases)} 条）")
    else:
        task.start("testcases")
        try:
            testcases = stage3_testcases(flat_tps, req_path)
            task.done("testcases", testcases)
        except Exception as e:
            task.fail("testcases", str(e))
            testcases = []
            print(f"  [s11] 用例生成异常: {e}")

    tc_out   = RUN_DIR / f"testcases{_sfx}.json"
    xlsx_out = RUN_DIR / f"testcases{_sfx}.xlsx"

    tc_out.write_text(json.dumps(testcases, ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        xlsx_ok = export_excel(testcases, xlsx_out)
    except Exception as e:
        print(f"  [warn] Excel 生成失败: {e}")
        import traceback; traceback.print_exc()
        xlsx_ok = False

    # ④ 测分文档生成（本地，零 token）
    from gen_report import generate_report
    report_out = RUN_DIR / f"report{_sfx}.md"
    try:
        report_md = generate_report(
            {"meta": {"requirement": str(req_path), "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                      "total": len(flat_tps),
                      "by_source": {"REQ": req_count, "KB": kb_count, "RISK": risk_count}},
             "review": review, "testpoints": flat_tps},
            testcases,
            report_out
        )
        report_out.write_text(report_md, encoding="utf-8")
        report_ok = True
    except Exception as e:
        print(f"  [warn] 测分文档生成失败: {e}")
        report_ok = False

    task.done("export", {"testcases": str(tc_out.name), "excel": str(xlsx_out.name),
                         "report": str(report_out.name) if report_ok else ""})

    print(f"\n{'='*52}")
    print(f"  ③ 测试用例生成完成")
    print(f"     总数: {len(testcases)} 条")
    print(f"     JSON:  {tc_out.name}")
    if xlsx_ok:
        print(f"     Excel: {xlsx_out.name}")
    if report_ok:
        print(f"     测分:  {report_out.name}")
    print(f"\n  [s07] 最终任务状态: {task.summary()}")
    print(f"{'='*52}\n")


if __name__ == "__main__":
    main()
